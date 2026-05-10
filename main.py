import os
import json
import re
import asyncio
import shutil
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from urllib.parse import urlparse
from pathlib import Path

import httpx
try:
    import openpyxl
except ImportError:
    openpyxl = None
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel
import anthropic

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory="static"), name="static")
client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

TEMPLATE_PATH = "static/tiktok_template.xlsx"
EXPORTS_DIR = Path("exports")
EXPORTS_DIR.mkdir(exist_ok=True)
DATA_DIR = Path(os.environ.get("DATA_DIR", "/data"))
DATA_DIR.mkdir(exist_ok=True, parents=True)
TT_HISTORY_FILE = DATA_DIR / "tiktok_history.json"
META_HISTORY_FILE = DATA_DIR / "meta_history.json"
KREATIVE_HISTORY_FILE = DATA_DIR / "kreative_history.json"
FORECAST_ENTRIES_FILE = DATA_DIR / "forecast_entries.json"
FORECAST_DELETED_FILE = DATA_DIR / "forecast_deleted.json"
FORECAST_HISTORY_FILE = DATA_DIR / "forecast_history.json"
SPOROCANJE_FILE = DATA_DIR / "sporocanje_common.json"

# ─── ECONT GEO (cities / streets / quarters lookup) ──────────────────────────
def _load_econt_geo():
    for p in [DATA_DIR / "econt_geo.json", Path("econt_geo.json"), Path("static/econt_geo.json")]:
        if p.exists():
            try:
                with open(p, encoding="utf-8") as f:
                    data = json.load(f)
                print(f"[econt_geo] Loaded from {p} — {len(data.get('city_by_id',{}))} cities, {len(data.get('streets_by_city',{}))} cities w/ streets")
                return data
            except Exception as e:
                print(f"[econt_geo] Failed to load {p}: {e}")
    print("[econt_geo] NOT FOUND — street validation disabled")
    return {}

ECONT_GEO = _load_econt_geo()


def econt_lookup_city(zip_code: str, city_name: str) -> dict | None:
    """Vrni city entry (id, name_bg, name_en, zip) glede na ZIP ali ime mesta."""
    if not ECONT_GEO:
        return None
    zip_map = ECONT_GEO.get("zip_to_city_id", {})
    city_map = ECONT_GEO.get("city_by_id", {})
    name_map = ECONT_GEO.get("name_to_city_ids", {})
    city_key = city_name.lower().strip()

    # 1. Poskusi ZIP + ime skupaj — poišči city ki ima ta ZIP IN ime se ujema
    if zip_code and city_name:
        # Najdi vse city ki imajo ta ZIP
        candidates = []
        for cid, cdata in city_map.items():
            if cdata.get("zip") == str(zip_code).strip():
                candidates.append((cid, cdata))
        if candidates:
            # Med kandidati poišči najboljše ujemanje po imenu
            for cid, cdata in candidates:
                en = (cdata.get("name_en") or "").lower()
                bg = (cdata.get("name_bg") or "").lower()
                if city_key == en or city_key == bg:
                    return {"id": int(cid), **cdata}
            # Delno ujemanje
            for cid, cdata in candidates:
                en = (cdata.get("name_en") or "").lower()
                bg = (cdata.get("name_bg") or "").lower()
                if city_key in en or en in city_key or city_key in bg or bg in city_key:
                    return {"id": int(cid), **cdata}
            # ZIP ustreza ampak ime ne — vrni prvega (fallback)
            cid, cdata = candidates[0]
            return {"id": int(cid), **cdata}

    # 2. Samo ZIP lookup
    if zip_code:
        cid = zip_map.get(str(zip_code).strip())
        if cid:
            return {"id": cid, **city_map.get(str(cid), {})}

    # 3. Samo ime lookup (lowercase, točno)
    cids = name_map.get(city_key)
    if cids:
        cid = cids[0]
        return {"id": cid, **city_map.get(str(cid), {})}

    # 4. Delno ujemanje po imenu
    for k, ids in name_map.items():
        if city_key and (city_key in k or k in city_key):
            cid = ids[0]
            return {"id": cid, **city_map.get(str(cid), {})}
    return None


def econt_get_streets_context(city_id, street_query: str, max_results: int = 12) -> str:
    """Vrni seznam ulic iz mesta kot kontekst za AI — fuzzy match na street_query."""
    if not ECONT_GEO or not city_id:
        return ""
    streets = ECONT_GEO.get("streets_by_city", {}).get(str(city_id), [])
    quarters = ECONT_GEO.get("quarters_by_city", {}).get(str(city_id), [])
    if not streets and not quarters:
        return ""
    # Fuzzy match — poišči ulice ki vsebujejo ključne besede iz poizvedbe
    q = re.sub(r'(ul\.|bul\.|zh\.k\.|kv\.|bl\.|vh\.|et\.|ap\.|\d+)', '', street_query.lower()).strip()
    words = [w for w in q.split() if len(w) > 2]
    scored = []
    for s in streets:
        en = s[1].lower() if s[1] else ""
        bg = s[0].lower() if s[0] else ""
        score = sum(1 for w in words if w in en or w in bg)
        if score > 0:
            scored.append((score, s))
    scored.sort(key=lambda x: -x[0])
    top_streets = [s for _, s in scored[:max_results]]
    # Če ni zadetkov — vrni prvih N ulic (za osnovno validacijo)
    if not top_streets:
        top_streets = streets[:8]
    lines = []
    if quarters:
        q_sample = quarters[:6]
        lines.append("Četrti: " + ", ".join(f"{q[1]}" for q in q_sample if q[1]))
    lines.append("Ulice (vzorec):")
    for s in top_streets:
        lines.append(f"  {s[1]} / {s[0]}")
    return "\n".join(lines)

# ─── BRAND DOMAIN MAPS ───────────────────────────────────────────────────────

BRAND_DOMAINS = {
    "maaarket": {"sl":"www.maaarket.si","hr":"www.maaarket.hr","rs":"www.maaarket.rs","hu":"www.maaarket.hu","cz":"www.maaarket.cz","sk":"www.maaarket.sk","pl":"www.maaarket.pl","gr":"www.maaarket.gr","ro":"www.maaarket.ro","bg":"www.maaarket.bg"},
    "fluxigo":  {"sl":"www.fluxigo.si","hr":"www.fluxigo.hr","rs":"www.fluxigo.rs","hu":"www.fluxigo.hu","cz":"www.fluxigo.cz","sk":"www.fluxigo.sk","pl":"www.fluxigo.pl","gr":"www.fluxigo.gr","ro":"www.fluxigo.ro","bg":"www.fluxigo.bg"},
    "easyzo":   {"sl":"www.easyzo.si","hr":"www.easyzo.hr","rs":"www.easyzo.rs","hu":"www.easyzo.hu","cz":"www.easyzo.cz","sk":"www.easyzo.sk","pl":"www.easyzo.pl","gr":"www.easyzo.gr","ro":"www.easyzo.ro","bg":"www.easyzo.bg"},
    "zipply":   {"sl":"www.zipply.si","hr":"www.zipply.hr","rs":"www.zipply.rs","hu":"www.zipply.hu","cz":"www.zipply.cz","sk":"www.zipply.sk","pl":"www.zipply.pl","gr":"www.zipply.gr","ro":"www.zipply.ro","bg":"www.zipply.bg"},
    "thundershop": {"sl":"www.thundershop.si","hr":"www.thundershop.hr","rs":"www.thundershop.rs","hu":"www.thundershop.hu","cz":"www.thundershop.cz","sk":"www.thundershop.sk","gr":"www.thundershop.gr","ro":"www.thundershop.ro","bg":"www.thundershop.bg"},
    "colibrishop": {"sl":"www.colibrishop.si","hr":"www.colibrishop.hr","rs":"www.colibrishop.rs","cz":"www.colibrishop.cz","sk":"www.colibrishop.sk","gr":"www.colibrishop.gr","ro":"www.colibrishop.ro","bg":"www.colibrishop.bg"},
}

MAAARKET_FEEDS = {
    "sl":"https://api.maaarket.si/storage/exports/sl/google.xml",
    "hr":"https://api.maaarket.hr/storage/exports/hr/google.xml",
    "rs":"https://api.maaarket.rs/storage/exports/sr/google.xml",
    "hu":"https://api.maaarket.hu/storage/exports/hu/google.xml",
    "pl":"https://api.maaarket.pl/storage/exports/pl/google.xml",
    "cz":"https://api.maaarket.cz/storage/exports/cs/google.xml",
    "sk":"https://api.maaarket.sk/storage/exports/sk/google.xml",
    "gr":"https://api.maaarket.gr/storage/exports/el/google.xml",
    "bg":"https://api.maaarket.bg/storage/exports/bg/google.xml",
    "ro":"https://api.maaarket.ro/storage/exports/ro/google.xml",
}

G = "http://base.google.com/ns/1.0"
feed_by_lang: dict = {}
slug_to_id: dict = {}
last_fetch: Optional[datetime] = None
CACHE_TTL_HOURS = 24


def is_cache_stale():
    return last_fetch is None or datetime.now() - last_fetch > timedelta(hours=CACHE_TTL_HOURS)


def extract_slug(url: str) -> Optional[str]:
    path = urlparse(url).path.rstrip('/')
    parts = [p for p in path.split('/') if p]
    return parts[-1].lower() if parts else None


def parse_feed(xml_content: str) -> dict:
    products = {}
    try:
        root = ET.fromstring(xml_content)
        channel = root.find('channel')
        if channel is None:
            return products
        for item in channel.findall('item'):
            gid_el = item.find(f'{{{G}}}id')
            link_el = item.find(f'{{{G}}}link')
            if gid_el is None or not gid_el.text or link_el is None or not link_el.text:
                continue
            g_id = gid_el.text.strip()
            url = link_el.text.strip()
            path = urlparse(url).path
            products[g_id] = {"url": url, "path": path}
    except ET.ParseError:
        pass
    return products


async def fetch_all_feeds():
    global feed_by_lang, slug_to_id, last_fetch
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching XML feeds...")
    async with httpx.AsyncClient(timeout=30.0) as hc:
        tasks = {lang: hc.get(url) for lang, url in MAAARKET_FEEDS.items()}
        new_cache = {}
        for lang, task in tasks.items():
            try:
                resp = await task
                new_cache[lang] = parse_feed(resp.text) if resp.status_code == 200 else {}
                print(f"  ✓ {lang}: {len(new_cache.get(lang,{}))} products")
            except Exception as e:
                new_cache[lang] = {}
                print(f"  ✗ {lang}: {e}")
    feed_by_lang = new_cache
    new_slug_to_id = {}
    for lang, lang_feed in feed_by_lang.items():
        for g_id, data in lang_feed.items():
            slug = extract_slug(data["url"])
            if slug and slug not in new_slug_to_id:
                new_slug_to_id[slug] = g_id
    slug_to_id = new_slug_to_id
    last_fetch = datetime.now()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Done. Slug index: {len(slug_to_id)}")


async def ensure_cache_fresh():
    if is_cache_stale():
        await fetch_all_feeds()


def detect_brand(url: str) -> Optional[str]:
    if not url:
        return None
    domain = urlparse(url).netloc.lower().replace("www.", "")
    for brand, lang_map in BRAND_DOMAINS.items():
        for d in lang_map.values():
            if domain == d.replace("www.", ""):
                return brand
    return None


def find_product_urls(source_url: Optional[str]) -> dict:
    if not source_url:
        return {}
    brand = detect_brand(source_url) or "maaarket"
    slug = extract_slug(source_url)
    if not slug:
        return {}
    g_id = slug_to_id.get(slug)
    if not g_id:
        return {}
    target_domains = BRAND_DOMAINS.get(brand, BRAND_DOMAINS["maaarket"])
    result = {}
    for lang, products in feed_by_lang.items():
        if lang not in target_domains or g_id not in products:
            continue
        result[lang] = f"https://{target_domains[lang]}{products[g_id]['path']}"
    return result


# ─── STARTUP ─────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup_event():
    await fetch_all_feeds()
    asyncio.create_task(daily_refresh())


async def daily_refresh():
    while True:
        await asyncio.sleep(CACHE_TTL_HOURS * 3600)
        await fetch_all_feeds()


# ─── PROMPT BUILDERS ─────────────────────────────────────────────────────────

def build_meta_prompt(user_msg: str, pt_count: int, hl_count: int) -> str:
    pt_ph = ", ".join([f'"PT {i+1}"' for i in range(pt_count)])
    hl_ph = ", ".join([f'"HL {i+1}"' for i in range(hl_count)])
    return f"""{user_msg}

Ustvari Meta oglase za FB/Instagram v 10 jezikih.

Primary Text ({pt_count}x na jezik): 2-3 vrstice, 2-3 emoji-ji, prodajni ton, brez cen, vsak DRUGAČEN.
Headline ({hl_count}x na jezik): MAX 5 BESED, 1 emoji na začetku, brez cen, vsak DRUGAČEN.
EMOJI PRAVILO: Uporabljaj SAMO te emoji-je ki so zagotovo podprti na vseh napravah:
✅ ⭐ 🔥 💪 🎯 👍 ❤️ 💥 🚀 ✨ 💡 🎁 💰 👌 🙌 😍 💎 🏆 ⚡ 🌟 👏 💫 🛒 📦 🔑 💯 😊 🤩 🌿 🌱 🍃 🌸 🌻 🌞 🍀 🎉 🎊 🛍️ 💚 💙 🧡 💜 🤍 🖤
NE uporabljaj: redkih, novejših ali manj znanih emoji-jev ki se lahko prikažejo kot □

Jeziki: SL (izvirnik), HR (latinica), RS (SAMO latinica!), HU, CZ, SK, PL, GR (grška pisava), RO (latinica), BG (SAMO cirilica!).

Vrni SAMO JSON brez markdown:
{{
  "product": "ime",
  "sl": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "hr": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "rs": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "hu": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "cz": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "sk": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "pl": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "gr": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "ro": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "bg": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}}
}}"""


def build_tiktok_prompt(user_msg: str) -> str:
    return f"""{user_msg}

Ustvari TikTok oglasne tekste v 10 jezikih.

Pravila:
- Točno 4 variante na jezik
- Vsaka varianta MAX 80 znakov (strogo!)
- Brez emojiev, brez # in {{}}
- Kratek, direkten, akcijski stil
- Vsaka varianta drugačen pristop (korist, socialni dokaz, nujnost, radovednost)
- Brez "kakovost/dostava/zaloga" strukture — bodi kreativen

KRITIČNO PREPOVEDANO:
- NE omenjaj cen ali EUR (kot "6,99 EUR", "samo X EUR", itd.)
- NE omenjaj popustov ali procentov (kot "minus 62%", "do -50%", itd.)
- NE izmišljaj specifičnih številk (kot "100 funkcij", "375 kupcev") — uporabljaj generične izraze ("tisoči kupcev", "številne funkcije")
- NE omenjaj "danes/jutri" v povezavi z akcijo

Jeziki: SL (izvirnik), HR (latinica), RS (SAMO latinica!), HU, CZ, SK, PL, GR (grška pisava), RO (latinica), BG (SAMO cirilica!).

KRITIČNO VAŽNO — FORMAT:
- Vsak jezik mora imeti TOČNO 4 oklepaje: [var1],[var2],[var3],[var4]
- Vejica ZNOTRAJ variante je PREPOVEDANA — če bi napisal "Brez kuhinje, brez stresa" → napiši "Brez kuhinje in brez stresa"
- Vejica se sme pojaviti SAMO med oklepaji kot separator: ],[
- Vrni IZKLJUČNO in SAMO JSON — nobenih uvodnih besed, nobenih razlag, nobenih komentarjev, nobenih markdown backticks. Prva in zadnja stvar v odgovoru mora biti {{ in }}. Nič drugega.
{{
  "product": "ime",
  "sl": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "hr": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "rs": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "hu": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "cz": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "sk": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "pl": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "gr": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "ro": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "bg": "[tekst1],[tekst2],[tekst3],[tekst4]"
}}\n"""


# ─── GENERATE HELPERS ────────────────────────────────────────────────────────

def parse_json_response(text: str) -> Optional[dict]:
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text).strip()
    # Try direct parse first
    match = re.search(r'\{[\s\S]*\}', text)
    if not match:
        return None
    json_str = match.group()
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        # Try to fix truncated JSON by finding last complete key-value
        try:
            # Remove trailing incomplete content after last complete string value
            fixed = re.sub(r',\s*"[^"]*":\s*"[^"]*$', '', json_str)
            fixed = re.sub(r',\s*"[^"]*":\s*$', '', fixed)
            if not fixed.endswith('}'):
                fixed = fixed.rstrip(',\n\r\t ') + '}'
            return json.loads(fixed)
        except Exception:
            return None


async def call_claude(prompt: str, model: str, tools=None, max_tokens: int = 4000) -> str:
    loop = asyncio.get_event_loop()
    kwargs = {"model": model, "max_tokens": max_tokens, "messages": [{"role": "user", "content": prompt}]}
    if tools:
        kwargs["tools"] = tools

    for attempt in range(3):
        try:
            msg = await loop.run_in_executor(None, lambda: client.messages.create(**kwargs))
            return "".join(b.text for b in msg.content if hasattr(b, "text"))
        except anthropic.RateLimitError:
            if attempt < 2:
                wait = (attempt + 1) * 20
                print(f"  Rate limit, waiting {wait}s...")
                await asyncio.sleep(wait)
            else:
                raise
        except Exception as e:
            raise
    return ""


async def generate_meta_sl_only(user_msg: str, mode: str, source_url: Optional[str],
                                pt_count: int, hl_count: int) -> dict:
    """Generira samo SL tekste brez prevajanja — za streaming mode."""
    product_urls = find_product_urls(source_url)
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []
    pt_ph = ", ".join([f'"PT {i+1}"' for i in range(pt_count)])
    hl_ph = ", ".join([f'"HL {i+1}"' for i in range(hl_count)])
    sl_prompt = f"""{user_msg}

Ustvari Meta oglase SAMO v slovenščini.
Primary Text ({pt_count}x): 2-3 vrstice, 2-3 emoji-ji, prodajni ton, brez cen, vsak DRUGAČEN.
Headline ({hl_count}x): MAX 5 BESED, 1 emoji na začetku, brez cen, vsak DRUGAČEN.
EMOJI PRAVILO: Uporabljaj SAMO te emoji-je ki so zagotovo podprti na vseh napravah:
✅ ⭐ 🔥 💪 🎯 👍 ❤️ 💥 🚀 ✨ 💡 🎁 💰 👌 🙌 😍 💎 🏆 ⚡ 🌟 👏 💫 🛒 📦 🔑 💯 😊 🤩 🌿 🌱 🍃 🌸 🌻 🌞 🍀 🎉 🎊 🛍️ 💚 💙 🧡 💜 🤍 🖤
NE uporabljaj: redkih, novejših ali manj znanih emoji-jev ki se lahko prikažejo kot □
Vrni SAMO JSON: {{"product": "ime", "pt": [{pt_ph}], "hl": [{hl_ph}]}}"""

    sl_text = await call_claude(sl_prompt, "claude-sonnet-4-6", tools if tools else None, 1500)
    sl_data = parse_json_response(sl_text)
    if not sl_data:
        return {"error": "Napaka pri generiranju SL tekstov."}
    return {
        "product": sl_data.get("product", "Izdelek"),
        "sl": {"pt": sl_data.get("pt", []), "hl": sl_data.get("hl", [])},
        "product_urls": product_urls,
    }


async def generate_meta_one(user_msg: str, mode: str, source_url: Optional[str],
                            pt_count: int, hl_count: int, qmode: str) -> dict:
    product_urls = find_product_urls(source_url)
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []

    if qmode == "fast":
        pt_ph = ", ".join([f'"PT {i+1}"' for i in range(pt_count)])
        hl_ph = ", ".join([f'"HL {i+1}"' for i in range(hl_count)])
        sl_prompt = f"""{user_msg}

Ustvari Meta oglase SAMO v slovenščini.
Primary Text ({pt_count}x): 2-3 vrstice, 2-3 emoji-ji, prodajni ton, brez cen, vsak DRUGAČEN.
Headline ({hl_count}x): MAX 5 BESED, 1 emoji na začetku, brez cen, vsak DRUGAČEN.
EMOJI PRAVILO: Uporabljaj SAMO te emoji-je ki so zagotovo podprti na vseh napravah:
✅ ⭐ 🔥 💪 🎯 👍 ❤️ 💥 🚀 ✨ 💡 🎁 💰 👌 🙌 😍 💎 🏆 ⚡ 🌟 👏 💫 🛒 📦 🔑 💯 😊 🤩 🌿 🌱 🍃 🌸 🌻 🌞 🍀 🎉 🎊 🛍️ 💚 💙 🧡 💜 🤍 🖤 🟢 🔵 🟡
NE uporabljaj: redkih, novejših ali manj znanih emoji-jev ki se lahko prikažejo kot □
Vrni SAMO JSON: {{"product": "ime", "pt": [{pt_ph}], "hl": [{hl_ph}]}}"""

        sl_text = await call_claude(sl_prompt, "claude-sonnet-4-6", tools if tools else None, 1500)
        sl_data = parse_json_response(sl_text)
        if not sl_data:
            print(f"SL parse failed. Raw response: {sl_text[:500]}")
            return {"error": "Napaka pri generiranju SL tekstov."}

        sl_pts = sl_data.get("pt", [])
        sl_hls = sl_data.get("hl", [])
        trans_prompt = f"""Prevedi Meta oglase iz slovenščine v 9 jezikov. Ohrani ŠTEVILO in POZICIJO emoji-jev točno kot v originalu.

Primary Texts: {json.dumps(sl_pts, ensure_ascii=False)}
Headlines: {json.dumps(sl_hls, ensure_ascii=False)}

PRAVILA PREVAJANJA PO JEZIKIH:

HR (hrvaščina, latinica): Natural marketing ton. Pazi na "č/ć/š/ž/đ".

RS (srbščina, SAMO LATINICA — NIKOLI cirilica!): Natural marketing ton. Pazi na "č/ć/š/ž/đ".

HU (madžarščina): KRITIČNO - to ni indoevropski jezik, NE prevajaj dobesedno. Uporabljaj aglutinacijo (končnice) pravilno. CTA kot "Naroči zdaj" = "Rendeld meg most". Pazi da stavek ni predolg (madžarski stavki so lahko za 20-30% daljši).

CZ (češčina): Natural. Pazi na sklonjenje samostalnikov (akuzativ po glagolih).

SK (slovaščina): Podobno češčini. Natural.

PL (poljščina): Pazi na sklone (7 padežev). CTA "Naroči" = "Zamów". Uporabi neformalni ti-vi.

GR (grščina, grška pisava!): KRITIČNO - kompleksna slovnica s skloni (nominativ/akuzativ). CTA "Naroči zdaj" = "Παράγγειλε τώρα". Izogibaj se predolgih stavkov. Pazi na spol samostalnikov. Uporabljaj natural marketing grščino, ne dobesedni prevod.

RO (romunščina, latinica): Natural. Pazi na "ă/â/î/ș/ț". CTA "Naroči" = "Comandă".

BG (bolgarščina, SAMO CIRILICA — NIKOLI latinica!): Natural. CTA "Naroči" = "Поръчай".

SPLOŠNA PRAVILA:
- Ohrani prodajni/energičen ton
- Prevodi morajo zveneti kot da jih je pisal materni govorec, ne robot
- Ohrani ŠTEVILO emoji-jev (če ima SL 3 emoji, mora imeti tudi prevod 3)
- Headlines: ohrani MAX 5 besed tudi v prevodu
- Ne prevajaj blagovnih znamk, imen izdelkov, če so v originalu

Vrni SAMO JSON:
{{"hr":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"rs":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"hu":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"cz":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"sk":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"pl":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"gr":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"ro":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"bg":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}}}}"""

        trans_text = await call_claude(trans_prompt, "claude-haiku-4-5-20251001", None, 5000)
        trans_data = parse_json_response(trans_text)
        if not trans_data:
            return {"error": "Napaka pri prevajanju."}

        result = {"product": sl_data.get("product","Izdelek"), "sl": {"pt": sl_pts, "hl": sl_hls}, "product_urls": product_urls}
        result.update(trans_data)
        return result
    else:
        prompt = build_meta_prompt(user_msg, pt_count, hl_count)
        text = await call_claude(prompt, "claude-sonnet-4-6", tools if tools else None)
        data = parse_json_response(text)
        if not data:
            return {"error": "Claude ni vrnil veljavnega JSON."}
        data["product_urls"] = product_urls
        return data


async def generate_tiktok_one(user_msg: str, mode: str, source_url: Optional[str]) -> dict:
    product_urls = find_product_urls(source_url)
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []
    prompt = build_tiktok_prompt(user_msg)
    text = await call_claude(prompt, "claude-sonnet-4-6", tools if tools else None)
    data = parse_json_response(text)
    if not data:
        print(f"TikTok JSON parse failed. Raw (first 800 chars): {text[:800]}")
        return {"error": "Claude ni vrnil veljavnega JSON."}
    data["product_urls"] = product_urls
    return data


# ─── TIKTOK XLSX BUILDER ─────────────────────────────────────────────────────

COUNTRY_TO_LANG = {
    "BG": "bg", "CZ": "cz", "GR": "gr", "SK": "sk", "RS": "rs",
    "RO": "ro", "HU": "hu", "HR": "hr", "PL": "pl", "SLO": "sl"
}


def normalize_tiktok_text(raw: str) -> str:
    """Zagotovi format [var1],[var2],[var3],[var4] ne glede na vhodni format."""
    raw = raw.strip()
    # Pravilni AI format: [var1],[var2],... — pusti kot je
    if re.match(r'^\[.*\],\[.*\]', raw):
        return raw
    # Format ],[  brez prefiksa/sufiksa oklepaja
    if '],[' in raw:
        parts = re.split(r'\]\s*,\s*\[', raw)
        parts = [p.strip().strip('[]') for p in parts if p.strip().strip('[]')]
        return ','.join(f'[{p}]' for p in parts)
    # Brez oklepajev — AI pozabil dodati []. Splittaj po vejici.
    # Varno ker pravilo narekuje max 80 znakov / varianto in vsebinska vejica ni pričakovana.
    if '[' not in raw:
        parts = [p.strip() for p in raw.split(',') if p.strip()]
        return ','.join(f'[{p}]' for p in parts)
    # En sam oklepaj npr. [t1,t2,t3,t4] — pusti kot je, TikTok bo rešil
    return raw


def build_tiktok_xlsx(sku: str, brand: str, video_names: str,
                      texts_by_lang: dict, urls_by_lang: dict,
                      skip_rs: bool = False) -> str:
    if not Path(TEMPLATE_PATH).exists():
        raise FileNotFoundError("TikTok template not found. Upload tiktok_template.xlsx to static/")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['Ads']
    headers = [cell.value for cell in ws[1]]

    col_campaign = headers.index('Campaign Name') + 1
    col_bc_id    = headers.index('Business Center ID of the identity') + 1
    col_video    = headers.index('Video Name') + 1
    col_text     = headers.index('Text') + 1
    col_url      = headers.index('Web URL') + 1
    col_ag       = headers.index('Ad Group Name') + 1

    base_bc_raw = ws.cell(row=2, column=col_bc_id).value or ''
    single_id = base_bc_raw.split(',')[0].strip()
    videos = [v.strip() for v in re.findall(r'\[([^\]]+)\]', video_names)]
    new_bc_id = ','.join([single_id] * len(videos)) if videos else single_id
    today = datetime.now().strftime('%-d_%-m_%Y')
    new_campaign = f'[{brand}] Smart+ {sku} - {today}'

    # Zberemo vrstice za brisanje (ne brišemo med iteracijo!)
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        country = ws.cell(row=r, column=col_ag).value
        if not country:
            continue
        lang = COUNTRY_TO_LANG.get(country)
        if skip_rs and lang == 'rs':
            rows_to_delete.append(r)
            continue
        ws.cell(row=r, column=col_campaign).value = new_campaign
        ws.cell(row=r, column=col_bc_id).value = new_bc_id
        ws.cell(row=r, column=col_video).value = video_names
        if lang and lang in texts_by_lang:
            ws.cell(row=r, column=col_text).value = normalize_tiktok_text(texts_by_lang[lang])
        url = (urls_by_lang.get(lang) if lang else None) or next(iter(urls_by_lang.values()), '')
        if url:
            ws.cell(row=r, column=col_url).value = url

    # Zbriši RS vrstice po iteraciji (v obratnem vrstnem redu da ne zamešamo indeksov)
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    out_path = str(EXPORTS_DIR / f"tiktok_{sku}_{uuid.uuid4().hex[:8]}.xlsx")
    wb.save(out_path)
    return out_path



def build_master_xlsx(skus: list) -> str:
    """Združi več SKU-jev v en master XLS — vsak SKU doda svoje vrstice (koliko jih je v template-u)."""
    if not Path(TEMPLATE_PATH).exists():
        raise FileNotFoundError("TikTok template not found.")

    # Load template to get headers and template rows
    wb_tmpl = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_tmpl = wb_tmpl['Ads']
    headers = [cell.value for cell in ws_tmpl[1]]

    col_campaign = headers.index('Campaign Name') + 1
    col_bc_id    = headers.index('Business Center ID of the identity') + 1
    col_video    = headers.index('Video Name') + 1
    col_text     = headers.index('Text') + 1
    col_url      = headers.index('Web URL') + 1
    col_ag       = headers.index('Ad Group Name') + 1

    base_bc_raw = ws_tmpl.cell(row=2, column=col_bc_id).value or ''
    single_id = base_bc_raw.split(',')[0].strip()

    # Collect template rows (row 2 onwards) as base structure
    tmpl_rows = []
    for row in ws_tmpl.iter_rows(min_row=2, values_only=False):
        country = row[col_ag - 1].value
        if not country:
            continue
        tmpl_rows.append({
            'country': country,
            'row_data': [cell.value for cell in row],
            'row_num': row[0].row,
        })

    # Create new workbook
    wb_out = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_out = wb_out['Ads']

    # Clear all data rows
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    today = datetime.now().strftime('%-d_%-m_%Y')
    out_row = 2

    for sku_entry in skus:
        sku = sku_entry.get('sku', '')
        brand = sku_entry.get('brand', '')
        video_names = sku_entry.get('videos', '')
        texts_by_lang = sku_entry.get('texts', {})
        urls_by_lang = sku_entry.get('urls', {})
        fallback_url = sku_entry.get('url') or sku_entry.get('source_url') or ''
        print(f"[master] SKU={sku} url={fallback_url!r} urls={urls_by_lang}")
        # Fallback: če urls prazen, uporabi url za vse jezike
        if not urls_by_lang and fallback_url:
            all_langs = list(set(list(COUNTRY_TO_LANG.values()) + ['sl','hr','rs','hu','cz','sk','pl','gr','ro','bg']))
            urls_by_lang = {lang: fallback_url for lang in all_langs}
        print(f"[master] SKU={sku} urls_by_lang keys={list(urls_by_lang.keys())[:5]}")

        videos = [v.strip() for v in re.findall(r'\[([^\]]+)\]', video_names)]
        new_bc_id = ','.join([single_id] * len(videos)) if videos else single_id
        new_campaign = f'[{brand}] Smart+ {sku} - {today}'

        for tmpl_row in tmpl_rows:
            country = tmpl_row['country']
            lang = COUNTRY_TO_LANG.get(country)

            # Copy template row values
            orig_row = ws_tmpl[tmpl_row['row_num']]
            for col_idx, orig_cell in enumerate(orig_row, 1):
                ws_out.cell(row=out_row, column=col_idx).value = orig_cell.value

            # Fill in our data
            ws_out.cell(row=out_row, column=col_campaign).value = new_campaign
            ws_out.cell(row=out_row, column=col_bc_id).value = new_bc_id
            ws_out.cell(row=out_row, column=col_video).value = video_names
            if lang and lang in texts_by_lang:
                ws_out.cell(row=out_row, column=col_text).value = normalize_tiktok_text(texts_by_lang[lang])
            # URL: lang-specifičen ali fallback
            url = (urls_by_lang.get(lang) if lang else None) or fallback_url or next(iter(urls_by_lang.values()), '')
            if url:
                ws_out.cell(row=out_row, column=col_url).value = url

            out_row += 1

    out_path = str(EXPORTS_DIR / f"master_{uuid.uuid4().hex[:8]}.xlsx")
    wb_out.save(out_path)
    return out_path


class MasterXlsxRequest(BaseModel):
    skus: List[dict]  # [{sku, brand, url, videos, texts, urls}]


@app.post("/build-master-xlsx")
async def build_master_xlsx_endpoint(req: MasterXlsxRequest):
    """Sestavi master XLS iz že shranjenih tekstov — brez API klicev."""
    skus_data = []
    for entry in req.skus:
        if not entry.get('videos'):
            continue
        skus_data.append({
            'sku':    entry.get('sku', ''),
            'brand':  entry.get('brand', ''),
            'videos': entry.get('videos', ''),
            'texts':  entry.get('texts', {}),
            'urls':   entry.get('urls', {}),
        })

    if not skus_data:
        return {"error": "Ni veljavnih SKU-jev (manjkajo video imena)."}

    try:
        path = build_master_xlsx(skus_data)
        return {"status": "ok", "file": path}
    except FileNotFoundError as e:
        return {"error": str(e)}


@app.post("/generate-master-xlsx")
async def generate_master_xlsx(req: MasterXlsxRequest):
    await ensure_cache_fresh()

    skus_data = []
    for entry in req.skus:
        url = entry.get('url', '')
        sku = entry.get('sku', '')
        brand = entry.get('brand', '')
        videos = entry.get('videos', '')

        if not videos:
            continue  # preskoči SKU brez video imen

        # Generate TikTok texts
        user_msg = f"Preberi to stran in ustvari TikTok oglase: {url}"
        data = await generate_tiktok_one(user_msg, "url", url)
        if "error" in data:
            continue

        texts_by_lang = {lang: data[lang] for lang in ["sl","hr","rs","hu","cz","sk","pl","gr","ro","bg"] if lang in data}
        urls_by_lang = data.get("product_urls", {})

        skus_data.append({
            'sku': sku, 'brand': brand, 'videos': videos,
            'texts': texts_by_lang, 'urls': urls_by_lang
        })

        if len(skus_data) < len(req.skus):
            await asyncio.sleep(15)  # rate limit

    if not skus_data:
        return {"error": "Ni veljavnih SKU-jev za generiranje."}

    try:
        path = build_master_xlsx(skus_data)
        return {"status": "ok", "file": path}
    except FileNotFoundError as e:
        return {"error": str(e)}


    input: str
    mode: str
    pt_count: int = 1
    hl_count: int = 1
    source_url: Optional[str] = None
    qmode: str = "sonnet"


class MultiAdRequest(BaseModel):
    products: List[dict]
    pt_count: int = 1
    hl_count: int = 1
    qmode: str = "sonnet"


class TikTokRequest(BaseModel):
    source_url: str
    sku: str
    brand: str
    video_names: str
    skip_rs: bool = False
    product_data: Optional[dict] = None  # maaarket API podatki iz frontenda


# ─── ROUTES ──────────────────────────────────────────────────────────────────

@app.get("/tiktok-history")
async def get_tiktok_history():
    if TT_HISTORY_FILE.exists():
        try:
            return json.loads(TT_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/tiktok-history")
async def save_tiktok_history(data: dict):
    try:
        history = data.get("history", [])
        TT_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}

@app.get("/meta-history")
async def get_meta_history():
    if META_HISTORY_FILE.exists():
        try:
            return json.loads(META_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/meta-history")
async def save_meta_history(data: dict):
    try:
        history = data.get("history", [])
        META_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}

@app.get("/sporocanje-common")
async def sporocanje_get_common(brand: str = "maaarket"):
    """Vrne seznam pogostih odgovorov za določen brand."""
    file = DATA_DIR / f"sporocanje_{brand}.json"
    if file.exists():
        try:
            data = json.loads(file.read_text(encoding="utf-8"))
            return {"ok": True, "answers": data.get("answers", [])}
        except:
            pass
    return {"ok": True, "answers": []}

@app.post("/sporocanje-save")
async def sporocanje_save(data: dict):
    """Shrani odgovor v pogosto bazo za določen brand."""
    brand = data.get("brand", "maaarket")
    file = DATA_DIR / f"sporocanje_{brand}.json"
    existing = {"answers": []}
    if file.exists():
        try:
            existing = json.loads(file.read_text(encoding="utf-8"))
        except:
            pass
    answers = existing.get("answers", [])
    reply_sl = (data.get("reply_sl") or "").strip().lower()
    found = False
    for a in answers:
        if (a.get("reply_sl") or "").strip().lower() == reply_sl:
            a["count"] = a.get("count", 1) + 1
            a["ts"] = data.get("ts", a.get("ts"))
            found = True
            break
    if not found:
        answers.insert(0, {
            "question": data.get("question", ""),
            "reply_sl": data.get("reply_sl", ""),
            "reply_translated": data.get("reply_translated", ""),
            "reply_hr": data.get("reply_hr", ""),
            "lang": data.get("lang", ""),
            "brand": brand,
            "count": 1,
            "ts": data.get("ts", 0),
        })
    answers.sort(key=lambda x: -x.get("count", 1))
    answers = answers[:200]
    existing["answers"] = answers
    file.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "total": len(answers)}

@app.post("/ai-proxy")
async def ai_proxy(data: dict):
    """Proxy za AI klice iz frontenda (za Sporočanje)."""
    prompt = data.get("prompt", "")
    max_tokens = min(int(data.get("max_tokens", 500)), 800)
    model = data.get("model", "claude-haiku-4-5-20251001")
    if not prompt:
        return {"content": [{"text": ""}]}
    msg = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}]
    )
    return {"content": [{"type": "text", "text": msg.content[0].text}]}

@app.post("/kayako-kb-search")
async def kayako_kb_search(data: dict):
    """Poišče top N relevantnih Q&A parov iz KB za dano vprašanje."""
    brand   = data.get("brand", "maaarket")
    query   = data.get("query", "").lower().strip()
    top_n   = min(int(data.get("top_n", 8)), 20)
    if not query:
        return {"ok": True, "results": []}
    kb_file = KB_FILES.get(brand)
    if not kb_file or not kb_file.exists():
        return {"ok": True, "results": []}
    try:
        kb = json.loads(kb_file.read_text(encoding="utf-8"))
        pairs = kb.get("qa_pairs", [])
    except:
        return {"ok": True, "results": []}
    # Keyword scoring
    words = [w for w in query.split() if len(w) > 3]
    scored = []
    for p in pairs:
        hay = (p.get("subject","") + " " + p.get("question","") + " " + p.get("answer","")).lower()
        score = sum(1 for w in words if w in hay)
        if score > 0:
            scored.append((score, p))
    scored.sort(key=lambda x: -x[0])
    results = [p for _, p in scored[:top_n]]
    return {"ok": True, "results": results, "total_kb": len(pairs)}


@app.get("/forecast-entries")
async def get_forecast_entries():
    """Vrne entries samo če so za danes."""
    from datetime import datetime
    try:
        import pytz
        lj = pytz.timezone("Europe/Ljubljana")
        today = datetime.now(lj).strftime("%Y-%m-%d")
    except:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    if not FORECAST_ENTRIES_FILE.exists():
        return {}
    try:
        data = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
        if data.get("date") == today:
            return data
        # Stari podatki — vrni prazen
        print(f"[forecast] entries are from {data.get('date')}, today is {today} — returning empty")
        return {}
    except:
        return {}

@app.post("/forecast-entries")
async def save_forecast_entries(data: dict):
    """Združi poslane entries z obstoječimi po času (multi-user safe).
    Spoštuje deleted-list: vnose ki so bili izbrisani po datumu deletion ne sprejme."""
    try:
        from datetime import datetime
        try:
            import pytz
            lj = pytz.timezone("Europe/Ljubljana")
            today = datetime.now(lj).strftime("%Y-%m-%d")
        except:
            today = datetime.utcnow().strftime("%Y-%m-%d")

        new_entries = data.get("entries", [])
        new_date = data.get("date", today)
        replace_mode = data.get("replace", False)

        existing = {}
        if FORECAST_ENTRIES_FILE.exists():
            try:
                existing = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
            except:
                existing = {}

        if existing.get("date") != today:
            existing = {"date": today, "entries": []}

        if new_date != today:
            print(f"[forecast] ignored save with non-today date {new_date}")
            return {"status": "ok", "note": "date mismatch"}

        # Naloži deleted-list za danes
        deleted = {}
        if FORECAST_DELETED_FILE.exists():
            try:
                all_deleted = json.loads(FORECAST_DELETED_FILE.read_text(encoding="utf-8"))
                deleted = all_deleted.get(today, {})  # {label: timestamp_ms}
            except:
                deleted = {}

        # REPLACE mode — zamenja namesto združi
        if replace_mode:
            # Filtriraj tiste ki so bili brisani
            filtered = [e for e in new_entries if e.get("label","") not in deleted]
            result = {"date": today, "entries": sorted(filtered, key=lambda e: e.get("label",""))}
            FORECAST_ENTRIES_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[forecast] REPLACE save: {len(filtered)} entries")
            return {"status": "ok", "replaced": len(filtered)}

        # MERGE mode
        merged = {}
        for e in existing.get("entries", []):
            lbl = e.get("label","")
            if lbl not in deleted:
                merged[lbl] = e
        for e in new_entries:
            lbl = e.get("label","")
            entry_date = e.get("_date", "")
            # Zavrni entry če _date obstaja in ni danes
            if entry_date and entry_date != today:
                print(f"[forecast] rejected entry '{lbl}' with _date={entry_date} (today={today})")
                continue
            if lbl not in deleted:
                merged[lbl] = e
        sorted_entries = sorted(merged.values(), key=lambda e: e.get("label",""))

        result = {"date": today, "entries": sorted_entries}
        FORECAST_ENTRIES_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[forecast] MERGE save: total {len(sorted_entries)} (filtered out {len(deleted)} deleted)")
        return {"status": "ok", "merged": len(sorted_entries)}
    except Exception as e:
        return {"error": str(e)}

@app.post("/forecast-entry-delete")
async def delete_forecast_entry(data: dict):
    """Izbriši vnos iz danes — doda na deleted-list ki traja čez seje."""
    try:
        from datetime import datetime
        try:
            import pytz
            lj = pytz.timezone("Europe/Ljubljana")
            today = datetime.now(lj).strftime("%Y-%m-%d")
        except:
            today = datetime.utcnow().strftime("%Y-%m-%d")

        label = data.get("label", "")
        if not label:
            return {"error": "missing label"}

        # Naloži deleted-list
        all_deleted = {}
        if FORECAST_DELETED_FILE.exists():
            try:
                all_deleted = json.loads(FORECAST_DELETED_FILE.read_text(encoding="utf-8"))
            except:
                all_deleted = {}

        # Dodaj v deleted za danes
        if today not in all_deleted:
            all_deleted[today] = {}
        import time
        all_deleted[today][label] = int(time.time() * 1000)

        # Pobriši stare dni (>7 dni) — sprosti prostor
        cutoff_keys = sorted(all_deleted.keys())
        if len(cutoff_keys) > 7:
            for k in cutoff_keys[:-7]:
                del all_deleted[k]

        FORECAST_DELETED_FILE.write_text(json.dumps(all_deleted, ensure_ascii=False, indent=2), encoding="utf-8")

        # Takoj odstrani iz entries fajla
        if FORECAST_ENTRIES_FILE.exists():
            try:
                existing = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
                if existing.get("date") == today:
                    existing["entries"] = [e for e in existing.get("entries",[]) if e.get("label","") != label]
                    FORECAST_ENTRIES_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
            except:
                pass

        print(f"[forecast] DELETED entry '{label}' for {today}")
        return {"status": "ok", "deleted": label}
    except Exception as e:
        return {"error": str(e)}

@app.get("/forecast-history")
async def get_forecast_history():
    if FORECAST_HISTORY_FILE.exists():
        try:
            hist = json.loads(FORECAST_HISTORY_FILE.read_text(encoding="utf-8"))
            # Normaliziraj stare sl-SI ključe → ISO format
            changed = False
            new_hist = {}
            for key, val in hist.items():
                # Prepoznaj sl-SI format: "8. 5. 2026"
                if '.' in key and len(key.split('.')) == 3:
                    try:
                        parts = [p.strip().strip('.') for p in key.split('.') if p.strip()]
                        if len(parts) == 3:
                            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                            iso_key = f"{y}-{m:02d}-{d:02d}"
                            if iso_key not in hist:
                                new_hist[iso_key] = val
                                changed = True
                            else:
                                new_hist[key] = val  # ISO verzija že obstaja
                        else:
                            new_hist[key] = val
                    except:
                        new_hist[key] = val
                else:
                    new_hist[key] = val
            if changed:
                FORECAST_HISTORY_FILE.write_text(json.dumps(new_hist, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"[forecast] normalized {sum(1 for k in new_hist if '-' in k)} history keys to ISO format")
                return new_hist
            return hist
        except:
            return {}
    return {}

@app.post("/forecast-history")
async def save_forecast_history(data: dict):
    try:
        FORECAST_HISTORY_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}

@app.get("/")
def root():
    return FileResponse("static/index.html")


@app.get("/cache-status")
async def cache_status():
    return {"last_fetch": last_fetch.isoformat() if last_fetch else None,
            "stale": is_cache_stale(),
            "products_per_lang": {lang: len(p) for lang, p in feed_by_lang.items()},
            "slug_index_size": len(slug_to_id)}


@app.post("/refresh-cache")
async def refresh_cache():
    await fetch_all_feeds()
    return {"status": "ok", "last_fetch": last_fetch.isoformat()}


@app.post("/generate")
async def generate(req: AdRequest):
    await ensure_cache_fresh()
    user_msg = f"Preberi to stran in ustvari Meta oglase: {req.input}" if req.mode == "url" else f"Na podlagi tega opisa ustvari Meta oglase:\n\n{req.input}"
    return await generate_meta_one(user_msg, req.mode, req.source_url, req.pt_count, req.hl_count, req.qmode)


@app.post("/generate-multi")
async def generate_multi(req: MultiAdRequest):
    await ensure_cache_fresh()
    results = []
    for i, p in enumerate(req.products):
        url = p.get("url", "").strip()
        mode = p.get("mode", "url")
        if not url:
            results.append({"error": "Prazen URL"})
            continue
        user_msg = f"Preberi to stran in ustvari Meta oglase: {url}" if mode == "url" else f"Na podlagi tega opisa:\n\n{url}"
        result = await generate_meta_one(user_msg, mode, url if mode == "url" else None,
                                         req.pt_count, req.hl_count, req.qmode)
        results.append(result)
        if i < len(req.products) - 1:
            await asyncio.sleep(15)
    return {"results": results}


@app.post("/generate-multi-stream")
async def generate_multi_stream(req: MultiAdRequest):
    """SSE streaming endpoint — pošilja rezultate batch po batch."""
    await ensure_cache_fresh()

    async def event_stream():
        for i, p in enumerate(req.products):
            url = p.get("url", "").strip()
            mode = p.get("mode", "url")
            if not url:
                yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': {'error': 'Prazen URL'}})}\n\n"
                continue

            # Notify frontend that this product is being processed
            yield f"data: {json.dumps({'type': 'loading', 'index': i, 'url': url})}\n\n"

            user_msg = f"Preberi to stran in ustvari Meta oglase: {url}" if mode == "url" else f"Na podlagi tega opisa:\n\n{url}"

            if req.qmode == "fast":
                # Step 1: SL generation
                yield f"data: {json.dumps({'type': 'progress', 'index': i, 'step': 'sl'})}\n\n"
                result = await generate_meta_sl_only(user_msg, mode, url if mode == "url" else None,
                                                      req.pt_count, req.hl_count)
                if "error" in result:
                    yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': result})}\n\n"
                    if i < len(req.products) - 1:
                        await asyncio.sleep(15)
                    continue

                # Send SL immediately
                yield f"data: {json.dumps({'type': 'partial', 'index': i, 'langs': ['sl'], 'data': result})}\n\n"

                sl_pts = result["sl"]["pt"]
                sl_hls = result["sl"]["hl"]
                product_urls = result.get("product_urls", {})
                product_name = result.get("product", "Izdelek")

                # Step 2: Parallel translation in 4 batches of 2-3 langs
                pt_ph = ", ".join([f'"PT {i2+1}"' for i2 in range(req.pt_count)])
                hl_ph = ", ".join([f'"HL {i2+1}"' for i2 in range(req.hl_count)])

                lang_batches = [
                    ["hr", "rs"],
                    ["hu", "cz"],
                    ["sk", "pl"],
                    ["gr", "ro", "bg"],
                ]

                lang_info = {
                    "hr": "HR (hrvaščina, latinica)",
                    "rs": "RS (srbščina, SAMO latinica!)",
                    "hu": "HU (madžarščina - aglutinacijski jezik, ne prevajaj dobesedno)",
                    "cz": "CZ (češčina)",
                    "sk": "SK (slovaščina)",
                    "pl": "PL (poljščina)",
                    "gr": "GR (grščina, grška pisava!)",
                    "ro": "RO (romunščina, latinica)",
                    "bg": "BG (bolgarščina, SAMO cirilica!)",
                }

                full_result = {
                    "product": product_name,
                    "product_urls": product_urls,
                    "sl": {"pt": sl_pts, "hl": sl_hls},
                }

                for batch in lang_batches:
                    yield f"data: {json.dumps({'type': 'progress', 'index': i, 'step': 'translating', 'langs': batch})}\n\n"

                    batch_json_keys = ", ".join([
                        f'"{lang}":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}}'
                        for lang in batch
                    ])
                    batch_lang_lines = "\n".join([f"- {lang_info[lang]}" for lang in batch])

                    batch_prompt = f"""Prevedi Meta oglase iz slovenščine v naslednje jezike. Ohrani ŠTEVILO in POZICIJO emoji-jev točno kot v originalu.

Primary Texts: {json.dumps(sl_pts, ensure_ascii=False)}
Headlines: {json.dumps(sl_hls, ensure_ascii=False)}

Prevedi SAMO v te jezike:
{batch_lang_lines}

SPLOŠNA PRAVILA:
- Ohrani prodajni/energičen ton
- Prevodi morajo zveneti kot materni govorec
- Ohrani ŠTEVILO emoji-jev
- Headlines: MAX 5 besed
- Ne prevajaj imen izdelkov/blagovnih znamk

Vrni SAMO JSON: {{{batch_json_keys}}}"""

                    # Retry do 3x
                    batch_data = None
                    for attempt in range(3):
                        try:
                            batch_text = await call_claude(batch_prompt, "claude-haiku-4-5-20251001", None, 3000)
                            batch_data = parse_json_response(batch_text)
                            if batch_data:
                                # Preveri da so vsi jeziki iz batcha prisotni
                                missing = [lang for lang in batch if lang not in batch_data]
                                if missing:
                                    print(f"[meta-stream] batch {batch} manjkajo jeziki: {missing}, retry {attempt+1}")
                                    batch_data = None
                                else:
                                    break
                            else:
                                print(f"[meta-stream] batch {batch} JSON parse fail, retry {attempt+1}, response[:200]: {batch_text[:200]}")
                        except Exception as e:
                            print(f"[meta-stream] batch {batch} error attempt {attempt+1}: {e}")
                        if attempt < 2:
                            await asyncio.sleep(5)

                    if batch_data:
                        full_result.update(batch_data)
                        yield f"data: {json.dumps({'type': 'partial', 'index': i, 'langs': batch, 'data': full_result})}\n\n"
                    else:
                        print(f"[meta-stream] batch {batch} FAILED po 3 poskusih — jeziki manjkajo v rezultatu")
                        yield f"data: {json.dumps({'type': 'batch_error', 'index': i, 'langs': batch, 'error': f'Prevod za {batch} ni uspel'})}\n\n"

                yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': full_result})}\n\n"

            else:
                # Kreativni način — en klic, pošlji ko konča
                result = await generate_meta_one(user_msg, mode, url if mode == "url" else None,
                                                  req.pt_count, req.hl_count, req.qmode)
                yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': result})}\n\n"

            if i < len(req.products) - 1:
                await asyncio.sleep(15)

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/generate-tiktok")
async def generate_tiktok(req: TikTokRequest):
    await ensure_cache_fresh()

    # Če imamo API podatke iz frontenda → uporabimo jih (brez web_search)
    if req.product_data and req.product_data.get("title"):
        pd = req.product_data

        # Zberi features opise
        features_text = ""
        for f in (pd.get("features") or []):
            t = f.get("title", "")
            c = f.get("content", "")
            if t or c:
                features_text += f"\n- {t}: {c}"

        # Reviews
        reviews_text = ""
        for r in (pd.get("comments") or [])[:3]:
            reviews_text += f'\n- "{r.get("content","")}" — {r.get("name","")}'

        sales = pd.get("sales_count", "")
        rating = ""
        try:
            rating = pd.get("rating", {}).get("total", {}).get("average", "")
        except:
            pass

        user_msg = f"""Napiši TikTok oglase za ta izdelek. Piši kot izkušen copywriter — ne naštevaj specifikacij, piši zgodbe in čustva.

IME: {pd.get('title', '')}
OPIS: {pd.get('lead', '') or pd.get('short_desc', '')}
PODROBNOSTI: {pd.get('content', '')[:800]}
{f"PREDNOSTI:{features_text}" if features_text else ""}
{f"PRODANO: {sales}x" if sales else ""}
{f"OCENE kupcev:{reviews_text}" if reviews_text else ""}
{f"POVPREČNA OCENA: {rating}/5" if rating else ""}

Ustvari TikTok oglase za ta izdelek."""

        print(f"[tiktok] API podatki: {pd.get('title','')[:50]} | {sales} prodaj | {rating}★")
        data = await generate_tiktok_one(user_msg, "text", req.source_url)
    else:
        # Fallback: stari način z web_search
        user_msg = f"Preberi to stran in ustvari TikTok oglase: {req.source_url}"
        print(f"[tiktok] Fallback web_search za: {req.source_url}")
        data = await generate_tiktok_one(user_msg, "url", req.source_url)

    if "error" in data:
        return data

    product_urls = data.get("product_urls", {})
    texts_by_lang = {lang: data[lang] for lang in ["sl","hr","rs","hu","cz","sk","pl","gr","ro","bg"] if lang in data}

    try:
        xlsx_path = build_tiktok_xlsx(
            sku=req.sku,
            brand=req.brand,
            video_names=req.video_names,
            texts_by_lang=texts_by_lang,
            urls_by_lang=product_urls,
            skip_rs=req.skip_rs
        )
        return {"status": "ok", "file": xlsx_path, "data": data}
    except FileNotFoundError as e:
        return {"error": str(e)}


@app.post("/extract-videos")
async def extract_videos(data: dict):
    """Extract video filenames from a base64 screenshot using Claude vision."""
    image_b64 = data.get("image")
    media_type = data.get("media_type", "image/png")
    if not image_b64:
        return {"error": "Ni slike."}
    loop = asyncio.get_event_loop()
    msg = await loop.run_in_executor(None, lambda: client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=500,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
            {"type": "text", "text": "Extract all video filenames from this screenshot. Return ONLY a JSON array of filenames, nothing else. Example: [\"VIDEO (1).mp4\", \"VIDEO (2).mp4\"]. Extract exactly as written, preserve spaces and capitalization."}
        ]}]
    ))
    text = msg.content[0].text.strip()
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text).strip()
    try:
        filenames = json.loads(text)
        formatted = ",".join([f"[{f}]" for f in filenames])
        return {"filenames": filenames, "formatted": formatted}
    except json.JSONDecodeError:
        return {"error": "Ni uspelo prebrati imen. Poskusi znova z jasnejšo sliko."}


@app.get("/download/{filename}")
def download_file(filename: str):
    path = EXPORTS_DIR / filename
    if not path.exists():
        return {"error": "File not found"}
    return FileResponse(str(path), filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/kreative-history")
async def get_kreative_history():
    if KREATIVE_HISTORY_FILE.exists():
        try:
            return json.loads(KREATIVE_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/kreative-history")
async def save_kreative_history(data: dict):
    try:
        history = data.get("history", [])
        KREATIVE_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}


# ─── KREATIVE ENDPOINTS ───────────────────────────────────────────────────────

@app.post("/analyze-product-kreative")
async def analyze_product_kreative(data: dict):
    """Prebere stran izdelka in generira A/B/C opcije za kreative."""
    url = data.get("url", "").strip()
    if not url:
        return {"error": "Manjka URL."}

    # Fetch page
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as hc:
            resp = await hc.get(url, headers={"User-Agent": "Mozilla/5.0"})
            html = resp.text if resp.status_code == 200 else ""
    except Exception as e:
        return {"error": f"Ne morem prebrati strani: {e}"}


    # Build prompt for Claude to analyze product
    tools = [{"type": "web_search_20250305", "name": "web_search"}]
    analysis_prompt = f"""Preberi to spletno stran izdelka: {url}

Na podlagi opisa izdelka generiraj strukturirane podatke za kreative FB oglasov.

Vrni SAMO JSON (brez markdown) v tej obliki:
{{
  "name": "IME_IZDELKA_CAPS",
  "aOptions": [
    {{"label": "MAIN VERSION", "text": "benefit1, benefit2, benefit3"}},
    {{"label": "HIGH-CONVERT", "text": "benefit1, benefit2, benefit3"}},
    {{"label": "PROBLEM-SOLUTION", "text": "benefit1, benefit2, benefit3"}},
    {{"label": "ULTRA SIMPLE", "text": "benefit1, benefit2"}},
    {{"label": "SOCIAL PROOF", "text": "benefit1, benefit2, benefit3"}}
  ],
  "bOptions": [
    {{"label": "BEST", "text": "kratko ime vibeа ozadja"}},
    {{"label": "SECOND OPTION", "text": "kratko ime vibeа ozadja"}},
    {{"label": "SCROLL STOPPER", "text": "kratko ime vibeа ozadja"}},
    {{"label": "BONUS", "text": "kratko ime vibeа ozadja"}},
    {{"label": "LIFESTYLE", "text": "kratko ime vibeа ozadja"}}
  ]
}}

PRAVILA:
- name: samo ime blagovne znamke/modela z CAPS
- aOptions: PRODAJNI UDARCI (ne lastnosti!), max 2-3 besede vsak, v angleščini
- bOptions: KONCEPT/VIBE specifičen za TA izdelek, max 3 besede, vedno akcija ali moment
- VSE mora biti v angleščini
- Vrni točno 5 aOptions in 5 bOptions

PRIMERI (uči se iz teh vzorcev):

WEEDZAP (weed removal tool):
aOptions: "Pull the Root, Stop the Weed, No Chemicals" | "Weeds Gone, Root and All, One Tool" | "Weeds Keep Coming Back, Pull the Root, Done for Good" | "Twist, Pull, Gone"
bOptions: "root pull satisfying moment" | "no chemicals angle" | "before/after garden" | "no back pain angle"

ASHIRAFLUX (smokeless fire pit):
aOptions: "Real Flame, No Smoke, Anywhere You Are" | "No Fireplace? No Problem, Instant Cozy, Zero Smoke" | "Missing That Fire Feeling, Warm Vibes Instantly" | "Fill, Light, Relax"
bOptions: "evening atmosphere shot" | "smoke vs no smoke" | "first light moment" | "indoor safe angle"

STAXA (steel organizer shelf):
aOptions: "Double Your Space, Zero Clutter, Instant Order" | "Messy Counter, One Shelf, Problem Solved" | "No Room, Stack It Up, Space Created" | "Stack, Store, Done"
bOptions: "before/after counter" | "multi-room tour" | "satisfying load test" | "steel vs plastic"

SIZZELA (electric frying pan):
aOptions: "No Stove Needed, Cook Anywhere, Instant Heat" | "One Pan, Any Meal, Zero Hassle" | "No Stove, No Smoke, No Problem" | "Plug & Sizzle, Done"
bOptions: "sizzle sound moment" | "speed cooking demo" | "steam lid reveal" | "no stove freedom"

PLANTDRILL (garden auger):
aOptions: "Drill, Plant, Done, No Digging" | "One Bit, Perfect Holes, Zero Effort" | "Back-Breaking Digging, One Drill Bit, Done" | "Drill, Drop, Grow"
bOptions: "speed demo" | "planting demo" | "satisfying drill moment"

SMARTFITNESS (EMS stimulator):
aOptions: "Train Anywhere, No Gym, Real Results" | "On the Couch, Still Training, Zero Effort" | "No Time to Work Out, Wear It, Feel It Work" | "Stick, Activate, Tone"
bOptions: "lifestyle demo" | "before/after body" | "reaction moment"

SOWSYNC (seed spacing tool):
aOptions: "Plant Smart, Perfect Spacing, Grow Better" | "Even Rows, No Waste, Strong Growth" | "Messy Planting, Seed Tool, Better Yield" | "Place, Press, Plant"
bOptions: "before/after planting" | "planting demo" | "grid effect"

Sedaj generiraj za izdelek na tej strani po ISTEM vzorcu:"""

    text = await call_claude(analysis_prompt, "claude-sonnet-4-6", tools, 800)
    result = parse_json_response(text)

    if not result:
        return {"error": "Ni uspelo analizirati izdelka. Poskusi znova."}

    return result


@app.post("/fetch-product-images")
async def fetch_product_images(data: dict):
    """Pobere slike izdelka s podane URL strani."""
    url = data.get("url", "").strip()
    if not url:
        return {"error": "Manjka URL."}
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as hc:
            resp = await hc.get(url, headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200:
                return {"error": f"Stran ni dostopna ({resp.status_code})."}
            html = resp.text

        # Extract img src attributes
        import re as _re
        srcs = _re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', html, _re.IGNORECASE)

        # Filter: only product-like images (skip icons, logos, tiny tracking pixels)
        from urllib.parse import urljoin
        base = url
        images = []
        seen = set()
        for src in srcs:
            if not src or src.startswith("data:"):
                continue
            full = urljoin(base, src)
            # Skip obviously non-product images
            skip_words = ["logo", "icon", "favicon", "sprite", "banner", "flag",
                         "payment", "badge", "star", "rating", "arrow", "tracking"]
            if any(w in full.lower() for w in skip_words):
                continue
            # Only jpg/png/webp
            if not any(ext in full.lower() for ext in [".jpg", ".jpeg", ".png", ".webp"]):
                continue
            if full not in seen:
                seen.add(full)
                images.append(full)
            if len(images) >= 20:
                break

        return {"images": images, "count": len(images)}
    except Exception as e:
        return {"error": str(e)}


@app.post("/generate-kreative")
async def generate_kreative(data: dict):
    """Generira kreative z Google Gemini (Nano Banana 2) API."""
    import base64, struct, zlib

    product_name = data.get("productName", "")
    a_options = data.get("aOptions", [])
    b_options = data.get("bOptions", [])
    count = data.get("count", 4)
    ref_images = data.get("images", [])  # base64 data URLs

    if not product_name or not a_options or not b_options:
        return {"error": "Manjkajo podatki (ime izdelka, A ali B opcije)."}

    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key:
        return {"error": "GEMINI_API_KEY ni nastavljen v Render environment variables."}

    # ── GEMINI FILE URI CACHE ──────────────────────────────────────────────────
    # Gemini Files API vrne URI ki velja 48h — cachiramo na disk da ne uploadamo vsakič
    import hashlib, json as _json, time as _time
    GEMINI_CACHE_FILE = DATA_DIR / "gemini_file_cache.json"

    def _load_cache():
        try:
            if GEMINI_CACHE_FILE.exists():
                return _json.loads(GEMINI_CACHE_FILE.read_text(encoding="utf-8"))
        except: pass
        return {}

    def _save_cache(cache):
        try: GEMINI_CACHE_FILE.write_text(_json.dumps(cache, ensure_ascii=False), encoding="utf-8")
        except: pass

    def _img_hash(b64_data):
        # Hash samo prvih 2000 znakov (dovolj za unikatnost, hitro)
        return hashlib.md5(b64_data[:2000].encode()).hexdigest()

    gemini_cache = _load_cache()
    now_ts = _time.time()
    cache_updated = False

    # Build combinations
    combos = []
    for a in a_options:
        for b in b_options:
            prompt = (
                f"From these reference images create a new FB ad creative. "
                f"Try a more intense '{b.get('text', '')}' background. "
                f"Do not include any text/words on the image except the device name in capital letters '{product_name}' — place it where it fits best or makes sense. "
                f"If possible (if you recognize any suitable English naming styles), you can also create a logo from the name. "
                f"Highlight (can be through icons or text in English) that it is: {a.get('text', '')}. "
                f"Keep all text and icons well within the image borders — nothing should be cut off at the edges. Square 1:1 format."
            )
            combos.append({
                "combo": f"{a.get('label','A')} × {b.get('label','B')}",
                "prompt": prompt,
            })

    # Prepare reference image parts for Gemini
    # Upload referenčne slike enkrat na Gemini Files API — z 48h cache
    file_uris = []
    if ref_images:
        async with httpx.AsyncClient(timeout=60.0) as hc:
            for img_data in ref_images:  # vse referenčne slike, cache poskrbi za optimizacijo
                try:
                    if "," in img_data:
                        header, b64 = img_data.split(",", 1)
                        mime = header.split(":")[1].split(";")[0]
                    else:
                        b64 = img_data
                        mime = "image/jpeg"

                    # Preveri cache — če URI še velja (< 46h star), preskoči upload
                    img_key = _img_hash(b64)
                    cached = gemini_cache.get(img_key)
                    if cached and (now_ts - cached.get("ts", 0)) < 46 * 3600:
                        # Cache hit — ne uploadamo znova
                        file_uris.append({"fileData": {"mimeType": cached["mime"], "fileUri": cached["uri"]}})
                        continue

                    # Cache miss — uploadaj
                    img_bytes = __import__("base64").b64decode(b64)
                    upload_url = f"https://generativelanguage.googleapis.com/upload/v1beta/files?key={gemini_key}"
                    resp = await hc.post(
                        upload_url,
                        content=img_bytes,
                        headers={"Content-Type": mime, "X-Goog-Upload-Content-Type": mime,
                                 "X-Goog-Upload-Protocol": "raw"}
                    )
                    if resp.status_code == 200:
                        uri = resp.json().get("file", {}).get("uri", "")
                        if uri:
                            file_uris.append({"fileData": {"mimeType": mime, "fileUri": uri}})
                            # Shrani v cache
                            gemini_cache[img_key] = {"uri": uri, "mime": mime, "ts": now_ts}
                            cache_updated = True
                except Exception:
                    continue

        if cache_updated:
            _save_cache(gemini_cache)

    # Fallback: če Files API ne dela, uporabi inline za prvo sliko
    if file_uris:
        image_parts = file_uris
    elif ref_images:
        try:
            img_data = ref_images[0]
            if "," in img_data:
                header, b64 = img_data.split(",", 1)
                mime = header.split(":")[1].split(";")[0]
            else:
                b64 = img_data; mime = "image/jpeg"
            image_parts = [{"inline_data": {"mime_type": mime, "data": b64}}]
        except Exception:
            image_parts = []
    else:
        image_parts = []

    async def generate_one_image(combo_prompt, combo_label, idx):
        api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-flash-image-preview:generateContent?key={gemini_key}"
        parts = list(image_parts) + [{"text": combo_prompt}]
        payload = {
            "contents": [{"parts": parts}],
            "generationConfig": {"responseModalities": ["IMAGE", "TEXT"]}
        }
        try:
            async with httpx.AsyncClient(timeout=120.0) as hc:
                resp = await hc.post(api_url, json=payload, headers={"Content-Type": "application/json"})
                result = resp.json()
            if resp.status_code != 200:
                return None, result.get("error", {}).get("message", str(result))
            for candidate in result.get("candidates", []):
                for part in candidate.get("content", {}).get("parts", []):
                    if "inlineData" in part:
                        inline = part["inlineData"]
                        mime = inline.get("mimeType", "image/png")
                        b64_data = inline.get("data", "")
                        return f"data:{mime};base64,{b64_data}", None
            return None, "Gemini ni vrnil slike: " + str(result)[:150]
        except Exception as e:
            return None, str(e)

    async def generate_combo(combo):
        """Generate `count` images for one combo in parallel."""
        tasks = [generate_one_image(combo["prompt"], combo["combo"], i) for i in range(count)]
        img_results = await asyncio.gather(*tasks)
        imgs = [img for img, err in img_results if img]
        errors = [err for img, err in img_results if err]
        if not imgs:
            return {"combo": combo["combo"], "images": [], "error": errors[0] if errors else "Ni slike"}
        return {"combo": combo["combo"], "images": imgs}

    # Vse kombinacije + vse slike vzporedno
    results = await asyncio.gather(*[generate_combo(combo) for combo in combos])
    return {"results": list(results)}


# ─── ASANA ENDPOINTS ──────────────────────────────────────────────────────────

ASANA_API = "https://app.asana.com/api/1.0"

def asana_headers():
    token = os.environ.get("ASANA_API_KEY", "")
    return {"Authorization": f"Bearer {token}", "Accept": "application/json"}

@app.get("/asana-search")
async def asana_search(q: str = ""):
    """Išče Asana taske po imenu."""
    if not q:
        return {"tasks": []}
    token = os.environ.get("ASANA_API_KEY", "")
    if not token:
        return {"error": "ASANA_API_KEY ni nastavljen."}
    try:
        async with httpx.AsyncClient(timeout=15.0) as hc:
            # Get workspaces first
            ws_resp = await hc.get(f"{ASANA_API}/workspaces", headers=asana_headers())
            workspaces = ws_resp.json().get("data", [])
            if not workspaces:
                return {"error": "Ni najdenih workspace-ov."}
            ws_gid = workspaces[0]["gid"]

            # Search tasks
            params = {"workspace": ws_gid, "text": q, "resource_type": "task",
                      "opt_fields": "gid,name,projects.name"}
            search_resp = await hc.get(f"{ASANA_API}/workspaces/{ws_gid}/tasks/search",
                                        params=params, headers=asana_headers())
            tasks_raw = search_resp.json().get("data", [])

            tasks = []
            for t in tasks_raw[:10]:
                projects = t.get("projects", [])
                proj_name = projects[0]["name"] if projects else ""
                tasks.append({"gid": t["gid"], "name": t["name"], "project": proj_name})

            return {"tasks": tasks}
    except Exception as e:
        return {"error": str(e)}


@app.post("/asana-attach")
async def asana_attach(data: dict):
    """Priloži slike (base64 data URLs) na Asana task."""
    task_id = data.get("task_id", "")
    image_urls = data.get("image_urls", [])

    if not task_id or not image_urls:
        return {"error": "Manjka task_id ali slike."}

    token = os.environ.get("ASANA_API_KEY", "")
    if not token:
        return {"error": "ASANA_API_KEY ni nastavljen."}

    attached = 0
    errors = []

    async with httpx.AsyncClient(timeout=60.0) as hc:
        for i, img_url in enumerate(image_urls):
            try:
                # Decode base64 data URL
                if img_url.startswith("data:"):
                    header, b64data = img_url.split(",", 1)
                    mime = header.split(":")[1].split(";")[0]
                    ext = mime.split("/")[1] if "/" in mime else "png"
                    img_bytes = __import__("base64").b64decode(b64data)
                else:
                    # Regular URL — fetch it
                    img_resp = await hc.get(img_url)
                    img_bytes = img_resp.content
                    mime = img_resp.headers.get("content-type", "image/png")
                    ext = "png"

                filename = f"kreativa_{i+1}.{ext}"

                # Upload to Asana as attachment
                files = {"file": (filename, img_bytes, mime)}
                attach_resp = await hc.post(
                    f"{ASANA_API}/tasks/{task_id}/attachments",
                    headers={"Authorization": f"Bearer {token}"},
                    files=files
                )

                if attach_resp.status_code in (200, 201):
                    attached += 1
                else:
                    errors.append(f"Slika {i+1}: {attach_resp.text[:100]}")

            except Exception as e:
                errors.append(f"Slika {i+1}: {str(e)}")

    return {"attached": attached, "errors": errors, "total": len(image_urls)}


# ─── LOKALIZACIJA ENDPOINT ───────────────────────────────────────────────────

LANG_NAMES = {
    "HR": "Croatian", "RS": "Serbian (Latin script)",
    "HU": "Hungarian", "CZ": "Czech", "SK": "Slovak",
    "PL": "Polish", "RO": "Romanian", "BG": "Bulgarian",
    "GR": "Greek", "SL": "Slovenian"
}

@app.post("/localize-kreativa")
async def localize_kreativa(data: dict):
    """Prevede tekst na hero kreativu v izbrane jezike z Gemini."""
    image_data = data.get("image", "") or (data.get("images", [None])[0] or "")
    images_data = data.get("images", [])
    if not images_data and image_data:
        images_data = [image_data]
    languages = data.get("languages", [])
    asana_task_id = data.get("asana_task_id")
    sku = data.get("sku", "SKU").strip().upper()
    brand = data.get("brand", "").strip()

    if not images_data or not languages:
        return {"error": "Manjka slika ali jeziki."}

    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key:
        return {"error": "GEMINI_API_KEY ni nastavljen."}

    # Decode base64 image
    try:
        if "," in image_data:
            header, b64 = image_data.split(",", 1)
            mime = header.split(":")[1].split(";")[0]
        else:
            b64 = image_data
            mime = "image/jpeg"
    except Exception as e:
        return {"error": f"Napaka pri dekodiranju slike: {e}"}

    async def translate_one(lang_code, img_b64, img_mime, img_idx):
        lang_name = LANG_NAMES.get(lang_code, lang_code)
        api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-flash-image-preview:generateContent?key={gemini_key}"
        brand_note = f" Do NOT translate these brand/product names (keep exactly as-is): {brand}." if brand else ""
        prompt = (
            f"Edit this image to translate all visible text into {lang_name}. "
            f"Keep EVERYTHING exactly the same — same people, same background, same layout, same design, same product, same icons, same colors, same fonts. "
            f"ONLY translate the text that is NOT a brand name or logo."
            f"{brand_note} "
            f"Do NOT translate or modify any brand names, logos, or product names. "
            f"Keep all text in the same position, same style, same size."
        )
        payload = {
            "contents": [{"parts": [
                {"inline_data": {"mime_type": img_mime, "data": img_b64}},
                {"text": prompt}
            ]}],
            "generationConfig": {"responseModalities": ["IMAGE", "TEXT"]}
        }
        try:
            async with httpx.AsyncClient(timeout=120.0) as hc:
                resp = await hc.post(api_url, json=payload, headers={"Content-Type": "application/json"})
                result = resp.json()
            if resp.status_code != 200:
                return {"lang": lang_code, "lang_name": lang_name, "url": None,
                        "error": result.get("error", {}).get("message", str(result))[:200]}
            for candidate in result.get("candidates", []):
                for part in candidate.get("content", {}).get("parts", []):
                    if "inlineData" in part:
                        out_mime = part["inlineData"].get("mimeType", "image/png")
                        out_b64 = part["inlineData"].get("data", "")
                        img_url = f"data:{out_mime};base64,{out_b64}"
                        filename = f"{sku}_{lang_code}_v{img_idx}.png"
                        asana_ok = False
                        if asana_task_id:
                            try:
                                img_bytes = __import__("base64").b64decode(out_b64)
                                token = os.environ.get("ASANA_API_KEY", "")
                                async with httpx.AsyncClient(timeout=30.0) as hc2:
                                    attach_resp = await hc2.post(
                                        f"{ASANA_API}/tasks/{asana_task_id}/attachments",
                                        headers={"Authorization": f"Bearer {token}"},
                                        files={"file": (filename, img_bytes, out_mime)}
                                    )
                                asana_ok = attach_resp.status_code in (200, 201)
                            except Exception:
                                pass
                        return {"lang": lang_code, "lang_name": lang_name, "url": img_url, "filename": filename, "asana_ok": asana_ok}
            return {"lang": lang_code, "lang_name": lang_name, "url": None, "error": "Ni slike"}
        except Exception as e:
            return {"lang": lang_code, "lang_name": lang_name, "url": None, "error": str(e)}

    # Pripravi vse kombinacije slika × jezik vzporedno
    tasks = []
    for img_idx, img_data_item in enumerate(images_data, 1):
        try:
            if "," in img_data_item:
                header, b64 = img_data_item.split(",", 1)
                mime = header.split(":")[1].split(";")[0]
            else:
                b64 = img_data_item; mime = "image/jpeg"
        except Exception:
            continue
        for lang_code in languages:
            tasks.append(translate_one(lang_code, b64, mime, img_idx))
    results = await asyncio.gather(*tasks)
    return {"results": list(results)}


# ─── NAROČILNICE HISTORY ─────────────────────────────────────────────────────

NAROCILNICE_HISTORY_FILE = DATA_DIR / "narocilnice_history.json"

@app.get("/narocilnice-history")
async def get_narocilnice_history():
    if NAROCILNICE_HISTORY_FILE.exists():
        try:
            return json.loads(NAROCILNICE_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/narocilnice-history")
async def save_narocilnice_history(data: dict):
    try:
        history = []
        if NAROCILNICE_HISTORY_FILE.exists():
            try:
                history = json.loads(NAROCILNICE_HISTORY_FILE.read_text(encoding="utf-8"))
            except:
                history = []
        
        csv_text = data.get("csv", "")
        date = data.get("date", "")
        
        # Count negative rows
        rows = 0
        for line in csv_text.split('\n')[1:]:
            if line.strip():
                rows += 1
        
        history.append({"csv": csv_text, "date": date, "rows": rows})
        # Keep last 50
        if len(history) > 50:
            history = history[-50:]
        
        NAROCILNICE_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}


# ─── NAROČILNICE SKU LOOKUP ───────────────────────────────────────────────────

@app.post("/narocilnice-lookup")
async def narocilnice_lookup(data: dict):
    """Poišče URL izdelka v Maaarket XML feedu po SKU ali nazivu."""
    skus = data.get("skus", [])  # list of {sku, naziv}
    
    await ensure_cache_fresh()
    
    results = {}
    sl_feed = feed_by_lang.get("sl", {})
    
    for item in skus:
        sku = item.get("sku", "").strip().upper()
        naziv = item.get("naziv", "").strip().lower()
        
        found_url = None
        
        # Search by SKU in title/id
        for g_id, prod in sl_feed.items():
            prod_title = prod.get("title", "").strip()
            prod_url = prod.get("url", "")
            prod_slug = extract_slug(prod_url) or ""
            
            # Match SKU in title or slug
            if (sku and (sku.lower() in prod_title.lower() or sku.lower() in prod_slug.lower())):
                found_url = prod_url
                break
        
        # Fallback: search by naziv words (first 3 significant words)
        if not found_url and naziv:
            words = [w for w in naziv.split() if len(w) > 3][:3]
            if words:
                best_score = 0
                for g_id, prod in sl_feed.items():
                    prod_title = prod.get("title", "").strip().lower()
                    score = sum(1 for w in words if w in prod_title)
                    if score > best_score and score >= 2:
                        best_score = score
                        found_url = prod.get("url", "")
        
        if found_url:
            results[sku] = found_url
    
    return {"urls": results}


@app.post("/narocilnice-history-set")
async def set_narocilnice_history(data: dict):
    """Nastavi celotno zgodovino (za brisanje)."""
    try:
        history = data.get("history", [])
        NAROCILNICE_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}



# ─── FORECAST EOD (End-of-day končna naročila) ───────────────────────────────

FORECAST_EOD_FILE = DATA_DIR / "forecast_eod.json"

@app.get("/forecast-eod")
async def get_forecast_eod():
    if FORECAST_EOD_FILE.exists():
        try:
            return json.loads(FORECAST_EOD_FILE.read_text(encoding="utf-8"))
        except:
            pass
    return {}

@app.post("/forecast-eod")
async def save_forecast_eod(data: dict):
    try:
        FORECAST_EOD_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(data)}
    except Exception as e:
        return {"error": str(e)}


# ─── KARANTENA PDF PARSER ─────────────────────────────────────────────────────

from fastapi import UploadFile, File, Form
import io
import re as _re

@app.post("/parse-karantena-pdf")
async def parse_karantena_pdf(file: UploadFile = File(...)):
    """Parsira PDF karantene in vrne strukturirane podatke."""
    try:
        import pdfplumber
        content = await file.read()
        rows = []
        
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            full_text = ""
            for page in pdf.pages:
                # Poskusi extract table najprej
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if not row or not any(row):
                                continue
                            # Preskoči header vrstice
                            if row[0] and str(row[0]).strip().lower() in ('product_id', 'id', '#'):
                                continue
                            # Normalizacija vrstice
                            cells = [str(c).strip() if c else '' for c in row]
                            if len(cells) >= 3:
                                product_id = cells[0] if cells[0] else ''
                                sku = cells[1] if len(cells) > 1 else ''
                                title = cells[2] if len(cells) > 2 else ''
                                stock = cells[3] if len(cells) > 3 else '0'
                                stock_actual = cells[4] if len(cells) > 4 else '0'
                                position = cells[5] if len(cells) > 5 else ''
                                
                                # Preskoči header
                                if sku.lower() in ('product_sku', 'sku', ''):
                                    continue
                                
                                try: stock = int(float(stock))
                                except: stock = 0
                                try: stock_actual = int(float(stock_actual))
                                except: stock_actual = 0
                                
                                rows.append({
                                    'product_id': product_id,
                                    'sku': sku,
                                    'title': title,
                                    'stock': stock,
                                    'stock_actual': stock_actual,
                                    'position': position,
                                })
                else:
                    full_text += (page.extract_text() or "") + "\n"
            
            # Fallback: text parsing če ni tabel
            if not rows and full_text:
                lines = full_text.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split()
                    if len(parts) >= 3 and parts[0].isdigit():
                        product_id = parts[0]
                        sku = parts[1]
                        # Preskoči header
                        if sku.lower() in ('product_sku', 'sku'):
                            continue
                        # Poišči zadnji del ki je številka (stock)
                        stock = 0
                        stock_actual = 0
                        position = ''
                        title_parts = []
                        for i, p in enumerate(parts[2:], 2):
                            if _re.match(r'^\d+$', p):
                                stock = int(p)
                                if i+1 < len(parts) and _re.match(r'^\d+$', parts[i+1]):
                                    stock_actual = int(parts[i+1])
                                    if i+2 < len(parts):
                                        position = parts[i+2]
                                break
                            else:
                                title_parts.append(p)
                        title = ' '.join(title_parts)
                        rows.append({
                            'product_id': product_id,
                            'sku': sku,
                            'title': title,
                            'stock': stock,
                            'stock_actual': stock_actual,
                            'position': position,
                        })
        
        if not rows:
            return {"error": "Ni podatkov v PDF-u."}
        
        print(f"[karantena] Parsed {len(rows)} rows from PDF")
        return {"rows": rows, "count": len(rows)}
        
    except ImportError:
        return {"error": "pdfplumber ni nameščen. Dodaj ga v requirements.txt."}
    except Exception as e:
        import traceback
        print(f"[karantena] Error: {e}\n{traceback.format_exc()[-500:]}")
        return {"error": str(e)}


# ─── KARANTENA HISTORY ────────────────────────────────────────────────────────

KARANTENA_HISTORY_FILE = DATA_DIR / "karantena_history.json"

@app.get("/karantena-history")
async def get_karantena_history():
    if KARANTENA_HISTORY_FILE.exists():
        try:
            return json.loads(KARANTENA_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/karantena-history")
async def save_karantena_history(data: dict):
    try:
        history = []
        if KARANTENA_HISTORY_FILE.exists():
            try:
                history = json.loads(KARANTENA_HISTORY_FILE.read_text(encoding="utf-8"))
            except:
                history = []
        history.append({
            "rows": data.get("rows", []),
            "filename": data.get("filename", ""),
            "date": data.get("date", "")
        })
        if len(history) > 30:
            history = history[-30:]
        KARANTENA_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}

@app.post("/karantena-history-set")
async def set_karantena_history(data: dict):
    try:
        history = data.get("history", [])
        KARANTENA_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}


# ─── VIDEO ADS — SCRIPT GENERATION ───────────────────────────────────────────

@app.post("/generate-video-scripts")
async def generate_video_scripts(data: dict):
    input_text = data.get("input", "").strip()
    duration = data.get("duration", 15)
    if not input_text:
        return {"error": "Manjka vnos."}

    mode = "url" if input_text.startswith("http") else "text"
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []
    # ~2.5 besede/sekundo za naravni govorni tempo, -1s buffer
    words = max(20, min(120, int((duration - 1) * 2.5)))

    prompt = f"""{'Preberi to stran in' if mode == 'url' else 'Na podlagi tega opisa'} ustvari voice over skripte za video oglas v 10 jezikih.

{'Stran: ' + input_text if mode == 'url' else 'Opis: ' + input_text}

Pravila:
- Točno {words} besed na jezik (±5)
- Naravni govorni slog, kot da govori prijatelj
- Poudarek na eni glavni koristi izdelka
- Brez cen, brez "klikni", brez "naroči"
- Konec z močno izjavo (ne pozivom k akciji)
- SL: slovenščina, HR: hrvaščina (latinica), RS: srbščina (SAMO latinica), HU: madžarščina, CZ: češčina, SK: slovaščina, PL: poljščina, GR: grščina (grška pisava), RO: romunščina, BG: bolgarščina (SAMO cirilica)

Vrni SAMO JSON brez markdown:
{{"product": "ime izdelka", "sl": "...", "hr": "...", "rs": "...", "hu": "...", "cz": "...", "sk": "...", "pl": "...", "gr": "...", "ro": "...", "bg": "..."}}"""

    try:
        text = await call_claude(prompt, "claude-sonnet-4-6", tools if tools else None, 10000)
        data_parsed = parse_json_response(text)
        if not data_parsed:
            return {"error": "Napaka pri generiranju skript."}
        scripts = {k: v for k, v in data_parsed.items() if k != "product"}
        return {"scripts": scripts, "product": data_parsed.get("product", "")}
    except Exception as e:
        return {"error": str(e)}


# ─── VIDEO ADS — ELEVENLABS AUDIO ────────────────────────────────────────────

ELEVENLABS_VOICES = {
    "sl": "pNInz6obpgDQGcFmaJgB",  # Adam — multilingual
    "hr": "pNInz6obpgDQGcFmaJgB",
    "rs": "pNInz6obpgDQGcFmaJgB",
    "hu": "pNInz6obpgDQGcFmaJgB",
    "cz": "pNInz6obpgDQGcFmaJgB",
    "sk": "pNInz6obpgDQGcFmaJgB",
    "pl": "pNInz6obpgDQGcFmaJgB",
    "gr": "pNInz6obpgDQGcFmaJgB",
    "ro": "pNInz6obpgDQGcFmaJgB",
    "bg": "pNInz6obpgDQGcFmaJgB",
}

def _parse_words(alignment: dict):
    """Iz ElevenLabs alignment podatkov izloci seznam (beseda, start, end)."""
    chars = alignment.get("characters", [])
    starts = alignment.get("character_start_times_seconds", [])
    ends = alignment.get("character_end_times_seconds", [])
    words = []
    cur_word, cur_start, cur_end = "", None, None
    for i, ch in enumerate(chars):
        if i >= len(starts):
            break
        t_start = starts[i]
        t_end = ends[i] if i < len(ends) else t_start + 0.1
        if ch in (' ', '\n', '\t'):
            if cur_word:
                words.append((cur_word, cur_start, cur_end))
            cur_word, cur_start, cur_end = "", None, None
        else:
            if cur_start is None:
                cur_start = t_start
            cur_end = t_end
            cur_word += ch
    if cur_word:
        words.append((cur_word, cur_start, cur_end))
    return words


def build_srt(alignment: dict) -> str:
    """Generira SRT iz ElevenLabs alignment (za download)."""
    words = _parse_words(alignment)
    if not words:
        return ""

    def fmt(s):
        h, m = int(s // 3600), int((s % 3600) // 60)
        sec, ms = int(s % 60), int((s - int(s)) * 1000)
        return f"{h:02d}:{m:02d}:{sec:02d},{ms:03d}"

    lines, i, idx = [], 0, 1
    while i < len(words):
        grp = [words[i]]
        i += 1
        while i < len(words) and len(grp) < 5 and (words[i][1] - grp[0][1]) < 3.0:
            grp.append(words[i]); i += 1
        lines.append(f"{idx}\n{fmt(grp[0][1])} --> {fmt(grp[-1][2])}\n{' '.join(w[0] for w in grp)}\n")
        idx += 1
    return "\n".join(lines)


def get_subtitle_style_for_format(width: int, height: int) -> dict:
    """Vrne optimalne subtitle nastavitve glede na video format."""
    if width == 0 or height == 0:
        # Default: assume 9:16
        return {"fontsize": 64, "marginv": 150, "outline": 5, "max_words": 4, "playresx": 1080, "playresy": 1920}

    ratio = width / height

    # MarginV kot % višine — podnapisi so vedno vidni ne glede na crop
    # Za pokončne videe damo podnapise višje (večji %) da ne padejo v safe-zone
    if ratio < 0.4:
        # Ultra ozki pokončni videi (6:19, 9:21 ipd.) — podnapisi zelo visoko
        # ker se spodnji del pogosto odreže pri predvajanju
        marginv = max(int(height * 0.35), 120)
        return {"fontsize": max(int(height * 0.038), 48), "marginv": marginv, "outline": 5, "max_words": 3, "playresx": width, "playresy": height}
    elif ratio < 0.7:
        # 9:16 vertical (TikTok, Reels, Stories) — 0.5625
        marginv = max(int(height * 0.08), 100)
        return {"fontsize": max(int(height * 0.038), 56), "marginv": marginv, "outline": 5, "max_words": 4, "playresx": width, "playresy": height}
    elif ratio < 1.2:
        # 1:1 square (Insta/FB feed)
        marginv = max(int(height * 0.07), 60)
        return {"fontsize": max(int(height * 0.045), 44), "marginv": marginv, "outline": 4, "max_words": 5, "playresx": width, "playresy": height}
    elif ratio < 1.6:
        # 4:3
        marginv = max(int(height * 0.07), 50)
        return {"fontsize": max(int(height * 0.05), 40), "marginv": marginv, "outline": 4, "max_words": 5, "playresx": width, "playresy": height}
    else:
        # 16:9 horizontal
        marginv = max(int(height * 0.06), 40)
        return {"fontsize": max(int(height * 0.055), 36), "marginv": marginv, "outline": 3, "max_words": 6, "playresx": width, "playresy": height}

def build_ass(alignment: dict, video_width: int = 1080, video_height: int = 1920) -> str:
    """Generira ASS karaoke podnapise — Stil B (bela+rumena, debela obroba), prilagojen formatu."""
    words = _parse_words(alignment)
    if not words:
        return ""

    style = get_subtitle_style_for_format(video_width, video_height)
    fontsize = style["fontsize"]
    marginv = style["marginv"]
    outline = style["outline"]
    max_words = style["max_words"]
    playresx = style["playresx"]
    playresy = style["playresy"]

    # ASS header — prilagojen video formatu
    header = f"""[Script Info]
ScriptType: v4.00+
PlayResX: {playresx}
PlayResY: {playresy}

[V4+ Styles]
Format: Name,Fontname,Fontsize,PrimaryColour,SecondaryColour,OutlineColour,BackColour,Bold,Italic,Underline,StrikeOut,ScaleX,ScaleY,Spacing,Angle,BorderStyle,Outline,Shadow,Alignment,MarginL,MarginR,MarginV,Encoding
Style: Default,Arial,{fontsize},&H00FFFFFF,&H00FFD600,&H00000000,&H00000000,-1,0,0,0,100,100,0,0,1,{outline},0,2,30,30,{marginv},1

[Events]
Format: Layer,Start,End,Style,Name,MarginL,MarginR,MarginV,Effect,Text
"""

    def fmt(s):
        h, m = int(s // 3600), int((s % 3600) // 60)
        sec, cs = int(s % 60), int((s - int(s)) * 100)
        return f"{h}:{m:02d}:{sec:02d}.{cs:02d}"

    # Grupiraj v vrstice (max max_words besed)
    lines_out = []
    i = 0
    while i < len(words):
        grp = [words[i]]; i += 1
        while i < len(words) and len(grp) < max_words and (words[i][1] - grp[0][1]) < 2.5:
            grp.append(words[i]); i += 1

        t_in = grp[0][1]
        t_out = grp[-1][2] + 0.05

        # Karaoke: vsaka beseda dobi {\kXX} tag (čas v centisekundah)
        parts = []
        for wi, (word, wstart, wend) in enumerate(grp):
            # Čas do naslednje besede ali konec
            if wi + 1 < len(grp):
                dur_cs = int((grp[wi+1][1] - wstart) * 100)
            else:
                dur_cs = int((wend - wstart) * 100) + 5
            dur_cs = max(dur_cs, 5)
            parts.append("{" + chr(92) + "k" + str(dur_cs) + "}" + word)

        karaoke_text = " ".join(parts)
        lines_out.append(f"Dialogue: 0,{fmt(t_in)},{fmt(t_out)},Default,,0,0,0,,{karaoke_text}")

    return header + "\n".join(lines_out)


@app.post("/generate-audio")
async def generate_audio(data: dict):
    text = data.get("text", "").strip()
    lang = data.get("lang", "sl")
    if not text:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Manjka tekst."}, status_code=400)

    api_key = os.environ.get("ELEVENLABS_API_KEY", "")
    if not api_key:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "ELEVENLABS_API_KEY ni nastavljen."}, status_code=400)

    voice_id = ELEVENLABS_VOICES.get(lang, "pNInz6obpgDQGcFmaJgB")

    try:
        async with httpx.AsyncClient(timeout=60.0) as hc:
            # Uporabi /with-timestamps endpoint za alignment podatke
            resp = await hc.post(
                f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}/with-timestamps",
                headers={
                    "xi-api-key": api_key,
                    "Content-Type": "application/json",
                },
                json={
                    "text": text,
                    "model_id": "eleven_multilingual_v2",
                    "voice_settings": {"stability": 0.5, "similarity_boost": 0.75}
                }
            )

        if resp.status_code != 200:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": f"ElevenLabs napaka {resp.status_code}: {resp.text[:300]}"}, status_code=400)

        result = resp.json()
        
        # Dekodira audio iz base64
        import base64
        audio_b64 = result.get("audio_base64", "")
        audio_bytes = base64.b64decode(audio_b64)
        
        # Generiraj SRT (za download) in ASS (za karaoke merge)
        alignment = result.get("alignment", {})
        srt = build_srt(alignment)
        ass = build_ass(alignment)

        return {
            "audio_base64": audio_b64,
            "srt": srt,
            "ass": ass,
            "lang": lang
        }

    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── MERGE VIDEO + AUDIO ──────────────────────────────────────────────────────

# ─── VIDEO SESSION CACHE — single upload za batch merge ───────────────────────

VIDEO_SESSION_DIR = Path("/tmp/video_sessions")
VIDEO_SESSION_DIR.mkdir(exist_ok=True, parents=True)

@app.post("/upload-video-session")
async def upload_video_session(video: UploadFile = File(...)):
    """Naloži video enkrat, vrni session_id za večkratno uporabo."""
    try:
        import uuid as _u
        session_id = _u.uuid4().hex[:16]
        session_path = VIDEO_SESSION_DIR / f"{session_id}.mp4"
        content_bytes = await video.read()
        session_path.write_bytes(content_bytes)
        # Avtomatsko počisti starejše od 30 min
        try:
            now = datetime.now().timestamp()
            for f in VIDEO_SESSION_DIR.glob("*.mp4"):
                if now - f.stat().st_mtime > 1800:
                    f.unlink()
        except: pass
        print(f"[video-session] uploaded {session_id} ({len(content_bytes)} bytes)")
        return {"session_id": session_id, "size": len(content_bytes)}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/cleanup-video-session")
async def cleanup_video_session(data: dict):
    """Počisti video session ko ni več potreben."""
    session_ids = data.get("session_ids", [])
    deleted = 0
    for sid in session_ids:
        try:
            p = VIDEO_SESSION_DIR / f"{sid}.mp4"
            if p.exists():
                p.unlink()
                deleted += 1
        except: pass
    return {"deleted": deleted}


@app.post("/merge-video-audio-session")
async def merge_video_audio_session(
    session_id: str = Form(...),
    audio: UploadFile = File(...),
    lang: str = Form("sl"),
    srt: UploadFile = File(None),
):
    """Merge z video iz cached session (faster — video že na strežniku)."""
    import subprocess, tempfile
    from fastapi.responses import JSONResponse as JR
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
    except Exception:
        pass

    session_video_path = VIDEO_SESSION_DIR / f"{session_id}.mp4"
    if not session_video_path.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": f"Video session {session_id} not found."}, status_code=404)

    try:
        with tempfile.TemporaryDirectory() as tmp:
            audio_path = f"{tmp}/audio.mp3"
            srt_path = f"{tmp}/subs.srt"
            ass_path = f"{tmp}/subs.ass"
            output_path = f"{tmp}/output_{lang}.mp4"
            video_path = str(session_video_path)

            with open(audio_path, "wb") as f:
                f.write(await audio.read())

            has_srt = False
            ass_content_orig = None
            if srt:
                srt_content = await srt.read()
                if srt_content.strip():
                    with open(srt_path, "wb") as f:
                        f.write(srt_content)
                    if srt_content.startswith(b'[Script Info]'):
                        ass_content_orig = srt_content
                        has_srt = 'ass'
                    else:
                        has_srt = 'srt'

            # Detect video dimensions
            video_width, video_height = 0, 0
            try:
                probe_cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0",
                             "-show_entries", "stream=width,height", "-of", "csv=p=0", video_path]
                probe_result = subprocess.run(probe_cmd, capture_output=True, timeout=10)
                if probe_result.returncode == 0:
                    dims = probe_result.stdout.decode().strip().split(',')
                    if len(dims) == 2:
                        video_width, video_height = int(dims[0]), int(dims[1])
            except: pass

            # Adapt ASS če imamo
            if has_srt == 'ass' and ass_content_orig and video_width > 0 and video_height > 0:
                style = get_subtitle_style_for_format(video_width, video_height)
                ass_text = ass_content_orig.decode('utf-8', errors='replace')
                ass_text = re.sub(r'PlayResX:\s*\d+', f'PlayResX: {style["playresx"]}', ass_text)
                ass_text = re.sub(r'PlayResY:\s*\d+', f'PlayResY: {style["playresy"]}', ass_text)
                new_style = f'Style: Default,Arial,{style["fontsize"]},&H00FFFFFF,&H00FFD600,&H00000000,&H00000000,-1,0,0,0,100,100,0,0,1,{style["outline"]},0,2,30,30,{style["marginv"]},1'
                ass_text = re.sub(r'Style:\s*Default,[^\n]+', new_style, ass_text)
                with open(ass_path, "wb") as f:
                    f.write(ass_text.encode('utf-8'))
            elif has_srt == 'ass' and ass_content_orig:
                with open(ass_path, "wb") as f:
                    f.write(ass_content_orig)

            if has_srt:
                sub_file = ass_path if has_srt == 'ass' else srt_path
                if has_srt == 'ass':
                    vf = f"ass={sub_file}"
                else:
                    s = get_subtitle_style_for_format(video_width, video_height)
                    vf = f"subtitles={sub_file}:force_style='FontName=Arial,FontSize={s['fontsize']},PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,Outline={s['outline']},Bold=1,Alignment=2,MarginV={s['marginv']}'"
                cmd = ["ffmpeg", "-y", "-i", video_path, "-i", audio_path, "-vf", vf,
                       "-map", "0:v:0", "-map", "1:a:0", "-c:v", "libx264", "-c:a", "aac",
                       "-shortest", output_path]
            else:
                cmd = ["ffmpeg", "-y", "-i", video_path, "-i", audio_path,
                       "-map", "0:v:0", "-map", "1:a:0", "-c:v", "copy", "-c:a", "aac",
                       "-shortest", output_path]

            result = subprocess.run(cmd, capture_output=True, timeout=180)
            if result.returncode != 0:
                err_msg = result.stderr.decode(errors='replace')[-400:]
                return JR({"error": f"FFmpeg napaka: {err_msg}"}, status_code=500)

            with open(output_path, "rb") as f:
                video_bytes = f.read()

            return StreamingResponse(
                iter([video_bytes]),
                media_type="video/mp4",
                headers={"Content-Disposition": f"attachment; filename=video_{lang}.mp4"}
            )
    except subprocess.TimeoutExpired:
        return JR({"error": "FFmpeg timeout."}, status_code=500)
    except Exception as e:
        return JR({"error": str(e)}, status_code=500)


# ─── MERGE VIDEO + AUDIO (full upload — fallback) ─────────────────────────────

@app.post("/merge-video-audio")
async def merge_video_audio(
    video: UploadFile = File(...),
    audio: UploadFile = File(...),
    lang: str = "sl",
    srt: UploadFile = File(None),
):
    """Spoji video + audio (+ opcijsko SRT podnapisi) z FFmpeg."""
    import subprocess, tempfile
    from fastapi.responses import JSONResponse as JR
    # Uporabi static-ffmpeg da dobimo ffmpeg binarko brez root
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
    except Exception:
        pass
    try:
        with tempfile.TemporaryDirectory() as tmp:
            video_path = f"{tmp}/input.mp4"
            audio_path = f"{tmp}/audio.mp3"
            srt_path = f"{tmp}/subs.srt"
            ass_path = f"{tmp}/subs.ass"
            output_path = f"{tmp}/output_{lang}.mp4"

            with open(video_path, "wb") as f:
                f.write(await video.read())
            with open(audio_path, "wb") as f:
                f.write(await audio.read())

            has_srt = False
            ass_content_orig = None
            if srt:
                srt_content = await srt.read()
                if srt_content.strip():
                    with open(srt_path, "wb") as f:
                        f.write(srt_content)
                    # Poskusi parsati kot ASS (začne z [Script Info])
                    if srt_content.startswith(b'[Script Info]'):
                        ass_content_orig = srt_content
                        has_srt = 'ass'
                    else:
                        has_srt = 'srt'

            # Detect video dimensions z ffprobe za prilagoditev podnapisov
            video_width, video_height = 0, 0
            try:
                probe_cmd = [
                    "ffprobe", "-v", "error",
                    "-select_streams", "v:0",
                    "-show_entries", "stream=width,height",
                    "-of", "csv=p=0",
                    video_path
                ]
                probe_result = subprocess.run(probe_cmd, capture_output=True, timeout=10)
                if probe_result.returncode == 0:
                    dims = probe_result.stdout.decode().strip().split(',')
                    if len(dims) == 2:
                        video_width = int(dims[0])
                        video_height = int(dims[1])
                        print(f"[merge] Video {video_width}x{video_height} ratio={video_width/video_height:.2f}")
            except Exception as e:
                print(f"[merge] ffprobe failed: {e}")

            # Če imamo ASS — adaptiraj nastavitve glede na format
            if has_srt == 'ass' and ass_content_orig and video_width > 0 and video_height > 0:
                style = get_subtitle_style_for_format(video_width, video_height)
                # Prepiši Style: Default vrstico v ASS s pravimi parametri za format
                ass_text = ass_content_orig.decode('utf-8', errors='replace')
                # Prepiši PlayResX/PlayResY
                ass_text = re.sub(r'PlayResX:\s*\d+', f'PlayResX: {style["playresx"]}', ass_text)
                ass_text = re.sub(r'PlayResY:\s*\d+', f'PlayResY: {style["playresy"]}', ass_text)
                # Prepiši Style: Default vrstico
                new_style = f'Style: Default,Arial,{style["fontsize"]},&H00FFFFFF,&H00FFD600,&H00000000,&H00000000,-1,0,0,0,100,100,0,0,1,{style["outline"]},0,2,30,30,{style["marginv"]},1'
                ass_text = re.sub(r'Style:\s*Default,[^\n]+', new_style, ass_text)
                with open(ass_path, "wb") as f:
                    f.write(ass_text.encode('utf-8'))
                print(f"[merge] Adapted ASS for {video_width}x{video_height}: fontsize={style['fontsize']}, marginv={style['marginv']}")
            elif has_srt == 'ass' and ass_content_orig:
                # Ni dimensions — uporabi originalni ASS
                with open(ass_path, "wb") as f:
                    f.write(ass_content_orig)

            if has_srt:
                # ASS karaoke ali SRT fallback
                sub_file = ass_path if has_srt == 'ass' else srt_path
                if has_srt == 'ass':
                    vf = f"ass={sub_file}"
                else:
                    # SRT fallback — prilagodi format glede na video
                    s = get_subtitle_style_for_format(video_width, video_height)
                    vf = f"subtitles={sub_file}:force_style='FontName=Arial,FontSize={s['fontsize']},PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,Outline={s['outline']},Bold=1,Alignment=2,MarginV={s['marginv']}'"
                cmd = [
                    "ffmpeg", "-y",
                    "-i", video_path,
                    "-i", audio_path,
                    "-vf", vf,
                    "-map", "0:v:0",
                    "-map", "1:a:0",
                    "-c:v", "libx264",
                    "-c:a", "aac",
                    "-shortest",
                    output_path
                ]
            else:
                # Samo audio zamenjava, video brez rekodiranja
                cmd = [
                    "ffmpeg", "-y",
                    "-i", video_path,
                    "-i", audio_path,
                    "-map", "0:v:0",
                    "-map", "1:a:0",
                    "-c:v", "copy",
                    "-c:a", "aac",
                    "-shortest",
                    output_path
                ]

            result = subprocess.run(cmd, capture_output=True, timeout=180)

            if result.returncode != 0:
                err_msg = result.stderr.decode(errors='replace')[-400:]
                print(f"[merge] FFmpeg error: {err_msg}")
                return JR({"error": f"FFmpeg napaka: {err_msg}"}, status_code=500)

            with open(output_path, "rb") as f:
                video_bytes = f.read()

            return StreamingResponse(
                iter([video_bytes]),
                media_type="video/mp4",
                headers={"Content-Disposition": f"attachment; filename=video_{lang}.mp4"}
            )
    except subprocess.TimeoutExpired:
        return JR({"error": "FFmpeg timeout."}, status_code=500)
    except Exception as e:
        return JR({"error": str(e)}, status_code=500)


# ─── VIDEO ADS HISTORY ────────────────────────────────────────────────────────

VADS_HISTORY_FILE = DATA_DIR / "vads_history.json"

def vads_cleanup_old():
    """Zbriše vnose starejše od 7 dni."""
    if not VADS_HISTORY_FILE.exists():
        return
    try:
        history = json.loads(VADS_HISTORY_FILE.read_text(encoding="utf-8"))
        cutoff = datetime.now() - timedelta(days=7)
        history = [h for h in history if datetime.strptime(h.get("date","1.1.2000 00:00"), "%d.%m.%Y %H:%M") > cutoff]
        VADS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
    except:
        pass

@app.get("/vads-history")
async def get_vads_history():
    vads_cleanup_old()
    if VADS_HISTORY_FILE.exists():
        try:
            return json.loads(VADS_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/vads-history")
async def save_vads_history(data: dict):
    try:
        history = []
        if VADS_HISTORY_FILE.exists():
            try:
                history = json.loads(VADS_HISTORY_FILE.read_text(encoding="utf-8"))
            except:
                history = []
        history.append({
            "input": data.get("input", ""),
            "product": data.get("product", ""),
            "scripts": data.get("scripts", {}),
            "date": data.get("date", "")
        })
        if len(history) > 50:
            history = history[-50:]
        VADS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}

@app.post("/vads-history-set")
async def set_vads_history(data: dict):
    try:
        history = data.get("history", [])
        VADS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}




# ─── ASANA ATTACH BINARY (za ZIP, video, audio) ───────────────────────────────

@app.post("/asana-attach-binary")
async def asana_attach_binary(
    task_id: str = Form(...),
    file: UploadFile = File(...),
    filename: str = Form(None)
):
    """Priloži binarni fajl (ZIP, MP4, MP3) na Asana task."""
    if not task_id:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Manjka task_id."}, status_code=400)

    token = os.environ.get("ASANA_API_KEY", "")
    if not token:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "ASANA_API_KEY ni nastavljen."}, status_code=400)

    try:
        content = await file.read()
        upload_filename = filename or file.filename or "attachment"
        mime = file.content_type or "application/octet-stream"

        async with httpx.AsyncClient(timeout=120.0) as hc:
            files = {"file": (upload_filename, content, mime)}
            attach_resp = await hc.post(
                f"{ASANA_API}/tasks/{task_id}/attachments",
                headers={"Authorization": f"Bearer {token}"},
                files=files
            )

        if attach_resp.status_code in (200, 201):
            return {"status": "ok", "filename": upload_filename, "size": len(content)}
        else:
            from fastapi.responses import JSONResponse
            return JSONResponse(
                {"error": f"Asana napaka {attach_resp.status_code}: {attach_resp.text[:200]}"},
                status_code=400
            )
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── ORODJA: Združevalnik SKU+količin ─────────────────────────────────────────

ORODJA_HISTORY_DIR = DATA_DIR / "orodja_history"
ORODJA_HISTORY_DIR.mkdir(exist_ok=True, parents=True)

ORODJA_HS_HISTORY_DIR = DATA_DIR / "orodja_hs_history"
ORODJA_HS_HISTORY_DIR.mkdir(exist_ok=True, parents=True)


def cleanup_orodja_history():
    """Briše datoteke starejše od 30 dni (CSV združevalnik)."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in ORODJA_HISTORY_DIR.glob("*.xlsx"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[orodja] cleanup err: {e}")


def cleanup_orodja_hs_history():
    """Briše datoteke starejše od 90 dni (HS+ uvoz)."""
    try:
        cutoff = datetime.now().timestamp() - (90 * 86400)
        for f in ORODJA_HS_HISTORY_DIR.glob("*.xlsx"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[orodja-hs] cleanup err: {e}")


@app.post("/orodja-merge-skus")
async def orodja_merge_skus(file: UploadFile = File(...)):
    """Sprejme CSV z SKU+Količina, združi dvojnike, vrne XLSX."""
    try:
        content_bytes = await file.read()
        text = content_bytes.decode('utf-8-sig', errors='replace')

        # Parsanje CSV
        import csv
        from io import StringIO
        reader = csv.reader(StringIO(text))
        headers = next(reader, None)
        if not headers:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Najdi indekse SKU in Količina kolon
        h_lower = [h.lower().strip() for h in headers]
        try:
            sku_idx = next(i for i, h in enumerate(h_lower) if h in ('sku', 'sku.'))
        except StopIteration:
            sku_idx = 1  # default druga kolona
        try:
            qty_idx = next(i for i, h in enumerate(h_lower) if 'količin' in h or 'kolicin' in h or h == 'qty' or h == 'quantity')
        except StopIteration:
            qty_idx = 3  # default

        # Združi po SKU
        sku_totals = {}
        for row in reader:
            if len(row) <= max(sku_idx, qty_idx):
                continue
            sku = (row[sku_idx] or '').strip()
            if not sku:
                continue
            try:
                qty = int(float((row[qty_idx] or '0').strip().replace(',', '.')))
            except:
                qty = 0
            sku_totals[sku] = sku_totals.get(sku, 0) + qty

        if not sku_totals:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Ne najdem SKU/količina kolon v CSV."}, status_code=400)

        # Generiraj XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Order"
        # Brez headerjev — samo SKU | količina
        for sku, qty in sorted(sku_totals.items()):
            ws.append([sku, qty])

        # Shrani v history
        cleanup_orodja_history()  # počisti stare najprej
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        out_filename = f"Order_{ts}.xlsx"
        out_path = ORODJA_HISTORY_DIR / out_filename
        wb.save(out_path)

        # Vrni datoteko
        from fastapi.responses import FileResponse
        return FileResponse(
            path=str(out_path),
            filename=out_filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"X-Skus-Total": str(len(sku_totals)), "X-Filename": out_filename}
        )
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/orodja-history")
async def orodja_history():
    """Vrne seznam datotek v orodja_history (urejen po datumu od najnovejše)."""
    cleanup_orodja_history()
    items = []
    try:
        for f in sorted(ORODJA_HISTORY_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = f.stat()
            items.append({
                "filename": f.name,
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    except Exception as e:
        print(f"[orodja] history err: {e}")
    return {"items": items[:100]}


@app.get("/orodja-download/{filename}")
async def orodja_download(filename: str):
    """Prenesi XLSX iz history."""
    # Sanitize filename — preprečimo path traversal
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HISTORY_DIR / filename
    if not f.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)

    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(f),
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.delete("/orodja-history/{filename}")
async def orodja_history_delete(filename: str):
    """Zbriši posamezno datoteko iz history."""
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HISTORY_DIR / filename
    if f.exists():
        try:
            f.unlink()
            return {"status": "ok"}
        except Exception as e:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": str(e)}, status_code=500)
    return {"status": "not_found"}


# ─── ORODJA: Uvoz HS+ PDF predračun ───────────────────────────────────────────

@app.post("/orodja-import-hs-pdf")
async def orodja_import_hs_pdf(file: UploadFile = File(...)):
    """Sprejme HS+ PDF predračun, vrne JSON s SKU + količinami.
    Uporablja Claude Vision za branje image-based PDF."""
    try:
        content_bytes = await file.read()
        items = []

        # Najprej poskus tekst-extraction (če je tekstovni PDF)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name

        try:
            try:
                with pdfplumber.open(tmp_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        if text.strip():
                            for line in text.split('\n'):
                                line = line.strip()
                                m = re.match(r'^(\d{12,14})\s+(.+?)\s+(\d+)\s+(?:KOS|kos|PCS|pcs)', line)
                                if m:
                                    opis = m.group(2).strip()
                                    tokens = [t.rstrip('.,;:') for t in opis.split()]
                                    upper_t = [t for t in tokens if t.isupper() and len(t) >= 3 and not t.isdigit()]
                                    sku = upper_t[-1] if upper_t else (tokens[-1] if tokens else opis)
                                    items.append({
                                        "ean": m.group(1),
                                        "opis": opis,
                                        "sku": sku,
                                        "kolicina": int(m.group(3)),
                                    })
            except: pass

            # Fallback: Claude Vision (PDF kot dokument)
            if not items:
                import base64
                pdf_b64 = base64.b64encode(content_bytes).decode('utf-8')

                prompt = """Preberi ta predračun in vrni VSA postavke v JSON formatu.
Za vsako postavko izloci:
- ean: 13-mestna številčna koda na začetku vrstice
- opis: celoten opis postavke
- sku: zadnja SVE-VELIKA-ČRKA beseda v opisu (npr. "HYDRASPRINK HYDRASPRINK" → SKU = "HYDRASPRINK"; "WHEELPLAY yellow WHEELPLAY" → SKU = "WHEELPLAY"; "TOPKNER 180x200 TOPKNER" → SKU = "TOPKNER")
- kolicina: število pred "KOS" oznako

Vrni IZKLJUČNO valid JSON v formatu:
{"items": [{"ean": "...", "opis": "...", "sku": "...", "kolicina": 350}, ...]}

Brez dodatnih komentarjev, samo JSON."""

                try:
                    client = anthropic.Anthropic()
                    response = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=8000,
                        messages=[{
                            "role": "user",
                            "content": [
                                {
                                    "type": "document",
                                    "source": {
                                        "type": "base64",
                                        "media_type": "application/pdf",
                                        "data": pdf_b64
                                    }
                                },
                                {"type": "text", "text": prompt}
                            ]
                        }]
                    )
                    text = "".join([b.text for b in response.content if hasattr(b, 'text')])
                    parsed = parse_json_response(text)
                    if parsed and 'items' in parsed:
                        for it in parsed['items']:
                            try:
                                qty = int(it.get('kolicina', 0))
                            except:
                                qty = 0
                            items.append({
                                "ean": str(it.get('ean', '')),
                                "opis": str(it.get('opis', '')),
                                "sku": str(it.get('sku', '')).strip(),
                                "kolicina": qty,
                            })
                except Exception as e:
                    print(f"[hs-pdf] Claude vision error: {e}")
        finally:
            try:
                os.unlink(tmp_path)
            except: pass

        if not items:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Iz PDF-ja ne morem prebrati postavk."}, status_code=400)

        # Skupna količina iz PDF-ja
        pdf_total = sum(int(it.get('kolicina', 0) or 0) for it in items)
        return {"items": items, "total": len(items), "pdf_total": pdf_total}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/orodja-export-hs-xlsx")
async def orodja_export_hs_xlsx(data: dict):
    """Sprejme [{sku, kolicina}], generira XLSX in shrani v history."""
    try:
        items = data.get("items", [])
        if not items:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Ni postavk."}, status_code=400)

        # Generiraj XLSX z headerji sku|stock (HS+ format)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime('%Y-%m-%d_%H-%M-%S-export')
        ws.append(["sku", "stock"])  # header
        for item in items:
            sku = (item.get("sku") or "").strip()
            qty = item.get("kolicina") or 0
            if not sku:
                continue
            try:
                qty_int = int(float(qty))
            except:
                qty_int = 0
            ws.append([sku, qty_int])

        cleanup_orodja_hs_history()
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        out_filename = f"HS_PLUS_{ts}.xlsx"
        out_path = ORODJA_HS_HISTORY_DIR / out_filename
        wb.save(out_path)

        from fastapi.responses import FileResponse
        return FileResponse(
            path=str(out_path),
            filename=out_filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"X-Filename": out_filename}
        )
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)



# ─── HS+ HISTORY ENDPOINTS ────────────────────────────────────────────────────

@app.get("/orodja-hs-history")
async def orodja_hs_history():
    """Vrne seznam HS+ datotek (90 dni)."""
    cleanup_orodja_hs_history()
    items = []
    try:
        for f in sorted(ORODJA_HS_HISTORY_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = f.stat()
            items.append({
                "filename": f.name,
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    except Exception as e:
        print(f"[orodja-hs] history err: {e}")
    return {"items": items[:200]}


@app.get("/orodja-hs-download/{filename}")
async def orodja_hs_download(filename: str):
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HS_HISTORY_DIR / filename
    if not f.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)

    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(f),
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.delete("/orodja-hs-history/{filename}")
async def orodja_hs_history_delete(filename: str):
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HS_HISTORY_DIR / filename
    if f.exists():
        try:
            f.unlink()
            return {"status": "ok"}
        except Exception as e:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": str(e)}, status_code=500)
    return {"status": "not_found"}



# ─── ORODJA: Kontrola cen — Stock CSV upload + match s PDF predračun ────────

STOCK_CSV_FILE = DATA_DIR / "stock_inventory.csv"
STOCK_CSV_META = DATA_DIR / "stock_inventory_meta.json"


@app.post("/orodja-stock-upload")
async def orodja_stock_upload(file: UploadFile = File(...)):
    """Naloži CSV zaloge — merge po SKU (sešteje stock iz več skladišč/uploadov)."""
    try:
        import csv as _csv
        from io import StringIO as _SIO

        content_bytes = await file.read()
        text = content_bytes.decode('utf-8-sig', errors='replace')
        sep = ';' if (text.split('\n', 1)[0].count(';') > text.split('\n', 1)[0].count(',')) else ','

        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        new_rows = [r for r in reader]
        if not new_rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Normalizacija: najdi SKU, stock, stock30, title stolpce
        sample = new_rows[0]
        keys = list(sample.keys())

        def find_col(*candidates):
            for c in candidates:
                for k in keys:
                    if k.strip().lower() == c.lower():
                        return k
            return None

        sku_col   = find_col('product_sku', 'sku')
        stock_col = find_col('stock', 'qty', 'quantity', 'kolicina', 'količina')
        s30_col   = find_col('stock30', 'stock_30', 'obrat30', 'obrat_30')
        title_col = find_col('title', 'naziv', 'name')
        price_col = find_col('price_netto', 'price', 'cena')
        pos_col   = find_col('position', 'pozicija', 'lokacija')
        note_col  = find_col('note', 'opomba', 'komentar')

        if not sku_col:
            return JSONResponse({"error": f"Ne najdem SKU stolpca. Najdeni: {keys}"}, status_code=400)

        # Merge znotraj CSV-ja — seštej stock za isti SKU (več lokacij)
        merged = {}
        added = 0
        updated = 0
        for row in new_rows:
            if not sku_col: continue
            sku = (row.get(sku_col) or '').strip()
            if not sku: continue
            try:
                new_stock = int(float((row.get(stock_col) or '0').replace(',', '.'))) if stock_col else 0
            except: new_stock = 0
            try:
                new_s30 = int(float((row.get(s30_col) or '0').replace(',', '.'))) if s30_col else 0
            except: new_s30 = 0
            title = (row.get(title_col) or '').strip() if title_col else ''

            if sku in merged:
                # Isti SKU, druga lokacija — seštej stock
                merged[sku]['stock'] += new_stock
                if new_s30 > merged[sku]['stock30']:
                    merged[sku]['stock30'] = new_s30
                if title and not merged[sku]['title']:
                    merged[sku]['title'] = title
                # Ohrani price/position/note iz prve lokacije (ali posodobi če prazno)
                for fld, src_col in [('price', price_col), ('position', pos_col), ('note', note_col), ('product_id', 'product_id')]:
                    new_val = (row.get(src_col) or '').strip() if src_col else ''
                    if new_val and not merged[sku].get(fld):
                        merged[sku][fld] = new_val
                updated += 1
            else:
                merged[sku] = {
                    'product_id': (row.get('product_id') or '').strip(),
                    'product_sku': sku,
                    'title': title,
                    'stock': new_stock,
                    'stock30': new_s30,
                    'price': (row.get(price_col) or '').strip() if price_col else '',
                    'position': (row.get(pos_col) or '').strip() if pos_col else '',
                    'note': (row.get(note_col) or '').strip() if note_col else '',
                }
                added += 1

        # Shrani nazaj kot CSV z vsemi kolonami
        out = _SIO()
        fieldnames = ['product_id', 'product_sku', 'title', 'stock', 'stock30', 'price', 'position', 'note']
        writer = _csv.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(merged.values())
        STOCK_CSV_FILE.write_text(out.getvalue(), encoding='utf-8')

        total = len(merged)
        meta = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "filename": file.filename,
            "rows": total,
            "rows_added": added,
            "rows_merged": updated,
        }
        # Ohrani seznam uploadov
        old_meta = {}
        if STOCK_CSV_META.exists():
            try: old_meta = json.loads(STOCK_CSV_META.read_text(encoding='utf-8'))
            except: pass
        uploads = old_meta.get('uploads', [])
        uploads.append({"filename": file.filename, "uploaded_at": meta["uploaded_at"], "added": added, "merged": updated})
        meta['uploads'] = uploads[-5:]  # zadnjih 5
        STOCK_CSV_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

        return {
            "status": "ok",
            "rows": total,
            "rows_added": added,
            "rows_merged": updated,
            "uploaded_at": meta["uploaded_at"],
            "filename": file.filename,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)



@app.get("/orodja-stock-status")
async def orodja_stock_status():
    """Vrne info o trenutno shranjeni zalogi."""
    if not STOCK_CSV_FILE.exists() or not STOCK_CSV_META.exists():
        return {"loaded": False}
    try:
        meta = json.loads(STOCK_CSV_META.read_text(encoding="utf-8"))
        return {"loaded": True, **meta}
    except Exception as e:
        return {"loaded": False, "error": str(e)}


@app.post("/orodja-stock-clear")
async def orodja_stock_clear():
    """Počisti zalogo (reset za nov upload)."""
    try:
        if STOCK_CSV_FILE.exists(): STOCK_CSV_FILE.unlink()
        if STOCK_CSV_META.exists(): STOCK_CSV_META.unlink()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/orodja-price-check")
async def orodja_price_check(file: UploadFile = File(...)):
    """Sprejme HS+ PDF + match s shranjeno zalogo, vrne primerjavo cen."""
    if not STOCK_CSV_FILE.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Najprej naloži CSV zaloge."}, status_code=400)

    try:
        # 1. Preberi PDF s Claude Vision (isti pristop kot orodja-import-hs-pdf)
        content_bytes = await file.read()
        items = []  # [{ean, opis, sku, kolicina, cena_pdf, popust_pct}]
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name

        try:
            # Najprej poskus pdfplumber
            try:
                with pdfplumber.open(tmp_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        if text.strip():
                            for line in text.split('\n'):
                                line = line.strip()
                                # Format: ean opis kolicina KOS cena popust DDV znesek
                                m = re.match(r'^(\d{12,14})\s+(.+?)\s+(\d+)\s+(?:KOS|kos)\s+([\d.,]+)\s+(\d+)?', line)
                                if m:
                                    opis = m.group(2).strip()
                                    tokens = [t.rstrip('.,;:') for t in opis.split()]
                                    upper_t = [t for t in tokens if t.isupper() and len(t) >= 3 and not t.isdigit()]
                                    sku = upper_t[-1] if upper_t else (tokens[-1] if tokens else opis)
                                    try:
                                        cena = float(m.group(4).replace(',', '.'))
                                    except:
                                        cena = 0
                                    try:
                                        popust = float(m.group(5)) if m.group(5) else 0
                                    except:
                                        popust = 0
                                    items.append({
                                        "ean": m.group(1),
                                        "opis": opis,
                                        "sku": sku,
                                        "kolicina": int(m.group(3)),
                                        "cena_pdf": cena,
                                        "popust_pct": popust,
                                    })
            except: pass

            # Fallback: Claude Vision
            if not items:
                import base64
                pdf_b64 = base64.b64encode(content_bytes).decode('utf-8')
                prompt = """Preberi ta predračun in vrni VSA postavke v JSON formatu.
Za vsako postavko izloci:
- ean: 13-mestna številka koda na začetku vrstice
- opis: celoten opis postavke
- sku: zadnja SVE-VELIKA-ČRKA beseda v opisu (npr. "HYDRASPRINK HYDRASPRINK" → SKU = "HYDRASPRINK"; "WHEELPLAY yellow WHEELPLAY" → SKU = "WHEELPLAY"; "TOPKNER 180x200 TOPKNER" → SKU = "TOPKNER_180x200")
- kolicina: število pred "KOS" oznako
- cena_pdf: cena enote (po KOS, npr. "2,67")
- popust_pct: popust v % (število iz Popust % stolpca, lahko je prazno → 0)

POMEMBNO: Pri SKU-jih z dimenzijami (TOPKNER, WHEELPLAY) vključi tudi dimenzijo/barvo, ker so to različni izdelki:
- "TOPKNER 180x200 TOPKNER" → "TOPKNER_180x200"
- "TOPKNER 160x200 TOPKNER" → "TOPKNER_160x200"
- "WHEELPLAY yellow WHEELPLAY" → "WHEELPLAY_yellow"
- "PLANTUP (white) PLANTUP" → "PLANTUP_white"
- "BEEWAX WOOD POLISH BEEWAX" → "BEEWAX"

Vrni IZKLJUČNO valid JSON v formatu:
{"items": [{"ean": "...", "opis": "...", "sku": "...", "kolicina": 350, "cena_pdf": 2.67, "popust_pct": 5}, ...]}

Brez dodatnih komentarjev, samo JSON."""

                try:
                    client = anthropic.Anthropic()
                    response = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=8000,
                        messages=[{
                            "role": "user",
                            "content": [
                                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                                {"type": "text", "text": prompt}
                            ]
                        }]
                    )
                    text_resp = "".join([b.text for b in response.content if hasattr(b, 'text')])
                    parsed = parse_json_response(text_resp)
                    if parsed and 'items' in parsed:
                        for it in parsed['items']:
                            try: qty = int(it.get('kolicina', 0))
                            except: qty = 0
                            try: cena = float(str(it.get('cena_pdf', 0)).replace(',', '.'))
                            except: cena = 0
                            try: popust = float(str(it.get('popust_pct', 0)).replace(',', '.'))
                            except: popust = 0
                            items.append({
                                "ean": str(it.get('ean', '')),
                                "opis": str(it.get('opis', '')),
                                "sku": str(it.get('sku', '')).strip(),
                                "kolicina": qty,
                                "cena_pdf": cena,
                                "popust_pct": popust,
                            })
                except Exception as e:
                    print(f"[price-check] Claude error: {e}")
        finally:
            try: os.unlink(tmp_path)
            except: pass

        if not items:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "PDF parse fail."}, status_code=400)

        # 2. Naloži zalogo iz CSV
        import csv
        from io import StringIO
        stock_text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        # Detect separator
        first_line = stock_text.split('\n')[0]
        delim = ';' if first_line.count(';') > first_line.count(',') else ','
        reader = csv.DictReader(StringIO(stock_text), delimiter=delim)

        # Map SKU → cena (lower-case za case-insensitive match)
        stock_map = {}
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if not sku:
                continue
            try:
                cena = float(str(row.get('price_netto') or row.get('price') or 0).replace(',', '.'))
            except:
                cena = 0
            title = (row.get('title') or '').strip()
            stock_map[sku.lower()] = {"sku": sku, "cena_zaloga": cena, "title": title}

        # Fuzzy match helper
        def find_stock_match(sku_pdf):
            """Najprej exact match, nato koren match (PLANTUP_white → PLANTUP), nato fuzzy."""
            sku_lower = sku_pdf.lower()
            # 1. Exact match
            if sku_lower in stock_map:
                return stock_map[sku_lower]

            # 2. Koren match — vzami osnovno besedo (PLANTUP_white → PLANTUP, COVERKA_2x3m → COVERKA)
            koren = re.split(r'[_\-\s]', sku_lower)[0]
            if not koren or len(koren) < 3:
                return None

            # Najdi vse SKU-je v zalogi ki začnejo s tem korenom
            candidates = [(k, v) for k, v in stock_map.items() if k.split('_')[0] == koren or k.split('-')[0] == koren or k == koren]

            if not candidates:
                return None

            # Če je samo eden, vrni ga
            if len(candidates) == 1:
                return candidates[0][1]

            # 3. Fuzzy match — pri več zadetkih izberi najbolj podobnega
            from difflib import SequenceMatcher
            best = None
            best_ratio = 0
            for k, v in candidates:
                ratio = SequenceMatcher(None, sku_lower, k).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best = v
            # Vrni samo če je dovolj podoben (>0.6)
            return best if best_ratio >= 0.6 else None

        # 3. Match in primerjava
        results = []
        for item in items:
            sku_pdf = item["sku"]
            stock = find_stock_match(sku_pdf)

            cena_pdf_neto = item["cena_pdf"] * (1 - item["popust_pct"] / 100)

            if stock:
                cena_zaloga = stock["cena_zaloga"]
                razlika = cena_pdf_neto - cena_zaloga
                razlika_pct = (razlika / cena_zaloga * 100) if cena_zaloga > 0 else 0
                if abs(razlika) < 0.001:
                    status = "match"
                elif razlika > 0:
                    status = "vecja"  # PDF cena VEČJA = SLABO za nas
                else:
                    status = "manjsa"  # PDF cena MANJŠA = DOBRO za nas
            else:
                cena_zaloga = None
                razlika = None
                razlika_pct = None
                status = "no_match"

            results.append({
                "sku": sku_pdf,
                "opis": item["opis"],
                "title_zaloga": stock["title"] if stock else None,
                "kolicina": item["kolicina"],
                "cena_pdf": item["cena_pdf"],
                "popust_pct": item["popust_pct"],
                "cena_pdf_neto": round(cena_pdf_neto, 4),
                "cena_zaloga": cena_zaloga,
                "razlika": round(razlika, 4) if razlika is not None else None,
                "razlika_pct": round(razlika_pct, 2) if razlika_pct is not None else None,
                "status": status,
            })

        return {"items": results, "total": len(results), "matched": sum(1 for r in results if r["status"] != "no_match")}
    except Exception as e:
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/orodja-stock-data")
async def orodja_stock_data():
    """Vrne celoten seznam zaloge iz shranjenega CSV."""
    if not STOCK_CSV_FILE.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Najprej naloži CSV zaloge."}, status_code=400)
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','

        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        items = []
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if not sku:
                continue
            items.append({
                "sku": sku,
                "product_sku": sku,
                "product_id": (row.get('product_id') or '').strip(),
                "title": (row.get('title') or '').strip(),
                "stock": (row.get('stock') or '0').strip(),
                "stock30": (row.get('stock30') or '0').strip(),
                "price": (row.get('price_netto') or row.get('price') or '0').strip(),
                "position": (row.get('position') or '').strip(),
                "note": (row.get('note') or '').strip(),
            })

        meta = {}
        if STOCK_CSV_META.exists():
            try:
                meta = json.loads(STOCK_CSV_META.read_text(encoding="utf-8"))
            except: pass

        return {"items": items, "total": len(items), **meta}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── ANALIZA: TikTok Ads upload ──────────────────────────────────────────────

TIKTOK_CREATIVE_FILE = DATA_DIR / "tiktok_creative_map.json"

@app.post("/tiktok-creative-upload")
async def tiktok_creative_upload(file: UploadFile = File(...)):
    """Naloži TikTok Ad level export — zgradi Video→SKU mapping."""
    try:
        import io, re as _re
        content = await file.read()
        fname = file.filename or ""

        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, r)) for r in rows_raw[1:] if any(v is not None for v in r)]
        else:
            import csv as _csv
            text = content.decode('utf-8-sig', errors='replace')
            sep = ';' if text.split('\n')[0].count(';') > text.split('\n')[0].count(',') else ','
            data_rows = list(_csv.DictReader(io.StringIO(text), delimiter=sep))

        def extract_sku(campaign):
            c = str(campaign or '')
            m = _re.search(r'SKU:\s*([A-Za-z0-9_]+)', c)
            if m: return smart_root(m.group(1)).upper()
            m = _re.search(r'Smart\+\s+([A-Za-z0-9_]+)', c, _re.I)
            if m: return smart_root(m.group(1)).upper()
            return ''

        # Zgradi mapping: {sku: [{video, cost, conversions, status}]}
        sku_map = {}
        skipped = 0
        for row in data_rows:
            video = str(row.get('Video') or '').strip()
            if not video or video == '-': skipped += 1; continue
            campaign = row.get('Campaign name') or row.get('Campaign Name') or ''
            sku = extract_sku(campaign)
            if not sku: skipped += 1; continue

            try: cost = float(str(row.get('Cost') or 0).replace(',', '.'))
            except: cost = 0
            try: conversions = int(float(str(row.get('Conversions') or 0).replace(',', '.')))
            except: conversions = 0
            status = str(row.get('Primary status') or '').strip().lower()

            if sku not in sku_map:
                sku_map[sku] = {}
            if video not in sku_map[sku]:
                sku_map[sku][video] = {'video': video, 'cost': 0, 'conversions': 0, 'status': status}
            sku_map[sku][video]['cost'] += cost
            sku_map[sku][video]['conversions'] += conversions
            if status == 'active': sku_map[sku][video]['status'] = 'active'

        # Pretvori v seznam in sortiraj po cost desc
        result = {}
        for sku, videos in sku_map.items():
            result[sku] = sorted(videos.values(), key=lambda x: x['cost'], reverse=True)

        TIKTOK_CREATIVE_FILE.write_text(
            json.dumps({'map': result, 'uploaded_at': __import__('datetime').datetime.now().isoformat(), 'filename': fname, 'skus': len(result)}, ensure_ascii=False),
            encoding='utf-8'
        )
        return {"ok": True, "skus": len(result), "videos": sum(len(v) for v in result.values()), "skipped": skipped}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/tiktok-creative-search")
async def tiktok_creative_search(sku: str = ""):
    """Poišči videe za SKU."""
    if not TIKTOK_CREATIVE_FILE.exists():
        return JSONResponse({"error": "Naloži Creative CSV najprej."}, status_code=400)
    try:
        data = json.loads(TIKTOK_CREATIVE_FILE.read_text(encoding='utf-8'))
        cmap = data.get('map', {})
        sku_up = sku.upper().strip()
        if not sku_up:
            return {"skus": list(cmap.keys()), "videos": []}
        # Išči po SKU ali korenu
        root = smart_root(sku_up)
        videos = cmap.get(sku_up) or cmap.get(root) or []
        # Fuzzy: če ni exact match, vrni SKU-je ki vsebujejo iskalni niz
        if not videos:
            matches = {k: v for k, v in cmap.items() if sku_up in k or root in k}
            return {"skus": list(matches.keys()), "videos": [], "fuzzy": True}
        return {"sku": sku_up, "videos": videos, "total": len(videos)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/tiktok-creative-info")
async def tiktok_creative_info():
    if not TIKTOK_CREATIVE_FILE.exists():
        return {"loaded": False}
    try:
        data = json.loads(TIKTOK_CREATIVE_FILE.read_text(encoding='utf-8'))
        return {"loaded": True, "skus": data.get('skus', 0), "uploaded_at": data.get('uploaded_at'), "filename": data.get('filename')}
    except:
        return {"loaded": False}


TIKTOK_ADS_FILE = DATA_DIR / "tiktok_ads_report.csv"
TIKTOK_ADS_META = DATA_DIR / "tiktok_ads_meta.json"

@app.post("/analiza-tiktok-upload")
async def analiza_tiktok_upload(file: UploadFile = File(...)):
    """Naloži TikTok XLSX/CSV Campaign Report."""
    try:
        content = await file.read()
        fname = file.filename or ""

        # Preberi XLSX ali CSV
        import io
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw: return JSONResponse({"error": "Prazna datoteka."}, status_code=400)
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(r is not None for r in row)]
        else:
            import csv as _csv
            text = content.decode('utf-8-sig', errors='replace')
            sep = ';' if text.split('\n')[0].count(';') > text.split('\n')[0].count(',') else ','
            reader = _csv.DictReader(io.StringIO(text), delimiter=sep)
            data_rows = list(reader)

        # Kolone za TikTok
        def fcol(row, *candidates):
            for c in candidates:
                for k in row.keys():
                    if k.strip().lower() == c.lower():
                        return row[k]
            return None

        parsed = []
        for row in data_rows:
            campaign = str(fcol(row, 'Campaign name', 'campaign_name') or '').strip()
            if not campaign: continue
            status = str(fcol(row, 'Primary status', 'Status', 'status') or '').strip().lower()
            spend_raw = fcol(row, 'Cost', 'Spend', 'cost', 'spend')
            conv_raw = fcol(row, 'Conversions', 'conversions', 'Cost per conversion')
            cpc_raw = fcol(row, 'CPC (destination)', 'CPC')

            try: spend = float(str(spend_raw).replace(',', '.')) if spend_raw else 0
            except: spend = 0
            try: conversions = float(str(conv_raw).replace(',', '.')) if conv_raw else 0
            except: conversions = 0

            # Izvleci SKU iz campaign name — vzorec: [Maaarket] Smart+ SKU ali SKU: ABPULLER kjerkoli
            import re
            sku = ''
            m = re.search(r'Smart\+\s+([A-Z0-9_]+)', campaign, re.IGNORECASE)
            if m: sku = smart_root(m.group(1).strip())
            else:
                m = re.search(r'SKU:\s*([A-Z0-9_]+)', campaign, re.IGNORECASE)
                if m: sku = smart_root(m.group(1).strip())
            if not sku:
                # Splošno: vzemi zadnji ALL-CAPS blok
                words = campaign.split()
                for w in words:
                    cleaned = re.sub(r'[^A-Z0-9_]', '', w.upper())
                    if len(cleaned) >= 3 and cleaned not in {'ALL','SMART','ADS','TT','TIKTOK','NEW','OLD'}:
                        sku = smart_root(cleaned)
                        break

            parsed.append({
                'campaign': campaign, 'sku': sku.upper(),
                'status': 'active' if 'active' in status else 'inactive',
                'spend': spend, 'conversions': conversions,
                'currency': str(fcol(row, 'Currency', 'currency') or 'EUR').strip(),
                'create_time': str(fcol(row, 'Date Created', 'Campaign create time', 'Create time', 'create_time', 'Created time', 'Ad group create time') or '').strip(),
            })

        if not parsed:
            return JSONResponse({"error": "Ni veljavnih kampanj v datoteki."}, status_code=400)

        # Shrani — accumulate (merge po campaign imenu)
        import csv as _csv2
        existing = {}
        if TIKTOK_ADS_FILE.exists():
            t = TIKTOK_ADS_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for row in _csv2.DictReader(io.StringIO(t)):
                existing[row['campaign']] = row

        for row in parsed:
            existing[row['campaign']] = row

        out = io.StringIO()
        fieldnames = ['campaign', 'sku', 'status', 'spend', 'conversions', 'currency', 'create_time']
        writer = _csv2.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(existing.values())
        TIKTOK_ADS_FILE.write_text(out.getvalue(), encoding='utf-8')

        # Meta
        meta = {}
        if TIKTOK_ADS_META.exists():
            try: meta = json.loads(TIKTOK_ADS_META.read_text(encoding='utf-8'))
            except: pass
        uploads = meta.get('uploads', [])
        uploads.append({'filename': fname, 'rows': len(parsed), 'uploaded_at': __import__('datetime').datetime.now().isoformat()})
        TIKTOK_ADS_META.write_text(json.dumps({'uploads': uploads}, ensure_ascii=False), encoding='utf-8')

        return {"ok": True, "rows": len(existing), "new": len(parsed)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/analiza-tiktok-data")
async def analiza_tiktok_data():
    """Vrne TikTok kampanje z zalogo."""
    if not TIKTOK_ADS_FILE.exists():
        return JSONResponse({"error": "Naloži TikTok poročilo."}, status_code=400)
    try:
        import csv as _csv3, io as _io3
        text = TIKTOK_ADS_FILE.read_text(encoding='utf-8-sig', errors='replace')
        rows = list(_csv3.DictReader(_io3.StringIO(text)))

        # Dodaj zalogo iz stock CSV
        stock_map = {}
        stock_root_map = {}  # root → seznam variant
        if STOCK_CSV_FILE.exists():
            st = STOCK_CSV_FILE.read_text(encoding='utf-8-sig', errors='replace')
            sep = ';' if st.split('\n')[0].count(';') > st.split('\n')[0].count(',') else ','
            for r in _csv3.DictReader(_io3.StringIO(st), delimiter=sep):
                sku = (r.get('product_sku') or r.get('sku') or '').strip().upper()
                if sku:
                    entry = {
                        'stock': int(float(r.get('stock', 0) or 0)),
                        'stock30': int(float(r.get('stock30', 0) or 0)),
                        'title': r.get('title', ''),
                    }
                    stock_map[sku] = entry
                    # Dodaj pod koren (BATHFLEX_white → BATHFLEX)
                    root = smart_root(sku).upper()
                    if root not in stock_root_map:
                        stock_root_map[root] = []
                    stock_root_map[root].append(entry)

        def get_stock(sku):
            # 1. Točen match
            if sku in stock_map: return stock_map[sku]
            # 2. Koren match (BATHFLEX → BATHFLEX_white + BATHFLEX_black sešteto)
            root = smart_root(sku).upper()
            variants = stock_root_map.get(sku) or stock_root_map.get(root)
            if variants:
                return {
                    'stock': sum(v['stock'] for v in variants),
                    'stock30': sum(v['stock30'] for v in variants),
                    'title': variants[0]['title'],
                }
            return {}

        items = []
        for r in rows:
            sku = (r.get('sku') or '').strip().upper()
            st_data = get_stock(sku)
            items.append({
                'campaign': r.get('campaign', ''),
                'sku': sku,
                'title': st_data.get('title', ''),
                'status': r.get('status', 'inactive'),
                'spend': float(r.get('spend', 0) or 0),
                'conversions': float(r.get('conversions', 0) or 0),
                'currency': r.get('currency', 'EUR'),
                'stock': st_data.get('stock', 0),
                'stock30': st_data.get('stock30', 0),
                'create_time': r.get('create_time', ''),
            })

        meta = {}
        if TIKTOK_ADS_META.exists():
            try: meta = json.loads(TIKTOK_ADS_META.read_text(encoding='utf-8'))
            except: pass

        return {"items": items, "total": len(items), **meta}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/analiza-tiktok-clear")
async def analiza_tiktok_clear():
    if TIKTOK_ADS_FILE.exists(): TIKTOK_ADS_FILE.unlink()
    if TIKTOK_ADS_META.exists(): TIKTOK_ADS_META.unlink()
    return {"ok": True}

# ─── TIKTOK KREATIVE ──────────────────────────────────────────────────────────
TIKTOK_KR_FILE = DATA_DIR / "tiktok_kreative.csv"
TIKTOK_CL_FILE = DATA_DIR / "tiktok_creative_library.csv"  # Creative Library resolucije

@app.post("/analiza-ttcreative-library-upload")
async def analiza_ttcreative_library_upload(file: UploadFile = File(...)):
    """Naloži TikTok Creative Library CSV z Video name + Video ID + resolucijo."""
    try:
        content = await file.read()
        fname = file.filename or ""
        import io, re as _re

        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(r is not None for r in row)]
        else:
            import csv as _csv
            text = content.decode('utf-8-sig', errors='replace')
            sep = '\t' if '\t' in text.split('\n')[0] else (';' if ';' in text.split('\n')[0] else ',')
            data_rows = list(_csv.DictReader(io.StringIO(text), delimiter=sep))

        def fcol(row, *keys):
            for k in keys:
                for rk in row.keys():
                    if rk.strip().lower() == k.lower(): return row[rk]
            return None

        def extract_dims_from_str(s):
            # Podpira: 576 * 1024, 720 X 1280, 720x1280
            m = _re.search(r'(\d{3,5})\s*[*×xX]\s*(\d{3,5})', str(s or ''))
            if m:
                w, h = int(m.group(1)), int(m.group(2))
                if 100 < w < 10000 and 100 < h < 10000:
                    return w, h
            return None, None

        def clean_video_name(v):
            v = str(v or '').strip()
            if not v or v == '-': return ''
            parts = v.split(' ')
            if len(parts) > 1 and _re.match(r'^[a-f0-9]{32}$', parts[0]):
                return ' '.join(parts[1:])
            return v

        parsed = {}
        for row in data_rows:
            # Podpira oba formata: 'Video'/'Creative Name'
            video_raw = fcol(row, 'Creative Name', 'Video', 'video', 'Name', 'Ad name', 'Creative name')
            video_id = str(fcol(row, 'Video ID', 'Video material ID', 'video_id', 'ID') or '').strip()
            res_col = fcol(row, 'Resolution', 'Video resolution', 'Dimension', 'Size')

            if not video_raw or str(video_raw).strip() in ('-', '', 'None'): continue
            video = clean_video_name(video_raw)
            if not video: continue

            # Resolucija: iz stolpca (npr. "576 * 1024") ali iz video imena
            w, h = extract_dims_from_str(res_col) if res_col else (None, None)
            if not w:
                w, h = extract_dims_from_str(video_raw)

            if video not in parsed:
                parsed[video] = {'video': video, 'video_id': video_id, 'w': w or '', 'h': h or ''}
            elif w and not parsed[video]['w']:
                parsed[video]['w'] = w
                parsed[video]['h'] = h

        if not parsed:
            return JSONResponse({"error": "Ni veljavnih videov v datoteki."}, status_code=400)

        # Shrani / merge
        import csv as _csv2
        existing = {}
        if TIKTOK_CL_FILE.exists():
            t = TIKTOK_CL_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for r in _csv2.DictReader(io.StringIO(t)):
                existing[r['video']] = r
        for v, data in parsed.items():
            if v not in existing:
                existing[v] = data
            else:
                # Posodobi resolucijo če je manjka
                if data['w'] and not existing[v].get('w'):
                    existing[v]['w'] = data['w']
                    existing[v]['h'] = data['h']
                if data['video_id'] and not existing[v].get('video_id'):
                    existing[v]['video_id'] = data['video_id']

        out = io.StringIO()
        fn = ['video', 'video_id', 'w', 'h']
        writer = _csv2.DictWriter(out, fieldnames=fn, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(existing.values())
        TIKTOK_CL_FILE.write_text(out.getvalue(), encoding='utf-8')

        with_res = sum(1 for v in existing.values() if v.get('w'))
        return {"ok": True, "videos": len(existing), "with_resolution": with_res}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/analiza-ttcreative-library-data")
async def analiza_ttcreative_library_data():
    if not TIKTOK_CL_FILE.exists():
        return {"items": [], "total": 0}
    import csv as _csv3, io as _io3
    text = TIKTOK_CL_FILE.read_text(encoding='utf-8-sig', errors='replace')
    rows = list(_csv3.DictReader(_io3.StringIO(text)))
    return {"items": rows, "total": len(rows)}

@app.get("/analiza-ttkreative-search")
async def analiza_ttkreative_search(sku: str = ""):
    """Poišče vse videote za SKU — spoji Ads + Creative Library."""
    import csv as _csv4, io as _io4, re as _re4
    sku = sku.strip().upper()
    if not sku:
        return JSONResponse({"error": "Vpišite SKU."}, status_code=400)

    # 1. Ads data — video→SKU mapping
    ads_videos = {}  # video_name → {cost, conversions, status, campaign}
    if TIKTOK_KR_FILE.exists():
        text = TIKTOK_KR_FILE.read_text(encoding='utf-8-sig', errors='replace')
        for r in _csv4.DictReader(_io4.StringIO(text)):
            row_sku = (r.get('sku') or '').strip().upper()
            row_root = smart_root(row_sku).upper()
            search_root = smart_root(sku).upper()
            if row_sku == sku or row_root == search_root or row_sku.startswith(search_root):
                v = r.get('video', '').strip()
                if v and v not in ads_videos:
                    ads_videos[v] = {
                        'video': v,
                        'cost': float(r.get('cost', 0) or 0),
                        'conversions': float(r.get('conversions', 0) or 0),
                        'status': r.get('status', ''),
                        'w': r.get('w') or None,
                        'h': r.get('h') or None,
                    }
                elif v in ads_videos:
                    ads_videos[v]['cost'] += float(r.get('cost', 0) or 0)
                    ads_videos[v]['conversions'] += float(r.get('conversions', 0) or 0)

    # 2. Creative Library — resolucije + vsi videoti za ta SKU
    cl_map = {}  # video_name → {w, h, video_id}
    cl_sku_videos = []  # videoti iz CL ki so vezani na ta SKU (po imenu)
    search_root = smart_root(sku).upper()
    if TIKTOK_CL_FILE.exists():
        text = TIKTOK_CL_FILE.read_text(encoding='utf-8-sig', errors='replace')
        for r in _csv4.DictReader(_io4.StringIO(text)):
            v = r.get('video', '').strip()
            if v:
                cl_map[v] = r
                # Preveri ali video ime vsebuje SKU (npr. ABPULLER19.mp4, PILARAFIT (5).mp4)
                import re as _re5
                v_clean = _re5.sub(r'\.mp4$', '', v, flags=_re5.I)
                v_clean = _re5.sub(r'\s*\(\d+\)$', '', v_clean).strip()  # odstrani (5), (1)
                v_root = _re5.sub(r'\d+$', '', v_clean).upper().strip('_- ')
                if v_root == search_root or v_clean.upper() == search_root or v_clean.upper().startswith(search_root):
                    if v not in ads_videos:
                        cl_sku_videos.append(v)

    # 3. Spoji — Ads videoti + CL-only videoti
    results = []
    for v, data in ads_videos.items():
        cl = cl_map.get(v, {})
        w = data.get('w') or cl.get('w') or None
        h = data.get('h') or cl.get('h') or None
        try: w = int(w) if w else None
        except: w = None
        try: h = int(h) if h else None
        except: h = None
        results.append({
            'video': v,
            'video_id': cl.get('video_id', ''),
            'w': w, 'h': h,
            'cost': round(data['cost'], 2),
            'conversions': int(data['conversions']),
            'status': data['status'],
            'source': 'ads',
            'res_status': _res_status(w, h),
        })

    # Dodaj CL-only videote (brez spend, samo v Creative Library)
    for v in cl_sku_videos:
        cl = cl_map.get(v, {})
        try: w = int(cl.get('w')) if cl.get('w') else None
        except: w = None
        try: h = int(cl.get('h')) if cl.get('h') else None
        except: h = None
        results.append({
            'video': v,
            'video_id': cl.get('video_id', ''),
            'w': w, 'h': h,
            'cost': 0,
            'conversions': 0,
            'status': '',
            'source': 'library',
            'res_status': _res_status(w, h),
        })

    # Sortiraj po cost desc
    results.sort(key=lambda x: x['cost'], reverse=True)
    return {"sku": sku, "items": results, "total": len(results)}

def _res_status(w, h):
    if not w or not h: return "unknown"
    ratio = w / h
    if ratio < 0.7: min_w, min_h = 540, 960
    elif ratio < 1.3: min_w, min_h = 640, 640
    else: min_w, min_h = 960, 540
    if w < min_w or h < min_h: return "bad"
    if min(w, h) < 720: return "warn"
    return "ok"


@app.post("/analiza-ttkreative-upload")
async def analiza_ttkreative_upload(file: UploadFile = File(...)):
    """Naloži TikTok Ad level XLSX z Video + Campaign name stolpci."""
    try:
        content = await file.read()
        fname = file.filename or ""
        import io, re as _re

        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(r is not None for r in row)]
        else:
            import csv as _csv, io as _io
            text = content.decode('utf-8-sig', errors='replace')
            sep = ';' if text.split('\n')[0].count(';') > text.split('\n')[0].count(',') else ','
            data_rows = list(_csv.DictReader(_io.StringIO(text), delimiter=sep))

        def fcol(row, *keys):
            for k in keys:
                for rk in row.keys():
                    if rk.strip().lower() == k.lower(): return row[rk]
            return None

        def extract_dims(video):
            m = _re.search(r'(\d{3,4})\s*[xX×]\s*(\d{3,4})', str(video or ''))
            if m:
                w, h = int(m.group(1)), int(m.group(2))
                return w, h
            return None, None

        def clean_video_name(v):
            v = str(v or '').strip()
            if not v or v == '-': return ''
            parts = v.split(' ')
            if len(parts) > 1 and _re.match(r'^[a-f0-9]{32}$', parts[0]):
                return ' '.join(parts[1:])
            return v

        parsed = []
        for row in data_rows:
            campaign = str(fcol(row, 'Campaign name') or '').strip()
            video_raw = fcol(row, 'Video', 'video')
            if not video_raw or str(video_raw).strip() in ('-', '', 'None'): continue

            # SKU iz campaign name
            m = _re.search(r'Smart\+\s+([A-Z0-9_]+)', campaign, _re.I)
            if m: sku = smart_root(m.group(1)).upper()
            else:
                m = _re.search(r'SKU:\s*([A-Z0-9_]+)', campaign, _re.I)
                sku = smart_root(m.group(1)).upper() if m else ''
            if not sku: continue

            status = str(fcol(row, 'Primary status', 'Status') or '').strip()
            try: cost = float(str(fcol(row, 'Cost') or 0).replace(',', '.'))
            except: cost = 0
            try: conversions = float(str(fcol(row, 'Conversions') or 0).replace(',', '.'))
            except: conversions = 0

            # Razdeli po vejici — en oglas ima lahko več videov
            raw_str = str(video_raw).strip()
            video_parts = [v.strip() for v in raw_str.split(',') if v.strip() and v.strip() != '-']

            for vp in video_parts:
                video = clean_video_name(vp)
                if not video: continue
                w, h = extract_dims(vp)
                # Cost/conversions delimo enakomerno med videote
                n = len(video_parts)
                parsed.append({
                    'sku': sku, 'video': video,
                    'w': w or '', 'h': h or '',
                    'cost': round(cost / n, 4), 'conversions': round(conversions / n, 4),
                    'status': status, 'campaign': campaign,
                })

        if not parsed:
            return JSONResponse({"error": "Ni veljavnih videov v datoteki."}, status_code=400)

        # Dedup po sku+video, summiraj cost/conversions
        import csv as _csv2
        existing = {}
        if TIKTOK_KR_FILE.exists():
            t = TIKTOK_KR_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for r in _csv2.DictReader(io.StringIO(t)):
                key = r['sku'] + '||' + r['video']
                existing[key] = r
        for r in parsed:
            key = r['sku'] + '||' + r['video']
            if key in existing:
                try: existing[key]['cost'] = float(existing[key]['cost']) + r['cost']
                except: existing[key]['cost'] = r['cost']
                try: existing[key]['conversions'] = float(existing[key]['conversions']) + r['conversions']
                except: existing[key]['conversions'] = r['conversions']
            else:
                existing[key] = r

        out = io.StringIO()
        fieldnames = ['sku', 'video', 'w', 'h', 'cost', 'conversions', 'status', 'campaign']
        writer = _csv2.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(existing.values())
        TIKTOK_KR_FILE.write_text(out.getvalue(), encoding='utf-8')

        skus = len(set(r['sku'] for r in existing.values()))
        return {"ok": True, "videos": len(existing), "skus": skus}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/analiza-ttkreative-data")
async def analiza_ttkreative_data():
    if not TIKTOK_KR_FILE.exists():
        return JSONResponse({"error": "Ni podatkov."}, status_code=400)
    try:
        import csv as _csv3, io as _io3
        text = TIKTOK_KR_FILE.read_text(encoding='utf-8-sig', errors='replace')
        rows = list(_csv3.DictReader(_io3.StringIO(text)))
        items = []
        for r in rows:
            try: w = int(r['w']) if r.get('w') else None
            except: w = None
            try: h = int(r['h']) if r.get('h') else None
            except: h = None
            try: cost = float(r.get('cost', 0) or 0)
            except: cost = 0
            try: conv = float(r.get('conversions', 0) or 0)
            except: conv = 0
            items.append({
                'sku': r.get('sku', ''), 'video': r.get('video', ''),
                'w': w, 'h': h, 'cost': cost, 'conversions': conv,
                'status': r.get('status', ''), 'campaign': r.get('campaign', ''),
            })
        return {"items": items, "total": len(items)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)



# ─── ANALIZA: Meta Ads CSV upload ──────────────────────────────────────────────

META_ADS_FILE = DATA_DIR / "meta_ads_report.csv"
META_ADS_META = DATA_DIR / "meta_ads_meta.json"


# Stop words - skupne besede ki niso SKU
SKU_STOPWORDS = {
    'STOP', 'BIDCAP', 'COSTCAP', 'BID', 'CPA', 'BC', 'EX', 'WP', 'WV', 'EU', 'OFF',
    'LOCAL', 'OUTLET', 'MAAARKET', 'MULTIPLE', 'MAAARKET.HR', 'MAAARKET.RS', 'MAAARKET.SK',
    'NI', 'ZALOGE', 'NOVO', 'CATALOG', 'CATALOGSALE', 'CATEGORY', 'INTERESTED', 'AUTO',
    'ADVANTAGE', 'PROSPECTING', 'REMARKETING', 'CAMPAIGN', 'LOOKALIKE', 'BROAD', 'AAA',
    'DABA', 'COLD', 'KATALOG', 'INFLATED', 'INTERESTS', 'ABO', 'ZIPPLY', 'EASYZO', 'SUBAN',
    'INTEREST', 'TOFU', 'BOFU', 'MOFU', 'ALL'
}


def _load_known_skus():
    """Naloži SKU-je iz CSV zaloge (case-insensitive set)."""
    if not STOCK_CSV_FILE.exists():
        return set()
    try:
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','
        import csv as _csv
        from io import StringIO as _SIO
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        skus = set()
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if sku:
                skus.add(sku.upper())
                # Dodaj tudi koren (PLANTUP_white -> PLANTUP, Maaa61lightBrown -> Maaa61)
                koren = smart_root(sku).upper()
                if koren and len(koren) >= 4:
                    skus.add(koren)
        return skus
    except Exception as e:
        print(f"[meta] _load_known_skus err: {e}")
        return set()


def smart_root(s: str) -> str:
    """Pridobi koren SKU-ja:
    1. Razdeli po _-/presledek (PLANTUP_white → PLANTUP, COVERKA_2x3m → COVERKA)
    2. Camel-case prehod (Maaa61lightBrown → Maaa61)
    3. Digit-pred-male-črke (M261red → M261, Maaa6red → Maaa6)
    """
    if not s:
        return s
    base = re.split(r'[_\-\s]', s)[0]
    cut = None
    # Camel-case: lower → Upper
    m = re.search(r'([a-z])([A-Z])', base)
    if m:
        idx = m.start() + 1
        while idx > 1 and base[idx-1].islower():
            idx -= 1
        if idx > 0:
            cut = idx
    # Digit followed by lowercase (M261red, Maaa6red)
    if cut is None:
        m2 = re.search(r'(\d)([a-z])', base)
        if m2:
            cut = m2.start() + 1
    return base[:cut] if cut is not None else base


def extract_skus_from_text(text: str, known_skus: set = None) -> list[str]:
    """Izvleče SKU tokene iz teksta. Če je known_skus dan, vrača samo te.
    
    Match strategija:
    1. Exact match (case-insensitive) v known_skus
    2. Koren match - PLANTUP_white → PLANTUP
    3. Mixed-case dovoljen če je v known_skus (npr. Maaa61, silux74)
    """
    if not text:
        return []
    tokens = []
    
    # Pripravi case-insensitive lookup
    known_upper = set()
    if known_skus is not None:
        known_upper = {s.upper() for s in known_skus}
    
    for raw in text.split():
        # Odstrani emoji in posebne znake na začetku/koncu
        cleaned = re.sub(r'^[^\w]+|[^\w]+$', '', raw)
        if not cleaned or len(cleaned) < 4:
            continue
        # Ne sme vsebovati piko/decimal (filtrira 9.0BID, BID8.5)
        if '.' in cleaned:
            continue
        # Mora vsebovati vsaj 1 črko
        if not any(c.isalpha() for c in cleaned):
            continue
        # Skip stopwords (case-insensitive)
        if cleaned.upper() in SKU_STOPWORDS:
            continue
        # Skip čiste številke + max 1 črka (90D, 7D, 30D, 1ER ipd.)
        if re.match(r'^\d+[A-Z]?$', cleaned, re.IGNORECASE):
            continue
        
        cleaned_upper = cleaned.upper()
        
        # Če imamo known_skus, preverjamo pripadnost
        if known_skus is not None:
            # Exact match (case-insensitive)
            if cleaned_upper in known_upper:
                tokens.append(cleaned)
                continue
            # Koren match (PLANTUP_white → PLANTUP, Maaa61lightBrown → Maaa61)
            koren = smart_root(cleaned)
            if koren.upper() in known_upper:
                tokens.append(cleaned)
                continue
            # Le UPPERCASE besede (>=4 znaki) ki niso v knownh — ignoriraj
        else:
            # Brez known_skus - sprejmemo le UPPERCASE besede (legacy)
            if cleaned == cleaned.upper():
                tokens.append(cleaned)
    return tokens


@app.post("/analiza-meta-upload")
async def analiza_meta_upload(file: UploadFile = File(...)):
    """Sprejme CSV iz FB Ads Manager export, DODA k obstoječim (accumulate po Campaign name unikatnosti)."""
    try:
        import csv as _csv
        from io import StringIO as _SIO

        content_bytes = await file.read()
        new_text = content_bytes.decode('utf-8-sig', errors='replace')
        first_line = new_text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','

        # Preberi nove vrstice
        new_reader = _csv.DictReader(_SIO(new_text), delimiter=sep)
        new_rows = [r for r in new_reader if r.get('Campaign name', '').strip()]
        if not new_rows:
            return JSONResponse({"error": "CSV nima veljavnih vrstic."}, status_code=400)

        headers = list(new_rows[0].keys())

        # Preberi obstoječe vrstice (če obstajajo)
        existing_rows = []
        if META_ADS_FILE.exists():
            try:
                ex_text = META_ADS_FILE.read_text(encoding='utf-8-sig', errors='replace')
                ex_sep = ';' if ex_text.split('\n',1)[0].count(';') > ex_text.split('\n',1)[0].count(',') else ','
                ex_reader = _csv.DictReader(_SIO(ex_text), delimiter=ex_sep)
                existing_rows = [r for r in ex_reader if r.get('Campaign name', '').strip()]
            except: pass

        # Deduplikacija: ključ = Campaign name + Account name + Reporting starts
        def row_key(r):
            return (r.get('Campaign name','').strip(), r.get('Account name','').strip(), r.get('Reporting starts','').strip())

        existing_keys = {row_key(r) for r in existing_rows}
        added = [r for r in new_rows if row_key(r) not in existing_keys]
        merged = existing_rows + added

        # Shrani merged CSV
        out = _SIO()
        # Združi vse headerje (union)
        all_headers = list(dict.fromkeys(headers + [h for h in (existing_rows[0].keys() if existing_rows else []) if h not in headers]))
        writer = _csv.DictWriter(out, fieldnames=all_headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(merged)
        META_ADS_FILE.write_text(out.getvalue(), encoding='utf-8')

        # Zapiši meta (seznam vseh naloženih fileov)
        meta = {}
        if META_ADS_META.exists():
            try: meta = json.loads(META_ADS_META.read_text(encoding='utf-8'))
            except: pass
        uploads = meta.get('uploads', [])
        uploads.append({
            "filename": file.filename,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "rows_added": len(added),
            "rows_total": len(merged),
        })
        meta = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "filename": file.filename,
            "rows": len(merged),
            "rows_added": len(added),
            "uploads": uploads[-10:],  # ohrani zadnjih 10
        }
        META_ADS_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

        # Zberi accounte iz merged podatkov
        accounts = sorted(set(r.get('Account name','').strip() for r in merged if r.get('Account name','').strip()))

        return {
            "status": "ok",
            "rows_added": len(added),
            "rows_total": len(merged),
            "uploaded_at": meta["uploaded_at"],
            "filename": file.filename,
            "accounts": accounts,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/analiza-meta-clear")
async def analiza_meta_clear():
    """Počisti vse naložene Meta Ads CSV podatke."""
    try:
        if META_ADS_FILE.exists(): META_ADS_FILE.unlink()
        if META_ADS_META.exists(): META_ADS_META.unlink()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)



@app.get("/analiza-meta-data")
async def analiza_meta_data():
    """Vrne agregirane podatke iz Meta Ads CSV: SKU → metrike po accountu."""
    if not META_ADS_FILE.exists():
        return {"loaded": False}

    try:
        text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','

        import csv as _csv
        from io import StringIO as _SIO
        reader = _csv.DictReader(_SIO(text), delimiter=sep)

        # Agregirajmo po SKU (in kasneje v JS po accountu)
        sku_data = {}  # sku → {accounts: {acc: {spend, purchases, etc}}, campaigns: [...]}

        def _f(v):
            try:
                return float(str(v or '0').replace(',', '.'))
            except:
                return 0.0

        def _i(v):
            try:
                return int(float(str(v or '0').replace(',', '.')))
            except:
                return 0

        # Naloži znane SKU-je iz zaloge (za boljši filter)
        known_skus = _load_known_skus()

        for row in reader:
            campaign_name = (row.get('Campaign name') or '').strip()
            if not campaign_name:
                continue

            account = (row.get('Account name') or '').strip() or '—'
            spend = _f(row.get('Amount spent (EUR)'))
            purchases = _i(row.get('Purchases'))
            cpa = _f(row.get('Cost per purchase'))
            cpc = _f(row.get('CPC (cost per link click)'))
            ctr = _f(row.get('CTR (link click-through rate)'))
            atc = _i(row.get('Adds to cart'))
            freq = _f(row.get('Frequency'))
            # Status SAMO iz FB Campaign Delivery kolone (edina resnica)
            # Ime kampanje (@STOP, ⛔, OFF) se IGNORIRA — to so naši interni oznaki
            delivery = (row.get('Campaign Delivery') or '').strip().lower()
            if delivery == 'inactive':
                is_stopped = True
            elif delivery == 'active':
                is_stopped = False
            else:
                # Stolpec manjka ali neznana vrednost → privzeto aktivna
                # (raje napačno aktivna kot napačno pavzirana — ker FB sam ne ve)
                is_stopped = False

            # Izvleci SKU-je iz imena (filtrirano po znanih SKU iz zaloge)
            skus = extract_skus_from_text(campaign_name, known_skus if known_skus else None)
            # Dedupliciraj
            skus = list(dict.fromkeys(skus))

            for sku in skus:
                if sku not in sku_data:
                    sku_data[sku] = {
                        "sku": sku,
                        "campaigns": [],
                        "accounts": {},
                        "total_spend": 0,
                        "total_purchases": 0,
                        "total_atc": 0,
                        "total_clicks_value": 0,  # za uteženo CPC
                        "campaign_count": 0,
                        "stopped_count": 0,
                    }
                d = sku_data[sku]
                d["campaigns"].append({
                    "name": campaign_name,
                    "account": account,
                    "spend": spend,
                    "purchases": purchases,
                    "cpa": cpa,
                    "atc": atc,
                    "stopped": is_stopped,
                })
                if account not in d["accounts"]:
                    d["accounts"][account] = {"spend": 0, "purchases": 0, "campaigns": 0, "active": 0, "paused": 0}
                d["accounts"][account]["spend"] += spend
                d["accounts"][account]["purchases"] += purchases
                d["accounts"][account]["campaigns"] += 1
                if is_stopped:
                    d["accounts"][account]["paused"] += 1
                else:
                    d["accounts"][account]["active"] += 1
                d["total_spend"] += spend
                d["total_purchases"] += purchases
                d["total_atc"] += atc
                d["campaign_count"] += 1
                if is_stopped:
                    d["stopped_count"] += 1

        # Pripravi flat seznam za frontend
        items = []
        for sku, d in sku_data.items():
            avg_cpa = (d["total_spend"] / d["total_purchases"]) if d["total_purchases"] > 0 else None
            # Izračunaj per-account status (active = vsaj 1 aktivna, paused = vse pavzirane)
            accounts_with_status = []
            for k, v in d["accounts"].items():
                acc = {"name": k, **v}
                if v.get("active", 0) > 0:
                    acc["status"] = "active"
                elif v.get("paused", 0) > 0:
                    acc["status"] = "paused"
                else:
                    acc["status"] = "none"
                accounts_with_status.append(acc)
            items.append({
                "sku": sku,
                "total_spend": round(d["total_spend"], 2),
                "total_purchases": d["total_purchases"],
                "total_atc": d["total_atc"],
                "avg_cpa": round(avg_cpa, 2) if avg_cpa is not None else None,
                "campaign_count": d["campaign_count"],
                "stopped_count": d["stopped_count"],
                "active_count": d["campaign_count"] - d["stopped_count"],
                "accounts": accounts_with_status,
            })

        # Zberi vse accounte dinamično iz podatkov (ne hardkodiran whitelist)
        all_accounts = sorted(set(
            acc_name
            for d in sku_data.values()
            for acc_name in d["accounts"].keys()
            if acc_name and acc_name != '—'
        ))

        meta = {}
        if META_ADS_META.exists():
            try:
                meta = json.loads(META_ADS_META.read_text(encoding="utf-8"))
            except: pass

        return {
            "loaded": True,
            "items": items,
            "total_skus": len(items),
            "accounts": all_accounts,  # dynamic lista za frontend checkboxe
            **{k: v for k, v in meta.items() if k != 'accounts'},
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── ANALIZA: Obrat 14 dni ─────────────────────────────────────────────────────

OBRAT14_FILE = DATA_DIR / "obrat_14dni.txt"
OBRAT14_META = DATA_DIR / "obrat_14dni_meta.json"

# Whitelist accounts za Obrat 14dni view
TARGET_ACCOUNTS = ['Maaarket X', 'Maaarket ALL', 'Maaarket ALL2', 'Maaarket ALL3 + RS', 'Zipply.', 'si_SUBAN_Maaarket SK', 'Maaarket PL/RO', 'Maaarket HR', 'si_Suban_Maaarket HR', 'Easyzo', 'Thundershop ALL HU', 'ThunderShop HR', 'ThunderShop RS']
# Ko dobiš nova accounta, dodaj ju sem IN v AD_ACCOUNTS_CONFIG v index.html


@app.post("/analiza-obrat14-upload")
async def analiza_obrat14_upload(file: UploadFile = File(...)):
    """Sprejme TXT/TSV iz top obratov, shrani."""
    try:
        content_bytes = await file.read()
        OBRAT14_FILE.write_bytes(content_bytes)

        text = content_bytes.decode('utf-8-sig', errors='replace')
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        meta = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "filename": file.filename,
            "size": len(content_bytes),
            "rows": max(0, len(lines) - 1),
        }
        OBRAT14_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"status": "ok", "rows": meta["rows"], "uploaded_at": meta["uploaded_at"], "filename": file.filename}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/analiza-obrat14-data")
async def analiza_obrat14_data():
    """Vrne obrat 14dni podatke + match z FB Ads (po accountu, samo whitelist)."""
    if not OBRAT14_FILE.exists():
        return {"loaded": False}

    try:
        # Naloži obrat14
        text = OBRAT14_FILE.read_text(encoding="utf-8-sig", errors="replace")
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if len(lines) < 2:
            return {"loaded": False, "error": "Prazna datoteka."}

        # Detect separator (tab ali ;)
        first = lines[0]
        if '\t' in first:
            sep = '\t'
        elif ';' in first:
            sep = ';'
        else:
            sep = ','

        items = []
        # Skip header
        for line in lines[1:]:
            parts = line.split(sep)
            if len(parts) < 3:
                continue
            sku = parts[0].strip()
            naziv = parts[1].strip()
            try:
                kolicina = int(float(parts[2].strip().replace(',', '.')))
            except:
                kolicina = 0
            if not sku:
                continue
            items.append({
                "sku": sku,
                "naziv": naziv,
                "kolicina": kolicina,
            })

        # Match z FB Ads — uporabljaj že shranjen META_ADS_FILE
        sku_ads_map = {}  # sku.upper() → {account: {spend, purchases}, ...}
        if META_ADS_FILE.exists():
            try:
                fb_text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
                fb_first = fb_text.split('\n', 1)[0]
                fb_sep = ';' if fb_first.count(';') > fb_first.count(',') else ','
                import csv as _csv
                from io import StringIO as _SIO
                reader = _csv.DictReader(_SIO(fb_text), delimiter=fb_sep)

                # Pripravi seznam znanih SKU iz obrat14 (za extract_skus filter)
                # Vključimo VSE variante — uppercase (za standardni match) + originalne + korene
                known_skus_obrat = set()
                for it in items:
                    s = it["sku"]
                    known_skus_obrat.add(s)
                    known_skus_obrat.add(s.upper())
                    # Koren (PLANTUP_white -> PLANTUP, Maaa61lightBrown -> Maaa61, COVERKA_2x3m -> COVERKA)
                    koren = smart_root(s)
                    if koren and len(koren) >= 4:
                        known_skus_obrat.add(koren)
                        known_skus_obrat.add(koren.upper())

                def _f(v):
                    try: return float(str(v or '0').replace(',', '.'))
                    except: return 0.0
                def _i(v):
                    try: return int(float(str(v or '0').replace(',', '.')))
                    except: return 0

                for row in reader:
                    cname = (row.get('Campaign name') or '').strip()
                    if not cname:
                        continue
                    account = (row.get('Account name') or '').strip() or '—'
                    spend = _f(row.get('Amount spent (EUR)'))
                    purchases = _i(row.get('Purchases'))
                    
                    # Status SAMO iz FB Campaign Delivery kolone (edina resnica)
                    # Ime kampanje (@STOP, ⛔, OFF) se IGNORIRA — to so naši interni oznaki
                    delivery = (row.get('Campaign Delivery') or '').strip().lower()
                    if delivery == 'inactive':
                        is_active_campaign = False
                    elif delivery == 'active':
                        is_active_campaign = True
                    else:
                        # Stolpec manjka ali neznana vrednost → privzeto aktivna
                        is_active_campaign = True

                    # Izvleci SKU iz imena
                    skus = extract_skus_from_text(cname, known_skus_obrat)
                    skus = list(dict.fromkeys(skus))

                    # Za vsak SKU token iz kampanje, najdi VSE matching izdelke v 14dni
                    for sku_token in skus:
                        sku_upper = sku_token.upper()
                        token_koren = smart_root(sku_upper)
                        target_skus = []
                        
                        # 1. Exact (case-insensitive)
                        for it in items:
                            if it["sku"].upper() == sku_upper:
                                target_skus = [it["sku"]]
                                break
                        
                        if not target_skus:
                            # 2. Koren match — vsi izdelki s tem korenom
                            candidates = []
                            for it in items:
                                item_koren = smart_root(it["sku"]).upper()
                                if item_koren == token_koren or item_koren == sku_upper or token_koren == it["sku"].upper():
                                    candidates.append(it["sku"])
                            
                            if len(candidates) == 0:
                                continue
                            elif len(candidates) == 1:
                                target_skus = candidates
                            else:
                                # 3. Več zadetkov: če je sku_token preprosto koren (npr. "Maaa61") → vse variante
                                if sku_upper == token_koren:
                                    target_skus = candidates
                                else:
                                    # Fuzzy match — najdi najbolj podobnega
                                    from difflib import SequenceMatcher
                                    best, best_ratio = None, 0
                                    for c in candidates:
                                        ratio = SequenceMatcher(None, sku_upper, c.upper()).ratio()
                                        if ratio > best_ratio:
                                            best_ratio = ratio
                                            best = c
                                    if best and best_ratio >= 0.6:
                                        target_skus = [best]
                                    else:
                                        target_skus = candidates  # fallback: vse

                        # Zabeleži za vse target SKU-je
                        for target_sku in target_skus:
                            if target_sku not in sku_ads_map:
                                sku_ads_map[target_sku] = {}
                            if account not in sku_ads_map[target_sku]:
                                sku_ads_map[target_sku][account] = {"spend": 0, "purchases": 0, "campaigns": 0, "active": 0, "paused": 0}
                            # Razdelimo spend/purchases enakomerno če je več target SKU
                            split = max(1, len(target_skus))
                            sku_ads_map[target_sku][account]["spend"] += spend / split
                            sku_ads_map[target_sku][account]["purchases"] += purchases / split
                            sku_ads_map[target_sku][account]["campaigns"] += 1
                            if is_active_campaign:
                                sku_ads_map[target_sku][account]["active"] += 1
                            else:
                                sku_ads_map[target_sku][account]["paused"] += 1
            except Exception as e:
                print(f"[obrat14] FB match err: {e}")

        # Agregiraj — vsakemu accountu dodaj status field (active/paused/none)
        for it in items:
            accounts_data = sku_ads_map.get(it["sku"], {})
            for acc_name, acc_data in accounts_data.items():
                if acc_data.get("active", 0) > 0:
                    acc_data["status"] = "active"
                elif acc_data.get("paused", 0) > 0:
                    acc_data["status"] = "paused"
                else:
                    acc_data["status"] = "none"
            it["accounts"] = accounts_data
            it["has_ads"] = len(accounts_data) > 0
            # Skupni status
            statuses = [a.get("status") for a in accounts_data.values()]
            if "active" in statuses:
                it["overall_status"] = "active"
            elif "paused" in statuses:
                it["overall_status"] = "paused"
            else:
                it["overall_status"] = "none"

        meta = {}
        if OBRAT14_META.exists():
            try:
                meta = json.loads(OBRAT14_META.read_text(encoding="utf-8"))
            except: pass

        return {
            "loaded": True,
            "items": items,
            "total": len(items),
            "target_accounts": TARGET_ACCOUNTS,
            **meta,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════
# HS+ UVOZ — naročilnice z združevanjem SKU + history
# ═══════════════════════════════════════════════════════════════
HSUVOZ_DIR = DATA_DIR / "hsuvoz_history"
HSUVOZ_DIR.mkdir(exist_ok=True, parents=True)
HSUVOZ_CURRENT = DATA_DIR / "hsuvoz_current.json"


def hsuvoz_cleanup():
    """Briše JSON datoteke starejše od 30 dni."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in HSUVOZ_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[hsuvoz] cleanup err: {e}")


@app.post("/hsuvoz-upload")
async def hsuvoz_upload(file: UploadFile = File(...)):
    """Sprejme CSV naročilnic, združi količine po SKU, shrani v current + history."""
    import csv
    from io import StringIO
    try:
        content = (await file.read()).decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(StringIO(content))

        # Normalizacija headerjev
        rows = []
        for row in reader:
            norm = {k.strip().replace('\ufeff', ''): v.strip() for k, v in row.items()}
            rows.append(norm)

        if not rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Najdi kolone (ID naročila, SKU, Naziv, Količina)
        sample = rows[0]
        keys = list(sample.keys())

        def find_col(candidates):
            for c in candidates:
                for k in keys:
                    if c.lower() in k.lower():
                        return k
            return None

        id_col  = find_col(["id naročila", "id narocila", "id"])
        sku_col = find_col(["sku"])
        naz_col = find_col(["naziv", "name"])
        qty_col = find_col(["količina", "kolicina", "qty", "quantity"])

        if not sku_col or not qty_col:
            return JSONResponse({"error": f"Ne najdem SKU/Količina kolon. Najdene: {keys}"}, status_code=400)

        # Združi po SKU
        sku_map = {}  # sku → {naziv, qty, orders: [id,...]}
        for row in rows:
            sku = (row.get(sku_col) or "").strip()
            if not sku:
                continue
            try:
                qty = int(float((row.get(qty_col) or "0").replace(",", ".")))
            except:
                qty = 0
            naziv = (row.get(naz_col) or "").strip() if naz_col else ""
            order_id = (row.get(id_col) or "").strip() if id_col else ""

            if sku not in sku_map:
                sku_map[sku] = {"sku": sku, "naziv": naziv, "qty": 0, "orders": [], "done": False}
            sku_map[sku]["qty"] += qty
            if order_id and order_id not in sku_map[sku]["orders"]:
                sku_map[sku]["orders"].append(order_id)
            # Ohrani naziv (vzemi prvega ki ni prazen)
            if not sku_map[sku]["naziv"] and naziv:
                sku_map[sku]["naziv"] = naziv

        items = sorted(sku_map.values(), key=lambda x: -x["qty"])

        # Naloži obstoječe done state iz currenta (ohrani done flag)
        if HSUVOZ_CURRENT.exists():
            try:
                existing = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
                done_map = {it["sku"]: it.get("done", False) for it in existing.get("items", [])}
                for it in items:
                    if it["sku"] in done_map:
                        it["done"] = done_map[it["sku"]]
            except:
                pass

        ts = datetime.now(timezone.utc).isoformat()
        payload = {
            "uploaded_at": ts,
            "filename": file.filename,
            "total_skus": len(items),
            "items": items,
        }

        # Shrani current
        HSUVOZ_CURRENT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        # Shrani v history
        hsuvoz_cleanup()
        hist_file = HSUVOZ_DIR / f"hsuvoz_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.json"
        hist_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True, "total_skus": len(items), "uploaded_at": ts, "filename": file.filename}

    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/hsuvoz-data")
async def hsuvoz_data():
    """Vrne trenutne HS+ uvoz podatke."""
    try:
        if not HSUVOZ_CURRENT.exists():
            return {"loaded": False, "items": []}
        data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        return {"loaded": True, **data}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-set-done")
async def hsuvoz_set_done(sku: str = None, done: str = "1", request: Request = None):
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku")
                    done = str(body.get("done", True)).lower()
            except: pass
        done_bool = done not in ("0", "false", "False")
        if not HSUVOZ_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == sku:
                it["done"] = done_bool
                break
        HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-edit-sku")
async def hsuvoz_edit_sku(old_sku: str = None, new_sku: str = None, source: str = "current", request: Request = None):
    """Preimenuje SKU v current ali order."""
    try:
        if (not old_sku or not new_sku) and request:
            try:
                raw = await request.body()
                if raw:
                    body = __import__("json").loads(raw)
                    old_sku = old_sku or body.get("old_sku")
                    new_sku = new_sku or (body.get("new_sku") or "").strip()
                    source = body.get("source", source)
            except: pass
        new_sku = (new_sku or "").strip()
        if not new_sku:
            return JSONResponse({"error": "Nov SKU je prazen."}, status_code=400)
        file = HSUVOZ_CURRENT if source == "current" else HSUVOZ_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(file.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == old_sku:
                it["sku"] = new_sku
                break
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/hsuvoz-history")
async def hsuvoz_history():
    """Vrne seznam zgodovinskih uploadov (30 dni)."""
    hsuvoz_cleanup()
    try:
        items = []
        for f in sorted(HSUVOZ_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "original_filename": d.get("filename", f.name),
                    "uploaded_at": d.get("uploaded_at", ""),
                    "total_skus": d.get("total_skus", 0),
                    "size": f.stat().st_size,
                })
            except:
                pass
        return {"items": items}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-load-history")
async def hsuvoz_load_history(filename: str = None, request: Request = None):
    """Naloži zgodovinski upload kot current."""
    try:
        if not filename and request:
            try:
                raw = await request.body()
                if raw: filename = __import__("json").loads(raw).get("filename")
            except: pass
        fname = filename
        hist_file = HSUVOZ_DIR / fname
        if not hist_file.exists():
            return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)
        data = json.loads(hist_file.read_text(encoding="utf-8"))
        HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "total_skus": data.get("total_skus", 0)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════
# HS+ NAROČANJE — naročilo composer (ločen persistent state)
# ═══════════════════════════════════════════════════════════════
HSUVOZ_ORDER = DATA_DIR / "hsuvoz_order.json"

@app.post("/hsuvoz-move-to-order")
async def hsuvoz_move_to_order(sku: str = None, request: Request = None):
    """Premakne SKU iz 'za naročilo' v 'naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not HSUVOZ_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        # Poberi iz current
        current = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        item = next((it for it in current.get("items", []) if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja."}, status_code=404)

        # Dodaj v order (ali posodobi qty)
        order = {"items": []}
        if HSUVOZ_ORDER.exists():
            try: order = json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
            except: pass

        existing = next((it for it in order["items"] if it["sku"] == sku), None)
        if existing:
            existing["qty"] += item["qty"]
            existing["orders"] = list(set(existing.get("orders", []) + item.get("orders", [])))
        else:
            order["items"].append({**item, "done": False})

        HSUVOZ_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        # Odstrani iz current
        current["items"] = [it for it in current["items"] if it["sku"] != sku]
        HSUVOZ_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-move-back")
async def hsuvoz_move_back(sku: str = None, request: Request = None):
    """Vrne SKU iz naročila nazaj v 'za naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not HSUVOZ_ORDER.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        order = json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
        item = next((it for it in order["items"] if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja v naročilu."}, status_code=404)

        # Vrni v current
        current = {"items": []}
        if HSUVOZ_CURRENT.exists():
            try: current = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
            except: pass

        if not any(it["sku"] == sku for it in current["items"]):
            current["items"].append({**item, "done": False})
            HSUVOZ_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        order["items"] = [it for it in order["items"] if it["sku"] != sku]
        HSUVOZ_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-delete-item")
async def hsuvoz_delete_item(sku: str = None, source: str = "current", request: Request = None):
    """Zbriše SKU iz seznama. Sprejme query param ali JSON body."""
    try:
        # Poskusi dobiti iz JSON body če query param ni podan
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku") or (body.get("skus") or [None])[0]
                    source = body.get("source", source)
            except: pass

        if not sku:
            return JSONResponse({"error": "Manjka SKU."}, status_code=400)

        file = HSUVOZ_CURRENT if source == "current" else HSUVOZ_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)

        data = json.loads(file.read_text(encoding="utf-8"))
        before = len(data.get("items", []))
        data["items"] = [it for it in data.get("items", []) if str(it.get("sku","")).strip() != str(sku).strip()]
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "deleted": before - len(data["items"])}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/hsuvoz-order-data")
async def hsuvoz_order_data():
    """Vrne seznam SKU-jev v naročilu."""
    try:
        if not HSUVOZ_ORDER.exists():
            return {"items": []}
        return json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-order-clear")
async def hsuvoz_order_clear():
    """Počisti celotno naročilo."""
    try:
        HSUVOZ_ORDER.write_text(json.dumps({"items": []}, ensure_ascii=False), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/hsuvoz-current-clear")
async def hsuvoz_current_clear():
    """Počisti celoten seznam 'za naročilo'."""
    try:
        if HSUVOZ_CURRENT.exists():
            data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
            data["items"] = []
            HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── APTEL NAROČANJE ─────────────────────────────────────────────────────────
APTEL_DIR = DATA_DIR / "aptel_history"
APTEL_DIR.mkdir(exist_ok=True, parents=True)
APTEL_CURRENT = DATA_DIR / "aptel_current.json"


def aptel_cleanup():
    """Briše JSON datoteke starejše od 30 dni."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in APTEL_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[aptel] cleanup err: {e}")


@app.post("/aptel-upload")
async def aptel_upload(file: UploadFile = File(...)):
    """Sprejme CSV naročilnic, združi količine po SKU, shrani v current + history."""
    import csv
    from io import StringIO
    try:
        content = (await file.read()).decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(StringIO(content))

        # Normalizacija headerjev
        rows = []
        for row in reader:
            norm = {k.strip().replace('\ufeff', ''): v.strip() for k, v in row.items()}
            rows.append(norm)

        if not rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Najdi kolone (ID naročila, SKU, Naziv, Količina)
        sample = rows[0]
        keys = list(sample.keys())

        def find_col(candidates):
            for c in candidates:
                for k in keys:
                    if c.lower() in k.lower():
                        return k
            return None

        id_col  = find_col(["id naročila", "id narocila", "id"])
        sku_col = find_col(["sku"])
        naz_col = find_col(["naziv", "name"])
        qty_col = find_col(["količina", "kolicina", "qty", "quantity"])

        if not sku_col or not qty_col:
            return JSONResponse({"error": f"Ne najdem SKU/Količina kolon. Najdene: {keys}"}, status_code=400)

        # Združi po SKU
        sku_map = {}  # sku → {naziv, qty, orders: [id,...]}
        for row in rows:
            sku = (row.get(sku_col) or "").strip()
            if not sku:
                continue
            try:
                qty = int(float((row.get(qty_col) or "0").replace(",", ".")))
            except:
                qty = 0
            naziv = (row.get(naz_col) or "").strip() if naz_col else ""
            order_id = (row.get(id_col) or "").strip() if id_col else ""

            if sku not in sku_map:
                sku_map[sku] = {"sku": sku, "naziv": naziv, "qty": 0, "orders": [], "done": False}
            sku_map[sku]["qty"] += qty
            if order_id and order_id not in sku_map[sku]["orders"]:
                sku_map[sku]["orders"].append(order_id)
            # Ohrani naziv (vzemi prvega ki ni prazen)
            if not sku_map[sku]["naziv"] and naziv:
                sku_map[sku]["naziv"] = naziv

        items = sorted(sku_map.values(), key=lambda x: -x["qty"])

        # Naloži obstoječe done state iz currenta (ohrani done flag)
        if APTEL_CURRENT.exists():
            try:
                existing = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
                done_map = {it["sku"]: it.get("done", False) for it in existing.get("items", [])}
                for it in items:
                    if it["sku"] in done_map:
                        it["done"] = done_map[it["sku"]]
            except:
                pass

        ts = datetime.now(timezone.utc).isoformat()
        payload = {
            "uploaded_at": ts,
            "filename": file.filename,
            "total_skus": len(items),
            "items": items,
        }

        # Shrani current
        APTEL_CURRENT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        # Shrani v history
        aptel_cleanup()
        hist_file = APTEL_DIR / f"aptel_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.json"
        hist_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True, "total_skus": len(items), "uploaded_at": ts, "filename": file.filename}

    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/aptel-data")
async def aptel_data():
    """Vrne trenutne Aptel uvoz podatke."""
    try:
        if not APTEL_CURRENT.exists():
            return {"loaded": False, "items": []}
        data = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
        return {"loaded": True, **data}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-set-done")
async def aptel_set_done(sku: str = None, done: str = "1", request: Request = None):
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku")
                    done = str(body.get("done", True)).lower()
            except: pass
        done_bool = done not in ("0", "false", "False")
        if not APTEL_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == sku:
                it["done"] = done_bool
                break
        APTEL_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-edit-sku")
async def aptel_edit_sku(old_sku: str = None, new_sku: str = None, source: str = "current", request: Request = None):
    """Preimenuje SKU v current ali order."""
    try:
        if (not old_sku or not new_sku) and request:
            try:
                raw = await request.body()
                if raw:
                    body = __import__("json").loads(raw)
                    old_sku = old_sku or body.get("old_sku")
                    new_sku = new_sku or (body.get("new_sku") or "").strip()
                    source = body.get("source", source)
            except: pass
        new_sku = (new_sku or "").strip()
        if not new_sku:
            return JSONResponse({"error": "Nov SKU je prazen."}, status_code=400)
        file = APTEL_CURRENT if source == "current" else APTEL_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(file.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == old_sku:
                it["sku"] = new_sku
                break
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/aptel-history")
async def aptel_history():
    """Vrne seznam zgodovinskih uploadov (30 dni)."""
    aptel_cleanup()
    try:
        items = []
        for f in sorted(APTEL_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "original_filename": d.get("filename", f.name),
                    "uploaded_at": d.get("uploaded_at", ""),
                    "total_skus": d.get("total_skus", 0),
                    "size": f.stat().st_size,
                })
            except:
                pass
        return {"items": items}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-load-history")
async def aptel_load_history(filename: str = None, request: Request = None):
    """Naloži zgodovinski upload kot current."""
    try:
        if not filename and request:
            try:
                raw = await request.body()
                if raw: filename = __import__("json").loads(raw).get("filename")
            except: pass
        fname = filename
        hist_file = APTEL_DIR / fname
        if not hist_file.exists():
            return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)
        data = json.loads(hist_file.read_text(encoding="utf-8"))
        APTEL_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "total_skus": data.get("total_skus", 0)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════
# Aptel NAROČANJE — naročilo composer (ločen persistent state)
# ═══════════════════════════════════════════════════════════════
APTEL_ORDER = DATA_DIR / "aptel_order.json"

@app.post("/aptel-move-to-order")
async def aptel_move_to_order(sku: str = None, request: Request = None):
    """Premakne SKU iz 'za naročilo' v 'naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not APTEL_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        # Poberi iz current
        current = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
        item = next((it for it in current.get("items", []) if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja."}, status_code=404)

        # Dodaj v order (ali posodobi qty)
        order = {"items": []}
        if APTEL_ORDER.exists():
            try: order = json.loads(APTEL_ORDER.read_text(encoding="utf-8"))
            except: pass

        existing = next((it for it in order["items"] if it["sku"] == sku), None)
        if existing:
            existing["qty"] += item["qty"]
            existing["orders"] = list(set(existing.get("orders", []) + item.get("orders", [])))
        else:
            order["items"].append({**item, "done": False})

        APTEL_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        # Odstrani iz current
        current["items"] = [it for it in current["items"] if it["sku"] != sku]
        APTEL_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-move-back")
async def aptel_move_back(sku: str = None, request: Request = None):
    """Vrne SKU iz naročila nazaj v 'za naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not APTEL_ORDER.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        order = json.loads(APTEL_ORDER.read_text(encoding="utf-8"))
        item = next((it for it in order["items"] if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja v naročilu."}, status_code=404)

        # Vrni v current
        current = {"items": []}
        if APTEL_CURRENT.exists():
            try: current = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
            except: pass

        if not any(it["sku"] == sku for it in current["items"]):
            current["items"].append({**item, "done": False})
            APTEL_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        order["items"] = [it for it in order["items"] if it["sku"] != sku]
        APTEL_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-delete-item")
async def aptel_delete_item(sku: str = None, source: str = "current", request: Request = None):
    """Zbriše SKU iz seznama. Sprejme query param ali JSON body."""
    try:
        # Poskusi dobiti iz JSON body če query param ni podan
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku") or (body.get("skus") or [None])[0]
                    source = body.get("source", source)
            except: pass

        if not sku:
            return JSONResponse({"error": "Manjka SKU."}, status_code=400)

        file = APTEL_CURRENT if source == "current" else APTEL_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)

        data = json.loads(file.read_text(encoding="utf-8"))
        before = len(data.get("items", []))
        data["items"] = [it for it in data.get("items", []) if str(it.get("sku","")).strip() != str(sku).strip()]
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "deleted": before - len(data["items"])}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/aptel-order-data")
async def aptel_order_data():
    """Vrne seznam SKU-jev v naročilu."""
    try:
        if not APTEL_ORDER.exists():
            return {"items": []}
        return json.loads(APTEL_ORDER.read_text(encoding="utf-8"))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-order-clear")
async def aptel_order_clear():
    """Počisti celotno naročilo."""
    try:
        APTEL_ORDER.write_text(json.dumps({"items": []}, ensure_ascii=False), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/aptel-current-clear")
async def aptel_current_clear():
    """Počisti celoten seznam 'za naročilo'."""
    try:
        if APTEL_CURRENT.exists():
            data = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
            data["items"] = []
            APTEL_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ─── INVENTURA ────────────────────────────────────────────────────────────────

INVENTURA_DIR = DATA_DIR / "inventura"
INVENTURA_DIR.mkdir(exist_ok=True, parents=True)

INVENTURA_CURRENT = INVENTURA_DIR / "_current.json"

DEJAVU_REGULAR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
DEJAVU_BOLD    = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"


def inventura_cleanup():
    """Zbriše PDF-je in JSON-e starejše od 30 dni (ne _current.json)."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in list(INVENTURA_DIR.glob("*.pdf")) + list(INVENTURA_DIR.glob("*.json")):
            if f.name == "_current.json":
                continue
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[inventura] cleanup err: {e}")


@app.get("/inventura-current")
async def inventura_get_current():
    """Vrne trenutno aktivno inventuro z diska."""
    if not INVENTURA_CURRENT.exists():
        return {"ok": False, "items": []}
    try:
        data = json.loads(INVENTURA_CURRENT.read_text(encoding="utf-8"))
        return {"ok": True, **data}
    except Exception:
        return {"ok": False, "items": []}


@app.post("/inventura-save-current")
async def inventura_save_current(data: dict):
    """Shrani trenutno stanje inventure na disk."""
    try:
        INVENTURA_CURRENT.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/inventura-clear-current")
async def inventura_clear_current():
    """Pobriše trenutno aktivno inventuro z diska."""
    try:
        if INVENTURA_CURRENT.exists():
            INVENTURA_CURRENT.unlink()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/inventura-upload")
async def inventura_upload(file: UploadFile = File(...)):
    """Sprejme CSV izvoz, združi po SKU, vrne seznam."""
    try:
        import csv
        from io import StringIO
        content = (await file.read()).decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(StringIO(content))
        rows = [{k.strip().replace('\ufeff', ''): (v or '').strip() for k, v in row.items()} for row in reader]
        if not rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)
        keys = list(rows[0].keys())

        def fc(*cands):
            for c in cands:
                for k in keys:
                    if c.lower() in k.lower(): return k
            return None

        sku_col = fc("sku", "SKU")
        naz_col = fc("naziv", "name", "Naziv")
        pos_col = fc("pozicija", "position", "Pozicija")

        if not sku_col:
            return JSONResponse({"error": f"Ne najdem SKU stolpca. Najdeni: {keys}"}, status_code=400)

        sku_map = {}
        for row in rows:
            sku = (row.get(sku_col) or "").strip()
            if not sku: continue
            naziv    = (row.get(naz_col) or "").strip() if naz_col else ""
            pozicija = (row.get(pos_col) or "").strip() if pos_col else ""
            if sku not in sku_map:
                sku_map[sku] = {"sku": sku, "naziv": naziv, "pozicija": pozicija, "komentar": "", "kolicina_dejansko": None}
            if not sku_map[sku]["pozicija"] and pozicija: sku_map[sku]["pozicija"] = pozicija
            if not sku_map[sku]["naziv"] and naziv: sku_map[sku]["naziv"] = naziv

        items = sorted(sku_map.values(), key=lambda x: (x["pozicija"] or "zzz", x["sku"]))
        return {"ok": True, "total_skus": len(items), "filename": file.filename, "items": items}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/inventura-pdf")
async def inventura_pdf(data: dict):
    """Generira PDF inventurni list z DejaVu fontom."""
    try:
        items = data.get("items", [])
        title_text = data.get("title", "Inventurni list")
        datum = data.get("datum", datetime.now().strftime("%d. %m. %Y"))
        filename_hint = data.get("filename", f"inventura_{datetime.now().strftime('%Y-%m-%d_%H-%M')}")
        if not items:
            return JSONResponse({"error": "Ni postavk."}, status_code=400)

        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase.pdfmetrics import stringWidth

        pdfmetrics.registerFont(TTFont("DejaVu", DEJAVU_REGULAR))
        pdfmetrics.registerFont(TTFont("DejaVu-Bold", DEJAVU_BOLD))

        def trunc_ellipsis(s, font, font_size, max_width_pt):
            if not s: return ""
            for sep in [",", "(", " -"]:
                idx = s.find(sep)
                if 0 < idx < len(s):
                    candidate = s[:idx].strip()
                    if stringWidth(candidate, font, font_size) <= max_width_pt:
                        s = candidate; break
            if stringWidth(s, font, font_size) <= max_width_pt: return s
            while s and stringWidth(s + "...", font, font_size) > max_width_pt: s = s[:-1]
            return s.strip() + "..."

        ROW_H = 0.65 * cm
        HDR_H = 0.7  * cm

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)

        s_title = ParagraphStyle("t", fontSize=14, fontName="DejaVu-Bold", spaceAfter=4)
        s_sub   = ParagraphStyle("s", fontSize=9,  fontName="DejaVu", textColor=colors.HexColor("#64748b"), spaceAfter=12)
        s_num   = ParagraphStyle("n", fontSize=7.5, fontName="DejaVu", leading=10, alignment=1)
        s_cell  = ParagraphStyle("c", fontSize=7.5, fontName="DejaVu", leading=10)
        s_sku   = ParagraphStyle("k", fontSize=7.5, fontName="DejaVu-Bold", leading=10)
        s_kom   = ParagraphStyle("m", fontSize=7.5, fontName="DejaVu", textColor=colors.HexColor("#7c3aed"), leading=10)
        s_hdr   = ParagraphStyle("h", fontSize=7.5, fontName="DejaVu-Bold", textColor=colors.white, leading=10, alignment=1)

        story = [
            Paragraph(title_text, s_title),
            Paragraph(f"Datum: {datum}  |  Skupaj SKU-jev: {len(items)}", s_sub),
        ]

        col_widths = [0.7*cm, 4.2*cm, 5.8*cm, 2.0*cm, 3.5*cm, 1.8*cm]
        naziv_max_pt = 5.8 * 28.35 - 10
        kom_max_pt   = 3.5 * 28.35 - 10

        table_data = [[
            Paragraph("#", s_hdr), Paragraph("SKU", s_hdr), Paragraph("Naziv", s_hdr),
            Paragraph("Pozicija", s_hdr), Paragraph("Komentar", s_hdr), Paragraph("Fizično ✓", s_hdr),
        ]]
        for i, it in enumerate(items, 1):
            komentar_raw = str(it.get("komentar") or "").strip()
            naziv_short  = trunc_ellipsis(str(it.get("naziv") or ""), "DejaVu", 7.5, naziv_max_pt)
            komentar_short = trunc_ellipsis(komentar_raw, "DejaVu", 7.5, kom_max_pt) if komentar_raw else ""
            table_data.append([
                Paragraph(str(i), s_num),
                Paragraph(str(it.get("sku") or ""), s_sku),
                Paragraph(naziv_short, s_cell),
                Paragraph(str(it.get("pozicija") or "—"), s_cell),
                Paragraph(komentar_short, s_kom) if komentar_short else Paragraph("", s_cell),
                Paragraph("", s_cell),
            ])

        row_heights = [HDR_H] + [ROW_H] * (len(table_data) - 1)
        tbl = Table(table_data, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
        row_styles = [
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#1e293b")),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
            ("LINEBELOW",     (0,0), (-1,0), 1.5, colors.HexColor("#1e293b")),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 5),
            ("RIGHTPADDING",  (0,0), (-1,-1), 5),
            ("ALIGN",         (0,0), (0,-1), "CENTER"),
            ("ALIGN",         (3,1), (3,-1), "CENTER"),
            ("BOX",           (5,1), (5,-1), 0.8, colors.HexColor("#94a3b8")),
            ("BACKGROUND",    (5,1), (5,-1), colors.HexColor("#f0fdf4")),
        ]
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                row_styles.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#f8fafc")))
        tbl.setStyle(TableStyle(row_styles))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(
            "Navodilo: V stolpec 'Fizično ✓' vpišite dejansko stanje zaloge. "
            "Prazno = ni pregledano.  ✓ = potrjeno.  0 = ni na zalogi.",
            ParagraphStyle("f", fontSize=7, fontName="DejaVu", textColor=colors.HexColor("#94a3b8"))
        ))
        doc.build(story)

        inventura_cleanup()
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base = filename_hint.replace(".pdf", "")
        save_name = f"{base}_{ts}.pdf"
        buf.seek(0)
        (INVENTURA_DIR / save_name).write_bytes(buf.read())
        (INVENTURA_DIR / save_name.replace(".pdf", ".json")).write_text(
            json.dumps({"filename": save_name, "datum": datum, "items": items}, ensure_ascii=False),
            encoding="utf-8"
        )

        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={save_name}"})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/inventura-history")
async def inventura_history():
    """Vrne seznam shranjenih inventurnih PDF-jev."""
    inventura_cleanup()
    items = []
    try:
        for f in sorted(INVENTURA_DIR.glob("*.pdf"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = f.stat()
            items.append({
                "filename": f.name,
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    except Exception as e:
        print(f"[inventura] history err: {e}")
    return {"items": items[:50]}


@app.get("/inventura-history-download/{filename}")
async def inventura_history_download(filename: str):
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    f = INVENTURA_DIR / filename
    if not f.exists():
        return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)
    return FileResponse(str(f), filename=filename, media_type="application/pdf")


@app.get("/inventura-history-load/{filename}")
async def inventura_history_load(filename: str):
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    json_name = filename.replace(".pdf", ".json")
    f = INVENTURA_DIR / json_name
    if not f.exists():
        return JSONResponse({"error": "Podatki niso na voljo (star zapis)."}, status_code=404)
    return json.loads(f.read_text(encoding="utf-8"))


# ─── ODPREMA / AI ADDRESS VALIDATION ─────────────────────────────────────────

GEOAPIFY_KEY = os.environ.get("GEOAPIFY_API_KEY", "")
GOOGLE_MAPS_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "")
ODPREMA_DIR = DATA_DIR / "odprema"
ODPREMA_DIR.mkdir(exist_ok=True, parents=True)


@app.post("/odprema-google-validate")
async def odprema_google_validate(data: dict):
    """Google Address Validation API za problematične naslove."""
    if not GOOGLE_MAPS_KEY:
        return JSONResponse({"error": "GOOGLE_MAPS_API_KEY ni nastavljen."}, status_code=500)

    street = (data.get("street") or "").strip()
    city = (data.get("city") or "").strip()
    zip_code = (data.get("zip") or "").strip()
    order = data.get("order", "")

    address_line = f"{street}, {city}, {zip_code}, Bulgaria"

    payload = {
        "address": {
            "addressLines": [address_line],
            "regionCode": "BG",
            "languageCode": "bg",  # cirilica - Econt sprejme oboje
        },
        "enableUspsCass": False,
    }

    try:
        async with httpx.AsyncClient(timeout=10.0) as hc:
            resp = await hc.post(
                f"https://addressvalidation.googleapis.com/v1:validateAddress?key={GOOGLE_MAPS_KEY}",
                json=payload,
            )
            resp.raise_for_status()
            result = resp.json()

        verdict = result.get("result", {}).get("verdict", {})
        address = result.get("result", {}).get("address", {})
        components = address.get("addressComponents", [])

        # Izvleci komponente
        def get_comp(comp_type):
            for c in components:
                if comp_type in c.get("componentType", ""):
                    return c.get("componentName", {}).get("text", "")
            return ""

        fix_street = ""
        route = get_comp("route")
        street_nr = get_comp("street_number")
        subpremise = get_comp("subpremise")
        if route:
            fix_street = route
            if street_nr:
                fix_street += f" {street_nr}"
            if subpremise:
                fix_street += f", ap. {subpremise}"

        fix_city = get_comp("locality") or get_comp("administrative_area_level_2") or city
        fix_zip = get_comp("postal_code") or zip_code

        # Mesto ohranimo v latinici (original iz naše baze, ne Google cirilica)
        fix_city = city  # vedno originalno latinično ime mesta
        formatted = address.get("formattedAddress", "")

        validation_granularity = verdict.get("validationGranularity", "")
        has_unconfirmed = verdict.get("hasUnconfirmedComponents", False)

        if validation_granularity in ("PREMISE", "SUB_PREMISE", "ROUTE"):
            status = "CONFIRMED"
        elif validation_granularity in ("BLOCK", "PREMISE_PROXIMITY"):
            status = "PARTIALLY_CONFIRMED"
        else:
            status = "NOT_CONFIRMED"

        return {
            "order": order,
            "status": status,
            "fix_street": fix_street or street,
            "fix_city": fix_city,
            "fix_zip": fix_zip,
            "formatted": formatted,
            "granularity": validation_granularity,
            "has_unconfirmed": has_unconfirmed,
            "note": f"Google: {validation_granularity}",
        }

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


def odprema_cleanup():
    """Zbriše zapise starejše od 90 dni."""
    try:
        cutoff = datetime.now().timestamp() - (90 * 86400)
        for f in ODPREMA_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[odprema] cleanup err: {e}")


@app.post("/odprema-save")
async def odprema_save(data: dict):
    """Shrani batch podatke pošiljk na disk (90 dni)."""
    try:
        odprema_cleanup()
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"odprema_{ts}.json"
        payload = {
            "saved_at": datetime.now().isoformat(),
            "filename": data.get("filename", ""),
            "total": data.get("total", 0),
            "rows": data.get("rows", []),
            "validation": data.get("validation", {}),
        }
        (ODPREMA_DIR / filename).write_text(
            json.dumps(payload, ensure_ascii=False), encoding="utf-8"
        )
        return {"ok": True, "filename": filename}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/odprema-history")
async def odprema_history_list():
    """Vrne seznam shranjenih batch zapisov (90 dni)."""
    odprema_cleanup()
    items = []
    try:
        for f in sorted(ODPREMA_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "saved_at": data.get("saved_at", ""),
                    "original_file": data.get("filename", ""),
                    "total": data.get("total", 0),
                    "size": f.stat().st_size,
                })
            except:
                pass
    except Exception as e:
        print(f"[odprema] history err: {e}")
    return {"items": items}


@app.get("/odprema-history-load/{filename}")
async def odprema_history_load(filename: str):
    """Naloži shranjeni batch zapis."""
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    f = ODPREMA_DIR / filename
    if not f.exists():
        return JSONResponse({"error": "Ni najdeno."}, status_code=404)
    return json.loads(f.read_text(encoding="utf-8"))


@app.delete("/odprema-history-delete/{filename}")
async def odprema_history_delete(filename: str):
    """Zbriše shranjeni batch zapis."""
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    f = ODPREMA_DIR / filename
    if f.exists():
        f.unlink()
    return {"ok": True}


@app.get("/odprema-test")
async def odprema_test():
    """Debug: preveri Geoapify API ključ in naredi en testni klic."""
    if not GEOAPIFY_KEY:
        return {"ok": False, "error": "GEOAPIFY_API_KEY ni nastavljen", "key_len": 0}
    try:
        async with httpx.AsyncClient(timeout=10.0) as hc:
            resp = await hc.get(
                "https://api.geoapify.com/v1/geocode/search",
                params={"text": "Sofia, Bulgaria", "limit": 1, "apiKey": GEOAPIFY_KEY, "format": "json"}
            )
            return {
                "ok": resp.status_code == 200,
                "status_code": resp.status_code,
                "key_prefix": GEOAPIFY_KEY[:8] + "...",
                "response_preview": resp.text[:200]
            }
    except Exception as e:
        return {"ok": False, "error": str(e), "key_prefix": GEOAPIFY_KEY[:8] + "..."}


@app.post("/odprema-validate")
async def odprema_validate(data: dict):
    """
    Normalizira BG naslove z Claude AI — razčleni, popravi format, standardizira za Econt.
    """
    addresses = data.get("addresses", [])
    if not addresses:
        return JSONResponse({"error": "Ni naslovov za validacijo."}, status_code=400)

    async def normalize_one(addr: dict) -> dict:
        order = addr.get("order", "")
        zip_code = (addr.get("zip") or "").strip()
        city = (addr.get("city") or "").strip()
        street = (addr.get("street") or "").strip()
        street_nr = (addr.get("streetNr") or "").strip()

        # ── Geo lookup — poišči mesto in ulice iz Econt baze ──────────────────
        city_entry = econt_lookup_city(zip_code, city)
        streets_context = ""
        city_info = ""
        if city_entry:
            cid = city_entry.get("id")
            name_bg = city_entry.get("name_bg", "")
            name_en = city_entry.get("name_en", "")
            city_zip = city_entry.get("zip", "")
            city_info = f"\nVERIFICIRANO MESTO (Econt baza): {name_en} / {name_bg}, ZIP={city_zip} (cityID={cid})"
            streets_context = econt_get_streets_context(cid, street)
            if streets_context:
                streets_context = f"\nULICE TEGA MESTA (Econt baza — predlagaj SAMO iz tega seznama):\n{streets_context}"
        else:
            city_info = "\n⚠ Mesto ni najdeno v Econt bazi — previdno z validacijo."

        prompt = f"""Si ekspert za bolgarske poštne naslove za Econt Express dostavo. Normaliziraj vhodni naslov v standardni format ZA ECONT VMESNIK (latinica).

VHODNI NASLOV:
- Ulica/naslov: {street}
- Hišna številka: {street_nr}
- Mesto: {city}
- ZIP: {zip_code}
{city_info}{streets_context}

BOLGARSKE OKRAJŠAVE:
- ul. = ulitsa (ulica), bul. = bulevard
- zh.k. / jk / kv. / zk. / z.k. = zhilishten kompleks (stanovanjska četrt)
- bl. = blok, vh. = vhod, et. = etazh (nadstropje), ap. = apartament
- s. / selo = vas, gr. / grad = mesto
- ofis ekont / ekont / paketomat = Econt pisarna

KRITIČNA PRAVILA (po prioriteti):

1. LATINICA OBVEZNA — vse vrni v latinici, tudi ulice (Econt vmesnik je v latinici)
   **IZJEMA: fix_city VEDNO v latinici** — nikoli ne prevajaj ali transliteriraj imena mesta v cirilico, tudi če Google ali drug vir vrne cirilico. Mesto ohrani točno tako kot je napisano v originalnem naslovu ali v standardni latinični obliki.

2. HIŠNA ŠTEVILKA — NIKOLI NE BRIŠI!
   Hišno številko VEDNO ohrani v fix_street — stranka jo je vpisala in ve kje stanuje.
   Odstrani SAMO če je očiten placeholder: "nn", "NN", "N/A", "n/a" — to so sistemske vrednosti brez pomena.
   VSE ostalo ohrani: številke, črke, duplikate (65 65), 0, 36b, 4A itd.
   Če je hišna številka v polju streetNr, jo dodaj na konec fix_street.
   Primer: ulica="Vasil Levski", streetNr="36b" → fix_street="ul. Vasil Levski 36b"

3. ULICA IN MESTO — POMEMBNO!
   Če imaš seznam ulic tega mesta (zgoraj), ga uporabi za validacijo in popravke pravopisnih napak.
   NIKOLI ne predlagaj ulice iz drugega mesta.
   AMPAK: Econt baza ulic ni popolna — manjkajo nove ulice, manjše vasi, četrti brez ulic.
   Če ulica ni v seznamu, jo VSEENO potrdi (status FIXED ali OK) če je naslov smiselno formatiran.
   Status UNCLEAR nastavi SAMO če naslov je res nerazumljiv ali manjka ulica/hišna številka — NE samo zato ker ulice ni v bazi.

4. ZIP EKSTRAKCIJA — če je ZIP v polju ulice ali mesta, ga prestavi v fix_zip. Popravi ZIP SAMO če je očitno napačen (vsebuje črke, ni 4 cifre, ali je ZIP drugega mesta). Če si negotov → pusti originalni ZIP.

5. SOFIJSKE ČETRTI IN ZIP — poznaj pravilne ZIP-e za sofijske četrti:
   zh.k. Lulin → 1343, zh.k. Mladost → 1750/1784, zh.k. Lyulin → 1343
   zh.k. Druzhba → 1582, zh.k. Nadezhda → 1220, zh.k. Ovcha Kupel → 1618
   zh.k. Bukston/Bakston → 1618, zh.k. Borovo → 1680, zh.k. Lozenets → 1164
   zh.k. Dianabad → 1172, zh.k. Manastirski livadi → 1404
   Če stranka piše ZIP 1000 za četrt → popravi na pravilen ZIP četrti!

6. ULICA = IME MESTA → ni ulice
   Če je ime ulice enako imenu mesta (npr. "ul. Kardzhali" v mestu Kardzhali) → fix_street=""
   Če je naslov samo ime mesta brez ulice (npr. "Cerven bryag 3") → fix_street="", status UNCLEAR

7. ULICA V VEČ ČETRTIH → opozori
   Če ulica obstaja v več četrtih Sofije → status UNCLEAR, note mora vsebovati "Quarter needed: X, Y, Z"
   Znani primeri: ul. Elin Pelin (Lozenets/Dragalevtsi/Pancharevo)

8. ECONT OFFICE V LATINICI
   Vse Econt office naslove vrni v latinici:
   "Ofis Ekont" → "Econt office [mesto]"
   Primer: "Ekont Bokar, zh.k. Manastirski livadi" → "Econt office Bokar, zh.k. Manastirski livadi"

9. IME MESTA V ULICI → odstrani
   "Sofia bul. Bulgaria 102" → fix_street="bul. Bulgaria 102", fix_city="Sofia"
   "Bansko Glazne 6" → fix_street="ul. Glazne 6", fix_city="Bansko"

10. PODVOJENI DELI NASLOVA → ohrani, ne briši
   Podvojena hišna številka (npr. "65 65", "9 9") je verjetno resnična — ohrani kot je.
   Odstrani duplikat SAMO pri okrajšavah formata (npr. "bl. 5 bl. 5" → "bl. 5").
   "Bl. 503 vh.A ap 65 et 11" → "bl. 503, vh. A, et. 11, ap. 65"

11. TIPKARSKE NAPAKE V IMENIH MEST:
   Vitosa → Vitosha, Blgaria → Bulgaria, Sofiq → Sofia
   Plovdic → Plovdiv, Kustendil → Kyustendil, Vraca → Vratsa
   Carevo → Tsarevo, Dupnica → Dupnitsa, Krdzali → Kardzhali
   Trstenik → Trastenik, Satovca → Satovcha

12. MALE VASI BREZ ULICE → status UNCLEAR z opombo
    Če je naslov samo ime vasi brez ulice in hišne številke → UNCLEAR, note="No street - recommend Econt office [najbližje mesto]"

PRIMERI (few-shot):
Input: ulica="Sofia bul.Vitosa 38", mesto="Sofia", ZIP="1000"
Output: {{"status":"FIXED","fix_street":"bul. Vitosha 38","fix_city":"Sofia","fix_zip":"1000","note":"Removed city from street, fixed Vitosa→Vitosha"}}

Input: ulica="Vasil Levski", streetNr="36b", mesto="Starcevo", ZIP="4987"
Output: {{"status":"FIXED","fix_street":"ul. Vasil Levski 36b","fix_city":"Presoka","fix_zip":"4987","note":"Added ul. prefix, kept house number 36b, corrected city name"}}

Input: ulica="Zk.Borovo bl.5 vh B ET.6 AP.34 34", mesto="Sofia", ZIP="1000"
Output: {{"status":"FIXED","fix_street":"zh.k. Borovo, bl. 5, vh. B, et. 6, ap. 34","fix_city":"Sofia","fix_zip":"1680","note":"Fixed format, removed duplicate 34, corrected ZIP for Borovo"}}

Input: ulica="Lulin5 bl540vhb 65etz11", mesto="Sofia", ZIP="1000"
Output: {{"status":"FIXED","fix_street":"zh.k. Lulin 5, bl. 540, vh. B, et. 11, ap. 65","fix_city":"Sofia","fix_zip":"1343","note":"Parsed compressed format, corrected ZIP for Lulin"}}

Input: ulica="ul. Elin Pelin 18", mesto="Sofia", ZIP="1000"
Output: {{"status":"UNCLEAR","fix_street":"ul. Elin Pelin 18","fix_city":"Sofia","fix_zip":"1000","note":"Quarter needed: Lozenets (1164), Dragalevtsi (1415), Pancharevo (1137)"}}

Input: ulica="Cerven brag 3", mesto="Cerven bryag", ZIP="5980"
Output: {{"status":"UNCLEAR","fix_street":"","fix_city":"Cherven Bryag","fix_zip":"5980","note":"No street provided - recommend Econt office Cherven Bryag, ul. Hristo Botev 2"}}

Input: ulica="Sofia Manastirski livadi Ekont Bokar", mesto="Sofia", ZIP="1000"
Output: {{"status":"ECONT_OFFICE","fix_street":"Econt office Bokar, zh.k. Manastirski livadi","fix_city":"Sofia","fix_zip":"1404","note":"Econt office Bokar in Manastirski livadi"}}

Vrni SAMO JSON, brez razlag:
{{"status": "OK|FIXED|ECONT_OFFICE|UNCLEAR", "fix_street": "...", "fix_city": "...", "fix_zip": "4-mestni ZIP", "note": "..."}}"""

        loop = asyncio.get_event_loop()
        try:
            msg = await loop.run_in_executor(None, lambda: client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=400,
                messages=[{"role": "user", "content": prompt}]
            ))
            text = msg.content[0].text.strip()
            text = re.sub(r'```json\s*', '', text)
            text = re.sub(r'```\s*', '', text).strip()
            result = json.loads(text)
            
            status_map = {
                "OK": "CONFIRMED",
                "FIXED": "PARTIALLY_CONFIRMED", 
                "ECONT_OFFICE": "ECONT_DEPO",
                "UNCLEAR": "NOT_CONFIRMED"
            }
            
            fix_street_raw = result.get("fix_street", street)

            # Ekstrahiraj hišno številko iz fix_street če je streetNr nn/N/A
            fix_nr_out = street_nr
            is_placeholder = street_nr.lower().strip() in ("nn", "n/a", "")
            if is_placeholder and fix_street_raw:
                if not re.search(r'econt\s*office', fix_street_raw, re.IGNORECASE):
                    nr_match = re.search(r'^(.*?)[\s,]+(\d+[A-Za-z]?(?:\s+\d+[A-Za-z]?)?)$', fix_street_raw.strip())
                    if nr_match:
                        fix_nr_out = nr_match.group(2).strip()
                        fix_street_raw = nr_match.group(1).strip().rstrip(',')

            # Fallback: če fix_nr še vedno nn, poišči številko v ORIGINALNEM naslovu
            # (AI jo je morda pozabil prenesti v fix_street)
            if fix_nr_out.lower().strip() in ("nn", "n/a", "") and street:
                if not re.search(r'econt\s*office', fix_street_raw, re.IGNORECASE):
                    nr_fallback = re.search(r'\b(\d+[A-Za-z]?)\s*$', street.strip())
                    if nr_fallback:
                        fix_nr_out = nr_fallback.group(1)

            return {
                "order": order,
                "status": status_map.get(result.get("status", "UNCLEAR"), "NOT_CONFIRMED"),
                "confidence": 1.0 if result.get("status") == "OK" else 0.7 if result.get("status") == "FIXED" else 0.3,
                "formatted": f"{fix_street_raw}, {result.get('fix_city','')}, {result.get('fix_zip','')}",
                "fix_street": fix_street_raw,
                "fix_nr": fix_nr_out,
                "fix_city": result.get("fix_city", city),
                "fix_zip": result.get("fix_zip", zip_code),
                "note": result.get("note", ""),
                "original": {"zip": zip_code, "city": city, "street": street, "streetNr": street_nr},
            }
        except Exception as e:
            print(f"[odprema] AI error for {order}: {e}")
            return {
                "order": order, "status": "ERROR", "confidence": 0,
                "formatted": "", "fix_street": street, "fix_nr": street_nr,
                "fix_city": city, "fix_zip": zip_code,
                "error": str(e),
                "original": {"zip": zip_code, "city": city, "street": street, "streetNr": street_nr},
            }

    # Vzporedno — max 5 hkrati (Claude rate limit)
    semaphore = asyncio.Semaphore(5)

    async def limited(addr):
        async with semaphore:
            return await normalize_one(addr)

    results = await asyncio.gather(*[limited(addr) for addr in addresses])
    return {"results": list(results), "total": len(results)}

# ─── ECONT OFFICES CACHE ──────────────────────────────────────────────────────

ECONT_OFFICES_CACHE: list = []
ECONT_OFFICES_LAST_FETCH: float = 0
ECONT_OFFICES_TTL = 86400  # 24h cache

ECONT_API_URL = "https://ee.econt.com/services"
ECONT_DEMO_URL = "https://demo.econt.com/ee/services"
ECONT_USER = os.environ.get("ECONT_USER", "iasp-dev")
ECONT_PASS = os.environ.get("ECONT_PASS", "1Asp-dev")


async def econt_fetch_offices() -> list:
    """Pobere vse BG Econt office-e z API-ja, cachira 24h."""
    global ECONT_OFFICES_CACHE, ECONT_OFFICES_LAST_FETCH
    import time

    now = time.time()
    if ECONT_OFFICES_CACHE and (now - ECONT_OFFICES_LAST_FETCH) < ECONT_OFFICES_TTL:
        return ECONT_OFFICES_CACHE

    # Poskusi produkcijo, potem demo
    for base_url in [ECONT_API_URL, ECONT_DEMO_URL]:
        try:
            async with httpx.AsyncClient(timeout=20.0) as hc:
                resp = await hc.post(
                    f"{base_url}/Nomenclatures/NomenclaturesService.getOffices.json",
                    json={"countryCode": "BG"},
                    auth=(ECONT_USER, ECONT_PASS)
                )
                if resp.status_code == 200:
                    data = resp.json()
                    offices = data.get("offices", [])
                    if offices:
                        ECONT_OFFICES_CACHE = offices
                        ECONT_OFFICES_LAST_FETCH = now
                        print(f"[econt] Naloženih {len(offices)} BG officeov iz {base_url}")
                        return offices
        except Exception as e:
            print(f"[econt] Napaka pri {base_url}: {e}")

    return ECONT_OFFICES_CACHE  # Vrni star cache če API ne dela


def econt_find_nearest_office(offices: list, zip_code: str, city_name: str) -> dict | None:
    """Poišče najbližji Econt office glede na ZIP kodo ali ime mesta."""
    if not offices:
        return None

    zip_clean = (zip_code or "").strip()
    city_lower = (city_name or "").lower().strip()

    # 1. Točno ujemanje ZIP
    for o in offices:
        addr = o.get("address") or {}
        city = addr.get("city") or {}
        office_zip = str(city.get("postCode") or "").strip()
        if office_zip and office_zip == zip_clean:
            return o

    # 2. Ujemanje mesta (case-insensitive)
    if city_lower:
        for o in offices:
            addr = o.get("address") or {}
            city = addr.get("city") or {}
            office_city = (city.get("name") or "").lower().strip()
            office_city_en = (city.get("nameEn") or "").lower().strip()
            if city_lower in office_city or city_lower in office_city_en:
                return o

    # 3. ZIP prefix match (prvih 2 cifri = regija)
    if len(zip_clean) >= 2:
        prefix = zip_clean[:2]
        for o in offices:
            addr = o.get("address") or {}
            city = addr.get("city") or {}
            office_zip = str(city.get("postCode") or "").strip()
            if office_zip.startswith(prefix):
                return o

    return None


def econt_office_to_address(office: dict) -> dict:
    """Pretvori Econt office objekt v naslovne polje."""
    addr = office.get("address") or {}
    city = addr.get("city") or {}
    return {
        "name": office.get("name") or office.get("nameEn") or "",
        "street": addr.get("street") or addr.get("fullAddress") or "",
        "num": addr.get("num") or "",
        "city": city.get("name") or city.get("nameEn") or "",
        "zip": str(city.get("postCode") or ""),
        "full": f"{office.get('name','')} — {addr.get('fullAddress') or addr.get('street','')} {addr.get('num','')}, {city.get('name','')}".strip(),
    }


@app.get("/econt-offices")
async def get_econt_offices():
    """Vrne seznam vseh BG Econt officeov (cachiran 24h)."""
    offices = await econt_fetch_offices()
    return {"ok": True, "total": len(offices), "offices": offices}


@app.get("/econt-valid-zips")
async def get_econt_valid_zips():
    """Vrne set vseh veljavnih ZIP kod iz Econt geo baze. ZIP ki NI tukaj = suspended."""
    if not ECONT_GEO:
        return {"ok": False, "error": "econt_geo.json not loaded", "zips": []}
    zips = list(ECONT_GEO.get("zip_to_city_id", {}).keys())
    return {"ok": True, "total": len(zips), "zips": zips}


@app.post("/econt-nearest-office")
async def get_nearest_office(data: dict):
    """Za dani ZIP/mesto vrne najbližji Econt office z naslovom."""
    zip_code = data.get("zip", "")
    city = data.get("city", "")

    offices = await econt_fetch_offices()
    if not offices:
        return JSONResponse({"error": "Econt API ni dosegljiv, ni officeov v cacheju."}, status_code=503)

    office = econt_find_nearest_office(offices, zip_code, city)
    if not office:
        return {"ok": False, "office": None, "message": "Ni najden office za ta ZIP/mesto"}

    return {"ok": True, "office": econt_office_to_address(office), "raw": office}


@app.post("/odprema-econt-check")
async def odprema_econt_check(data: dict):
    """
    Double-check naslova direktno pri Econt API — preveri ali mesto + ulica obstajata.
    Klic gre iz Render serverja (whitelist IP). 
    Input: {"zip": "1000", "city": "Sofia", "street": "bul. Vitosha 38"}
    """
    zip_code = (data.get("zip") or "").strip()
    city_name = (data.get("city") or "").strip()
    street = (data.get("street") or "").strip()
    order = data.get("order", "")

    result = {"order": order, "zip": zip_code, "city": city_name}

    async with httpx.AsyncClient(timeout=15.0) as hc:
        for base_url in [ECONT_API_URL, ECONT_DEMO_URL]:
            try:
                # 1. getCities — preveri ali mesto obstaja
                r = await hc.post(
                    f"{base_url}/Nomenclatures/NomenclaturesService.getCities.json",
                    json={"countryCode": "BG", "name": city_name},
                    auth=(ECONT_USER, ECONT_PASS)
                )
                if r.status_code != 200:
                    continue

                cities = r.json().get("cities", [])
                city_match = None
                for c in cities:
                    if str(c.get("postCode", "")) == zip_code or \
                       (c.get("nameEn", "").lower() == city_name.lower()) or \
                       (c.get("name", "").lower() == city_name.lower()):
                        city_match = c
                        break

                result["city_found"] = bool(city_match)
                result["city_econt"] = city_match.get("nameEn", "") if city_match else ""
                result["city_zip"] = str(city_match.get("postCode", "")) if city_match else ""

                if not city_match:
                    result["status"] = "CITY_NOT_FOUND"
                    result["note"] = f"Mesto '{city_name}' ni v Econt bazi"
                    return result

                # 2. getStreets — preveri ali ulica obstaja v tem mestu
                if street:
                    # Ekstrahiraj samo ime ulice brez številke
                    street_name = re.sub(r'\b\d+\b.*$', '', street).strip().rstrip(',').strip()
                    street_name = re.sub(r'^(ul\.|bul\.|zh\.k\.|kv\.)\s*', '', street_name, flags=re.IGNORECASE).strip()

                    r2 = await hc.post(
                        f"{base_url}/Nomenclatures/NomenclaturesService.getStreets.json",
                        json={"cityID": city_match.get("id"), "name": street_name},
                        auth=(ECONT_USER, ECONT_PASS)
                    )
                    if r2.status_code == 200:
                        streets = r2.json().get("streets", [])
                        result["street_found"] = bool(streets)
                        result["street_matches"] = [s.get("nameEn", s.get("name", "")) for s in streets[:3]]
                        if streets:
                            result["status"] = "OK"
                            result["note"] = f"Mesto in ulica potrjena pri Econt"
                        else:
                            result["status"] = "STREET_NOT_FOUND"
                            result["note"] = f"Ulica '{street_name}' ni v Econt bazi za {city_name}"
                    else:
                        result["street_found"] = None
                        result["status"] = "OK"
                        result["note"] = "Mesto potrjeno, ulica ni preverjena"
                else:
                    result["street_found"] = None
                    result["status"] = "OK"
                    result["note"] = "Mesto potrjeno (ni ulice za preverjanje)"

                return result

            except Exception as e:
                result["error"] = str(e)
                continue

    result["status"] = "API_UNAVAILABLE"
    result["note"] = "Econt API ni dosegljiv (IP whitelist?)"
    return result



async def odprema_resolve_suspended(data: dict):
    """
    Za seznam suspended naslovov poišče najbližji Econt office.
    Input: [{"order": "...", "zip": "...", "city": "..."}, ...]
    """
    addresses = data.get("addresses", [])
    if not addresses:
        return JSONResponse({"error": "Ni naslovov."}, status_code=400)

    offices = await econt_fetch_offices()

    results = []
    for addr in addresses:
        order = addr.get("order", "")
        zip_code = addr.get("zip", "")
        city = addr.get("city", "")

        office = econt_find_nearest_office(offices, zip_code, city)
        if office:
            office_addr = econt_office_to_address(office)
            results.append({
                "order": order,
                "found": True,
                "office_name": office_addr["name"],
                "office_street": office_addr["street"],
                "office_num": office_addr["num"],
                "office_city": office_addr["city"],
                "office_zip": office_addr["zip"],
                "office_full": office_addr["full"],
            })
        else:
            results.append({
                "order": order,
                "found": False,
                "office_name": "",
                "office_street": "",
                "office_num": "",
                "office_city": "",
                "office_zip": zip_code,
                "office_full": "Ni najden — preveri ročno",
            })

    return {"ok": True, "results": results, "offices_available": len(offices)}


# ─── EMAIL OBVESTILA ──────────────────────────────────────────────────────────

SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
SMTP_FROM = os.environ.get("SMTP_FROM", "")


@app.post("/odprema-send-emails")
async def odprema_send_emails(data: dict):
    """
    Pošlje email obvestila strankam s suspended ZIP naslovi.
    Input: { "shipments": [{ "order": "...", "name": "...", "email": "...", "office": "...", "orig_city": "..." }, ...] }
    """
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASS, SMTP_FROM]):
        return JSONResponse({"error": "SMTP ni konfiguriran."}, status_code=500)

    shipments = data.get("shipments", [])
    if not shipments:
        return JSONResponse({"error": "Ni pošiljk za obveščanje."}, status_code=400)

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    sent = []
    failed = []

    for s in shipments:
        email = (s.get("email") or "").strip()
        if not email or "@" not in email:
            failed.append({"order": s.get("order"), "reason": "Ni e-mail naslova"})
            continue

        name = s.get("name", "")
        order = s.get("order", "")
        office = s.get("office", "najbližji Econt office")
        orig_city = s.get("orig_city", "")

        # Email v bolgarščini + angleščini
        subject = f"Вашата поръчка {order} - промяна на адрес за доставка / Your order {order} - delivery address change"

        html = f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
  <img src="https://maaarket.si/wp-content/uploads/2023/03/logo.png" style="height:40px;margin-bottom:20px" alt="Maaarket">
  
  <h2 style="color:#1a1a2e">Уважаеми {name},</h2>
  
  <p>Вашата поръчка с номер <strong>{order}</strong> не може да бъде доставена до посочения адрес 
  (<strong>{orig_city}</strong>), тъй като куриерската фирма <strong>Econt Express</strong> временно е 
  преустановила доставките до вашия район.</p>
  
  <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:16px;margin:20px 0">
    <strong>📦 Вашата поръчка ще ви чака на:</strong><br><br>
    <span style="font-size:16px;color:#1a1a2e">{office}</span>
  </div>
  
  <p>Моля, посетете горепосочения офис с <strong>личен документ</strong> за получаване на поръчката.</p>
  
  <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
  
  <h3 style="color:#1a1a2e">Dear {name},</h3>
  
  <p>Your order <strong>{order}</strong> cannot be delivered to the specified address 
  (<strong>{orig_city}</strong>) as <strong>Econt Express</strong> has temporarily suspended 
  deliveries to your area.</p>
  
  <div style="background:#d1ecf1;border:1px solid #bee5eb;border-radius:8px;padding:16px;margin:20px 0">
    <strong>📦 Your order will be waiting at:</strong><br><br>
    <span style="font-size:16px;color:#1a1a2e">{office}</span>
  </div>
  
  <p>Please visit the above office with a <strong>valid ID</strong> to collect your order.</p>
  
  <p style="color:#666;font-size:12px;margin-top:30px">
    Maaarket.eu | За въпроси / For questions: <a href="mailto:info@maaarket.bg">info@maaarket.bg</a>
  </p>
</div>
"""

        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = SMTP_FROM
            msg["To"] = email
            msg["Reply-To"] = "info@maaarket.bg"
            msg.attach(MIMEText(html, "html", "utf-8"))

            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_FROM, email, msg.as_string())

            sent.append({"order": order, "email": email})
            print(f"[email] Sent to {email} for order {order}")

        except Exception as e:
            print(f"[email] Failed {email} for {order}: {e}")
            failed.append({"order": order, "email": email, "reason": str(e)})

    return {
        "ok": True,
        "sent": len(sent),
        "failed": len(failed),
        "sent_list": sent,
        "failed_list": failed,
    }


# ─── KAYAKO CLASSIC REST API — KB IMPORT ─────────────────────────────────────
import hashlib
import hmac
import base64
import random

KAYAKO_API_URL = os.environ.get("KAYAKO_API_URL", "https://support.silux.si/api/index.php")
KAYAKO_API_KEY = os.environ.get("KAYAKO_API_KEY", "")
KAYAKO_SECRET  = os.environ.get("KAYAKO_SECRET", "")

KAYAKO_DEPT = {
    "silux":    1,
    "maaarket": 27,
}

KB_FILES = {
    "silux":    DATA_DIR / "kb_silux.json",
    "maaarket": DATA_DIR / "kb_maaarket.json",
}

MACROS_FILE = Path("macros_maaarket.json")

def _kayako_build_url(path: str) -> str:
    """Zgradi popoln Kayako URL z auth - signature je quote_plus encoded."""
    from urllib.parse import quote_plus
    salt = str(random.randint(1000000000, 9999999999))
    raw_sig = hmac.new(
        key=KAYAKO_SECRET.encode("utf-8"),
        msg=salt.encode("utf-8"),
        digestmod=hashlib.sha256
    ).digest()
    signature = quote_plus(base64.b64encode(raw_sig).decode("utf-8"))
    return (
        f"{KAYAKO_API_URL}"
        f"?e={path}"
        f"&apikey={KAYAKO_API_KEY}"
        f"&salt={salt}"
        f"&signature={signature}"
    )

# backwards compat
def _kayako_url(path: str) -> str:
    return _kayako_build_url(path)

def _xml_text(el, tag: str, default: str = "") -> str:
    node = el.find(tag)
    return (node.text or default).strip() if node is not None and node.text else default

def _parse_ticket_xml(xml_text: str) -> list[dict]:
    """Razčleni XML odgovor za ticket/tickete → lista dict."""
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return []
    tickets = []
    for t in root.findall("ticket"):
        ticket_id  = t.get("id", "")
        subject    = _xml_text(t, "subject")
        dept_id    = _xml_text(t, "departmentid")
        status_id  = _xml_text(t, "statusid")
        created    = _xml_text(t, "creationtime")
        email      = _xml_text(t, "email")
        fullname   = _xml_text(t, "fullname")
        # posti (konverzacija)
        posts = []
        for p in t.findall(".//post"):
            creator   = p.get("creator", _xml_text(p, "creator", "2"))
            staff_id  = _xml_text(p, "staffid", "0")
            p_fullname  = _xml_text(p, "fullname")
            contents_node = p.find("contents")
            contents = ""
            if contents_node is not None:
                contents = (contents_node.text or "").strip()
            # creator=1 = staff, creator=2 = user/stranka
            role = "staff" if (staff_id != "0" or str(creator) == "1") else "customer"
            if contents:
                posts.append({"role": role, "name": p_fullname, "text": contents})
        tickets.append({
            "id": ticket_id,
            "subject": subject,
            "dept_id": dept_id,
            "status_id": status_id,
            "created": created,
            "email": email,
            "fullname": fullname,
            "posts": posts,
        })
    return tickets

async def _fetch_tickets_batch(client_h: httpx.AsyncClient, dept_id: int, start: int, count: int = 50) -> list[dict]:
    """Potegne batch ticketov iz Kayako (samo header info, brez postov)."""
    path = f"/Tickets/Ticket/ListAll/{dept_id}/3/-1/-1/{count}/{start}/ticketid/DESC"
    url = _kayako_build_url(path)
    try:
        r = await client_h.get(url, timeout=30)
        print(f"[kayako] ListAll HTTP {r.status_code} | {url[:80]}")
        if r.status_code != 200:
            return []
        return _parse_ticket_xml(r.text)
    except Exception as e:
        print(f"[kayako] ListAll error: {e}")
        return []

async def _fetch_ticket_posts(client_h: httpx.AsyncClient, ticket_id: str) -> list[dict]:
    """Potegne posamezen ticket z vsemi posti."""
    path = f"/Tickets/Ticket/{ticket_id}"
    url = _kayako_build_url(path)
    try:
        r = await client_h.get(url, timeout=20)
        if r.status_code != 200:
            return []
        tickets = _parse_ticket_xml(r.text)
        return tickets[0]["posts"] if tickets else []
    except Exception as e:
        print(f"[kayako] Ticket {ticket_id} error: {e}")
        return []

def _tickets_to_kb(tickets_with_posts: list[dict]) -> dict:
    """
    Pretvori surove tickete v knowledge base format.
    Shrani samo tickete ki imajo vsaj 1 staff odgovor.
    Format: { "qa_pairs": [ {subject, question, answer, count:1} ] }
    """
    qa_pairs = []
    for t in tickets_with_posts:
        subject = t.get("subject", "")
        posts   = t.get("posts", [])
        if not posts:
            continue
        # Združi customer posti v eno vprašanje, staff posti v en odgovor
        customer_texts = [p["text"] for p in posts if p["role"] == "customer"]
        staff_texts    = [p["text"] for p in posts if p["role"] == "staff"]
        if not customer_texts or not staff_texts:
            continue
        question = " | ".join(customer_texts[:3])[:800]   # max 800 znakov
        answer   = staff_texts[-1][:800]                   # zadnji staff odgovor, max 800
        qa_pairs.append({
            "subject":  subject[:150],
            "question": question,
            "answer":   answer,
            "count":    1,
        })
    return {"qa_pairs": qa_pairs, "updated": datetime.now(timezone.utc).isoformat()}

# ─── MAKRI (KAYAKO MACROS) ────────────────────────────────────────────────────

@app.get("/macros")
async def get_macros():
    """Vrne vse makre iz /data/macros_maaarket.json"""
    if MACROS_FILE.exists():
        try:
            return json.loads(MACROS_FILE.read_text(encoding="utf-8"))
        except:
            pass
    return {"macros": [], "updated": ""}

@app.post("/macros-save")
async def save_macros(data: dict):
    """Shrani makre (za admin update)"""
    try:
        MACROS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "total": len(data.get("macros", []))}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/kayako-test")
async def kayako_test(brand: str = "silux"):
    """Test Kayako povezave — potegne samo 3 tickete da preverimo auth."""
    if not KAYAKO_API_KEY or not KAYAKO_SECRET:
        return {"ok": False, "error": "KAYAKO_API_KEY ali KAYAKO_SECRET nista nastavljena!"}
    dept_id = KAYAKO_DEPT.get(brand, 1)
    # Debug — pokaži točen URL ki ga kličemo
    path = f"/Tickets/Ticket/ListAll/{dept_id}/3/-1/-1/3/0/ticketid/DESC"
    debug_url = _kayako_build_url(path)
    print(f"[kayako] DEBUG URL: {debug_url}")
    async with httpx.AsyncClient() as h:
        tickets = await _fetch_tickets_batch(h, dept_id, start=0, count=3)
    if not tickets:
        return {
            "ok": False,
            "error": "Ni ticketov ali napaka pri povezavi",
            "debug_url_base": f"{KAYAKO_API_URL}?e={path}",
            "kayako_api_url_env": KAYAKO_API_URL,
        }
    return {
        "ok": True,
        "brand": brand,
        "dept_id": dept_id,
        "tickets_found": len(tickets),
        "sample": [{"id": t["id"], "subject": t["subject"], "posts_count": len(t["posts"])} for t in tickets],
    }

@app.get("/kayako-import")
async def kayako_import(
    brand: str = "maaarket",
    max_tickets: int = 500,
    batch_size: int = 50,
):
    """
    Importa tickete iz Kayako → shrani v /data/kb_{brand}.json
    Parametri:
      brand       = maaarket | silux
      max_tickets = koliko ticketov max (default 500)
      batch_size  = po koliko naenkrat (default 50, max 100)
    """
    if not KAYAKO_API_KEY or not KAYAKO_SECRET:
        return {"ok": False, "error": "KAYAKO_API_KEY ali KAYAKO_SECRET nista nastavljena v Render env vars!"}
    if brand not in KAYAKO_DEPT:
        return {"ok": False, "error": f"Neznan brand: {brand}"}

    dept_id    = KAYAKO_DEPT[brand]
    kb_file    = KB_FILES[brand]
    batch_size = min(batch_size, 100)

    # Naloži obstoječo KB če obstaja
    existing_kb = {"qa_pairs": [], "updated": ""}
    if kb_file.exists():
        try:
            existing_kb = json.loads(kb_file.read_text(encoding="utf-8"))
        except:
            pass
    existing_ids = {qa.get("ticket_id", "") for qa in existing_kb.get("qa_pairs", [])}

    all_tickets = []
    start = 0
    print(f"[kayako] Začenjam import za {brand} (dept={dept_id}, max={max_tickets})")

    async with httpx.AsyncClient() as h:
        # Faza 1: potegni liste ticketov
        while start < max_tickets:
            batch = await _fetch_tickets_batch(h, dept_id, start=start, count=batch_size)
            if not batch:
                break
            all_tickets.extend(batch)
            print(f"[kayako] ListAll: {start}–{start+len(batch)} ({len(batch)} ticketov)")
            if len(batch) < batch_size:
                break  # zadnja stran
            start += batch_size
            await asyncio.sleep(0.2)  # rate limit

        # Faza 2: za vsak ticket potegni posti (samo novi)
        new_count = 0
        for t in all_tickets:
            tid = t["id"]
            if tid in existing_ids:
                continue  # že imamo
            posts = await _fetch_ticket_posts(h, tid)
            t["posts"] = posts
            await asyncio.sleep(0.1)  # rate limit
            new_count += 1
            if new_count % 50 == 0:
                print(f"[kayako] Posti: {new_count}/{len(all_tickets)}")

    # Faza 3: pretvori v KB format
    new_qa = _tickets_to_kb([t for t in all_tickets if t["id"] not in existing_ids])

    # Dodaj ticket_id za deduplication
    for i, qa in enumerate(new_qa["qa_pairs"]):
        if i < len(all_tickets):
            qa["ticket_id"] = all_tickets[i]["id"]

    # Združi z obstoječim
    merged = existing_kb.get("qa_pairs", []) + new_qa["qa_pairs"]
    # Dedupliciraj po subject+question
    seen = set()
    deduped = []
    for qa in merged:
        key = qa["subject"] + "|" + qa["question"][:100]
        if key not in seen:
            seen.add(key)
            deduped.append(qa)

    final_kb = {
        "qa_pairs":   deduped,
        "updated":    datetime.now(timezone.utc).isoformat(),
        "brand":      brand,
        "total":      len(deduped),
    }
    kb_file.write_text(json.dumps(final_kb, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[kayako] Import končan: {len(new_qa['qa_pairs'])} novih, {len(deduped)} skupaj")
    return {
        "ok":         True,
        "brand":      brand,
        "new":        len(new_qa["qa_pairs"]),
        "total":      len(deduped),
        "skipped":    len(all_tickets) - new_count,
        "kb_file":    str(kb_file),
    }

@app.get("/kayako-kb-stats")
async def kayako_kb_stats(brand: str = "maaarket"):
    """Vrne statistiko knowledge base za brand."""
    kb_file = KB_FILES.get(brand)
    if not kb_file or not kb_file.exists():
        return {"ok": True, "brand": brand, "total": 0, "updated": None, "sample": []}
    try:
        kb = json.loads(kb_file.read_text(encoding="utf-8"))
        pairs = kb.get("qa_pairs", [])
        return {
            "ok":      True,
            "brand":   brand,
            "total":   len(pairs),
            "updated": kb.get("updated"),
            "sample":  pairs[:5],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ─── BETA ANALIZA — META ADS EXPORT HISTORY ──────────────────────────────────
BETA_DIR = DATA_DIR / "beta_exports"

@app.post("/beta-save-export")
async def beta_save_export(data: dict):
    """Shrani Meta Ads CSV export na disk z datumom."""
    try:
        BETA_DIR.mkdir(exist_ok=True)
        from datetime import datetime
        import pytz
        lj = pytz.timezone("Europe/Ljubljana")
        now = datetime.now(lj)
        date_str = now.strftime("%Y-%m-%d")
        ts_str = now.strftime("%Y-%m-%d_%H-%M")
        
        # Meta info
        filename = f"{ts_str}_{data.get('filename','export').replace(' ','_')[:40]}.json"
        payload = {
            "filename": data.get("filename", ""),
            "date": date_str,
            "timestamp": ts_str,
            "campaigns": data.get("campaigns", []),
            "total_spend": data.get("total_spend", 0),
            "total_purchases": data.get("total_purchases", 0),
        }
        (BETA_DIR / filename).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        
        # Pobriši stare (> 30 dni)
        cutoff = now.timestamp() - 30*24*3600
        for f in BETA_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
        
        return {"ok": True, "saved": filename}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/beta-export-history")
async def beta_export_history():
    """Seznam shranjenih exportov (zadnjih 30 dni)."""
    try:
        BETA_DIR.mkdir(exist_ok=True)
        files = sorted(BETA_DIR.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
        history = []
        for f in files:
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                history.append({
                    "filename": f.name,
                    "original": data.get("filename",""),
                    "date": data.get("date",""),
                    "timestamp": data.get("timestamp",""),
                    "total_spend": data.get("total_spend",0),
                    "total_purchases": data.get("total_purchases",0),
                    "campaign_count": len(data.get("campaigns",[])),
                })
            except:
                pass
        return {"ok": True, "exports": history}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/beta-export-load/{filename}")
async def beta_export_load(filename: str):
    """Naloži specifičen export."""
    try:
        f = BETA_DIR / filename
        if not f.exists():
            return {"ok": False, "error": "Datoteka ne obstaja"}
        data = json.loads(f.read_text(encoding="utf-8"))
        return {"ok": True, **data}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/forecast-fix-today")
async def forecast_fix_today():
    """Ročno popravi entries za danes — združi vse ključe."""
    from datetime import datetime
    try:
        import pytz
        lj = pytz.timezone("Europe/Ljubljana")
        today = datetime.now(lj).strftime("%Y-%m-%d")
        d_now = datetime.now(lj)
    except:
        d_now = datetime.utcnow()
        today = d_now.strftime("%Y-%m-%d")

    slsi_key = f"{d_now.day}. {d_now.month}. {d_now.year}"

    if not FORECAST_HISTORY_FILE.exists():
        return {"ok": False, "error": "History file ne obstaja"}

    hist = json.loads(FORECAST_HISTORY_FILE.read_text(encoding="utf-8"))

    # Združi vnose iz VSEH možnih ključev za danes
    all_raw = {}
    keys_used = []
    for key in [today, slsi_key]:
        if hist.get(key):
            keys_used.append(key)
            for e in hist[key]:
                time_key = f"{e.get('h',0):02d}:{e.get('m',0):02d}"
                if time_key not in all_raw:
                    all_raw[time_key] = e

    if not all_raw:
        return {"ok": False, "error": "Ni vnosov za danes", "tried": [today, slsi_key], "available": list(hist.keys())[-5:]}

    # Sortiraj po času
    sorted_entries = sorted(all_raw.values(), key=lambda e: e.get('h',0)*60 + e.get('m',0))

    recovered = {
        "date": today,
        "entries": [
            {
                "label": str(e.get("h",0)).zfill(2) + ":" + str(e.get("m",0)).zfill(2),
                "dejanski": e.get("rev", 0),
                "dejanskiOrd": e.get("ord", 0),
                "napoved": None,
                "napovedOrd": None,
            }
            for e in sorted_entries
        ]
    }
    FORECAST_ENTRIES_FILE.write_text(json.dumps(recovered, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "keys_used": keys_used, "entries_recovered": len(sorted_entries), "entries": recovered["entries"]}

@app.get("/forecast-clear-today")
async def forecast_clear_today():
    """Počisti entries za danes (admin endpoint)."""
    try:
        if FORECAST_ENTRIES_FILE.exists():
            data = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
            old_count = len(data.get("entries", []))
            FORECAST_ENTRIES_FILE.write_text(json.dumps({}, ensure_ascii=False, indent=2), encoding="utf-8")
            return {"ok": True, "cleared": old_count}
        return {"ok": True, "cleared": 0}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ─── SPAM FILTER ────────────────────────────────────────────────────────────
SPAM_DIR = DATA_DIR / "spam"
SPAM_DIR.mkdir(exist_ok=True, parents=True)
SPAM_CONFIRMED_FILE = SPAM_DIR / "confirmed.json"
SPAM_REJECTED_FILE = SPAM_DIR / "rejected.json"

def _spam_load_set(path: Path) -> set:
    if not path.exists(): return set()
    try:
        return set(json.loads(path.read_text(encoding="utf-8")))
    except: return set()

def _spam_save_set(path: Path, s: set):
    path.write_text(json.dumps(list(s), ensure_ascii=False, indent=2), encoding="utf-8")

SPAM_THRESHOLD = 80  # Spodnja meja za prikaz - samo >80% se vrne v ads.slx

@app.post("/spam-analyze")
async def spam_analyze():
    """Naloži zadnjih 100 ticketov iz vseh brand-ov, klasificira z AI in vrne SAMO tiste z score >80%."""
    if not KAYAKO_API_KEY or not KAYAKO_SECRET:
        return {"ok": False, "error": "Kayako API ni konfiguriran"}

    confirmed = _spam_load_set(SPAM_CONFIRMED_FILE)
    rejected  = _spam_load_set(SPAM_REJECTED_FILE)

    all_tickets = []
    async with httpx.AsyncClient() as client:
        # 100 odprtih ticketov iz vsakega branda = 200 skupaj
        for brand, dept_id in KAYAKO_DEPT.items():
            # Status filter: 1 = samo odprti (status_id=1)
            path = f"/Tickets/Ticket/ListAll/{dept_id}/1/-1/-1/100/0/ticketid/DESC"
            url = _kayako_build_url(path)
            try:
                r = await client.get(url, timeout=30)
                if r.status_code != 200:
                    print(f"[spam] ListAll {brand} HTTP {r.status_code}")
                    continue
                tickets = _parse_ticket_xml(r.text)
            except Exception as e:
                print(f"[spam] ListAll error {brand}: {e}")
                continue

            for t in tickets:
                tid = t.get("id", "")
                if tid in rejected or tid in confirmed:
                    continue  # že obdelano - preskoči
                t["brand"] = brand
                # Ne nalagamo postov — analiziramo samo naslov
                all_tickets.append(t)

    # AI klasifikacija
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY ni nastavljen"}

    client_h = anthropic.Anthropic(api_key=api_key)
    classified = []
    skipped = 0

    # Hitri pre-filter: označi za AI samo angleške naslove
    # Ne-angleški znaki, šumniki, in tipične besede v drugih jezikih = preskoči
    import unicodedata as _ud

    def _is_english_only(text: str) -> bool:
        """True če je naslov verjetno angleški (brez šumnikov, brez ne-ASCII črk)."""
        if not text or len(text.strip()) < 3:
            return False
        # Vsebuje šumnike ali cirilico ali grško → ni angleški
        for ch in text:
            if ch in 'čšžćđČŠŽĆĐńąęłóśźżŃĄĘŁÓŚŹŻáéíóúýäöüÁÉÍÓÚÝÄÖÜőűŐŰâîăşţÂÎĂŞŢ':
                return False
            # Ne-latinski alfabet (cirilica, grščina, arabščina, kitajščina...)
            if ch.isalpha() and ord(ch) > 591:
                return False
        # Tipične ne-angleške besede (običajno na začetku naslova)
        text_lower = text.lower()
        non_english_keywords = [
            'naroč', 'pošilj', 'dostav', 'plač', 'računa', 'izdelek', 'vraač',
            'reklam', 'kupon', 'paket', 'cena', 'ponudb', 'poizvedb',
            'narudžb', 'isporuk', 'plać', 'račun', 'proizvod', 'povrat',
            'narudžba', 'poručio', 'paypal',
            'поръчк', 'доставк', 'плащане', 'продукт',
            'objedn', 'doruč', 'platb', 'výrobok', 'vrátenie',
            'rendel', 'szállí', 'fizetés', 'termék', 'visszaküld',
            'zamówi', 'dostawa', 'płatność', 'produkt', 'zwrot',
            'comand', 'livrare', 'plat', 'produs', 'retur',
            'παραγγελ', 'αποστολ', 'πληρωμ', 'προϊόν', 'επιστροφ'
        ]
        for kw in non_english_keywords:
            if kw in text_lower:
                return False
        return True

    for t in all_tickets:
        subject = t.get("subject", "")
        from_email = t.get("email", "")
        from_name = t.get("fullname", "")

        # PRE-FILTER: samo angleški naslovi grejo skozi AI
        if not _is_english_only(subject):
            rejected.add(t.get("id",""))
            skipped += 1
            continue

        prompt = f"""Si strog klasifikator email spam-a. Analiziraj SAMO naslov ticket-a iz Kayako support sistema (e-trgovina).

POŠILJATELJ: {from_name} <{from_email}>
NASLOV: {subject}

Odgovori SAMO z JSON:
{{"score": <0-100, koliko verjetno je spam>, "reason": "<razlaga v slovenščini, max 1 stavek>"}}

Visok score (>80) = oglaševanje SEO/marketing/development storitev, B2B "we offer X services", "increase your traffic", phishing, prevare, "partnership opportunity", agencije ki ponujajo svoje storitve, generično angleško sporočilo brez konteksta naših izdelkov.
Srednji (50-80) = sumljivo a možno legitimno (npr. tuji kupec).
Nizek (<50) = pristno vprašanje o naročilih/izdelkih/dostavi (npr. order #12345, my order, refund request, where is my package)."""

        try:
            msg = client_h.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=200,
                messages=[{"role": "user", "content": prompt}]
            )
            txt = msg.content[0].text.strip() if msg.content else "{}"
            txt = re.sub(r'^```(?:json)?|```$', '', txt, flags=re.MULTILINE).strip()
            try:
                ai = json.loads(txt)
                score = int(ai.get("score", 0))
                reason = ai.get("reason", "")
            except:
                continue
        except Exception as e:
            print(f"[spam] AI error for {t.get('id','')}: {e}")
            continue

        # Samo nad threshold-om
        if score < SPAM_THRESHOLD:
            # Auto-zavrni tiste pod threshold-om da jih ne procesiramo več
            rejected.add(t.get("id",""))
            skipped += 1
            continue

        classified.append({
            "id": t.get("id", ""),
            "brand": t.get("brand", ""),
            "subject": subject,
            "from": f"{from_name} <{from_email}>" if from_email else from_name,
            "score": score,
            "reason": reason,
            "url": f"https://support.silux.si/staff/index.php?/Tickets/Ticket/View/{t.get('id','')}",
        })

    # Shrani auto-rejected da jih naslednji klic preskoči
    _spam_save_set(SPAM_REJECTED_FILE, rejected)

    return {"ok": True, "tickets": classified, "confirmed": list(confirmed), "stats": {"checked": len(all_tickets), "found": len(classified), "filtered_out": skipped}}

@app.post("/spam-confirm")
async def spam_confirm(data: dict):
    """Potrdi spam → premakni v Trash v Kayako, ali zavrni."""
    tid = str(data.get("id", ""))
    action = data.get("action", "")
    if not tid:
        return {"error": "missing id"}

    confirmed = _spam_load_set(SPAM_CONFIRMED_FILE)
    rejected  = _spam_load_set(SPAM_REJECTED_FILE)

    if action == "confirm":
        # Premakni v Trash v Kayako
        # Status ID za Trash = običajno 4 (Trash) ali 6 (Spam) - poskusimo Trash
        trash_status = int(os.environ.get("KAYAKO_TRASH_STATUS_ID", "4"))
        try:
            from urllib.parse import quote_plus
            url = _kayako_build_url(f"/Tickets/Ticket/{tid}")
            # PUT request z status ID
            async with httpx.AsyncClient() as client:
                form_data = {
                    "ticketstatusid": str(trash_status),
                }
                r = await client.put(url, data=form_data, timeout=20)
                print(f"[spam] Trash ticket {tid}: HTTP {r.status_code}")
                if r.status_code != 200:
                    print(f"[spam] response: {r.text[:200]}")
                    return {"ok": False, "error": f"Kayako error: {r.status_code}"}
        except Exception as e:
            print(f"[spam] Trash error: {e}")
            return {"ok": False, "error": str(e)}

        confirmed.add(tid)
        rejected.discard(tid)
        _spam_save_set(SPAM_CONFIRMED_FILE, confirmed)
        _spam_save_set(SPAM_REJECTED_FILE, rejected)
        return {"ok": True, "trashed": True}

    elif action == "reject":
        rejected.add(tid)
        confirmed.discard(tid)
    elif action == "unconfirm":
        confirmed.discard(tid)

    _spam_save_set(SPAM_CONFIRMED_FILE, confirmed)
    _spam_save_set(SPAM_REJECTED_FILE, rejected)
    return {"ok": True}

@app.post("/spam-clear-confirmed")
async def spam_clear_confirmed():
    """Počisti seznam potrjenih spam-ov."""
    _spam_save_set(SPAM_CONFIRMED_FILE, set())
    return {"ok": True}
